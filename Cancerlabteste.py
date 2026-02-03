# -*- coding: utf-8 -*-
"""
Created on Fri Jun 27 18:58:02 2025

@authors: victor fleck; Juliano Chaker; Fabio Pittella
"""

import streamlit as st
import json
import html as _html
import html, urllib.parse, mimetypes, base64
import pandas as pd
import numpy as np
import bcrypt 
from sqlalchemy import create_engine, Table, Column, Integer, String, MetaData, select
from datetime import datetime, date, timedelta
import os
import secrets
import smtplib
from email.message import EmailMessage
from sqlalchemy import DateTime
import urllib.parse as _uq
from html import escape
# top of file
from pathlib import Path
from textwrap import dedent
# tools/convert_to_webp.py
from PIL import Image


BASE_DIR = Path(__file__).parent
ICON_PATH = BASE_DIR / "static" / "cancerlab_icon.png"

st.set_page_config(
    page_title="CancerLab – Laboratory of Molecular Pathology of Cancer",
    page_icon=str(ICON_PATH) if ICON_PATH.exists() else "🧬",  # fallback se o PNG não existir
    layout="wide"
)

CANDIDATES = [
    "imghome1.jpg",
    "imghome2.jpg",
    "landmarks.jpg",
    "static/imgfront.png",
    # adicione mais se quiser...
]

GENERATE_SCALES = [1280, 768]   # deixe [] se não quiser tamanhos menores
PHOTO_QUALITY   = 82            # 75–85 é bom; 82 equilibra bem
WEBP_METHOD     = 6             # 0 rápido…6 melhor compactação

def has_alpha(img: Image.Image) -> bool:
    return img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info)

def to_webp(src: Path, dst: Path | None = None, max_width: int | None = None):
    if not src.exists():
        print(f"[skip] não encontrado: {src}")
        return
    if src.suffix.lower() == ".svg":
        print(f"[skip] svg não é suportado pelo Pillow: {src.name}")
        return

    img = Image.open(src)
    alpha = has_alpha(img)

    if max_width and img.width > max_width:
        img.thumbnail((max_width, 10_000), Image.LANCZOS)

    if dst is None:
        dst = src.with_suffix(".webp")

    if dst.exists() and dst.stat().st_mtime >= src.stat().st_mtime:
        print(f"[ok] já atualizado: {dst.name}")
        return

    save_args = {"method": WEBP_METHOD}
    if alpha or src.suffix.lower() == ".png":
        img = img.convert("RGBA")
        save_args["lossless"] = True      # logos/ícones com bordas nítidas
    else:
        img = img.convert("RGB")
        save_args["quality"] = PHOTO_QUALITY

    dst.parent.mkdir(parents=True, exist_ok=True)
    img.save(dst, "WEBP", **save_args)
    print(f"[ok] {src.name} -> {dst.name} ({img.width}x{img.height})")

def main():
    root = Path(__file__).resolve().parent.parent
    for rel in CANDIDATES:
        p = (root / rel).resolve()
        if not p.exists():
            print(f"[warn] não encontrado: {rel}")
            continue
        to_webp(p)
        for w in GENERATE_SCALES:
            to_webp(p, dst=p.with_name(f"{p.stem}-{w}.webp"), max_width=w)

if __name__ == "__main__":
    main()

from pathlib import Path


from pathlib import Path
import mimetypes

def prefer_webp(path_or_rel: str) -> str:
    p = Path(path_or_rel)
    if not p.is_absolute():
        p = (Path(__file__).parent / p).resolve()
    webp = p.with_suffix(".webp")
    return str(webp if webp.exists() else p)

def cached_img_uri(path_or_rel: str, mime: str | None = None) -> str:
    """
    Retorna data:URI para a imagem, preferindo .webp se existir.
    Aceita mime opcional; se não vier, deduz pelo sufixo do arquivo.
    """
    p = Path(prefer_webp(path_or_rel))
    mime = mime or (mimetypes.guess_type(str(p))[0] or "image/webp")
    return data_uri_cached(str(p), mime=mime)



DATABASE_URL = (
    "postgresql+psycopg2://"
    "postgres:nDxPchIFBWnEZercIqUuEDwxTZsLNBcL"
    "@switchback.proxy.rlwy.net:37511/railway"
)

metadata = MetaData()

@st.cache_data(show_spinner=False, max_entries=128)
def _data_uri_cached_core(path_str: str, mime: str | None, mtime_ns: int) -> str:
    p = Path(path_str)
    mime = mime or (mimetypes.guess_type(p.name)[0] or "image/jpeg")
    return f"data:{mime};base64," + base64.b64encode(p.read_bytes()).decode()

def data_uri_cached(path_or_rel: str, mime: str | None = None) -> str:
    p = Path(path_or_rel)
    if not p.is_absolute():
        p = (Path(__file__).parent / p).resolve()
    if not p.exists():
        raise FileNotFoundError(f"Asset not found: {path_or_rel}")
    return _data_uri_cached_core(str(p), mime, p.stat().st_mtime_ns)


@st.cache_resource(show_spinner=False)
def get_engine():
    return create_engine(
        DATABASE_URL,
        pool_pre_ping=True,
        pool_size=5,
        max_overflow=2,
        pool_recycle=1800,
    )

engine = get_engine()







# ====================== Public Page inicio ======================
PRIMARY_BLUE = "#4160BA"   # ajuste de cor 
TEXT_DARK = "#1b1f24"

# slug -> menu central 
PAGES_PUBLIC = [
    ("home", "home"),
    ("research", "research"),
    ("publications", "publications"),
    ("equipment", "equipment"),
    ("people", "people"),  
    ("partners", "partners"),
    ("positions", "open positions"),
    ("news", "news"), # + add this
    ("login", "Internal Access"),
]





def show_help_widget(
    name: str = "Victor Fleck",
    email: str = "Victor.fleck906438@gmail.com",
    role: str = "Undergraduate Pharmacy Student • Web Developer",
    phone: str | None = "+55 61 991281207",
    label: str = "Help",
    accent_css_var: str = "var(--bar-bg, #3f5eb5)",  # inherits your navbar blue
):

    # prefilled email
    subject = _uq.quote("Website issue – support request")
    body = _uq.quote(
        "Hi,\n\nI found a problem on the site.\n"
        "- What I was doing:\n- What happened:\n- Expected behavior:\n\nThanks!"
    )
    mailto = f"mailto:{_uq.quote(email.strip())}?subject={subject}&body={body}"

    # optional phone links
    tel_link = f"tel:{_uq.quote(phone)}" if phone else ""
    wa_link  = f"https://wa.me/{''.join([c for c in (phone or '') if c.isdigit()])}" if phone else ""

    st.markdown(
        f"""
        <style>
          :root {{
            --help-accent: {accent_css_var};
            --help-ink: #0f172a;
            --help-muted: #5a6881;
            --help-line: #e9edf6;
            --help-glass: rgba(255,255,255,.65);
            --help-glass-brd: rgba(255,255,255,.55);
          }}

          .help-root {{
            position: fixed; right: 14px; bottom: 14px; z-index: 1100;
            font-family: ui-sans-serif, system-ui, -apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial;
          }}
          .help-toggle {{ display:none; }}

          /* --- MINI FLOATING BUTTON --- */
          .help-fab {{
            position: fixed; right: 14px; bottom: 14px;
            width: 46px; height: 46px; border-radius: 999px;
            display:inline-flex; align-items:center; justify-content:center; gap:10px;
            background:
              radial-gradient(120% 120% at 50% 0%, rgba(255,255,255,.16), rgba(255,255,255,0)),
              var(--help-accent);
            color:#fff; border:1px solid rgba(255,255,255,.45);
            box-shadow: 0 12px 26px rgba(2,6,23,.18);
            cursor:pointer; user-select:none; text-decoration:none;
            transition: transform .15s ease, box-shadow .15s ease, background .2s ease;
            backdrop-filter: saturate(140%) blur(2px);
            -webkit-backdrop-filter: saturate(140%) blur(2px);
          }}
          .help-fab:hover {{ transform: translateY(-1px); box-shadow: 0 16px 34px rgba(2,6,23,.22); }}
          .help-fab svg {{ width:20px; height:20px; }}

          /* --- GLASS POPOVER (COMPACT) --- */
          .help-backdrop {{
            position: fixed; inset: 0; background: rgba(2,6,23,.08);
            opacity:0; pointer-events:none; transition: opacity .15s ease;
          }}
          .help-card {{
            position: fixed; right: 14px; bottom: 66px; width: min(92vw, 320px);
            background: var(--help-glass);
            color: var(--help-ink);
            border-radius: 16px;
            border:1px solid var(--help-glass-brd);
            box-shadow: 0 22px 44px rgba(2,6,23,.22);
            transform: translateY(8px) scale(.99);
            opacity:0; pointer-events:none;
            transition: transform .18s ease, opacity .18s ease, box-shadow .18s ease;
            backdrop-filter: blur(12px) saturate(120%);
            -webkit-backdrop-filter: blur(12px) saturate(120%);
            overflow:hidden;
          }}
          .help-header {{
            display:flex; align-items:center; justify-content:space-between; gap:10px;
            padding: 10px 12px;
            background:
              linear-gradient(180deg, rgba(255,255,255,.45), rgba(255,255,255,.15)),
              var(--help-accent);
            color:#fff; border-bottom:1px solid rgba(255,255,255,.45);
          }}
          .help-title {{ margin:0; font-weight:900; font-size:.98rem; letter-spacing:.2px; }}
          .help-close {{
            width:28px; height:28px; border-radius:10px; border:1px solid rgba(255,255,255,.6);
            background: rgba(255,255,255,.15); color:#fff;
            display:inline-flex; align-items:center; justify-content:center; cursor:pointer; text-decoration:none;
          }}
          .help-body {{ padding: 12px; }}
          .help-meta {{ font-size:.92rem; color:var(--help-muted); margin:6px 0 10px; }}
          .help-strong {{ font-weight:900; color:var(--help-ink); }}

          .row {{ display:flex; flex-wrap:wrap; gap:8px; }}
          .btn, .btn-ghost {{
            display:inline-flex; align-items:center; gap:8px; text-decoration:none; font-weight:800;
            padding:10px 12px; border-radius:12px; font-size:.9rem;
            transition: transform .12s ease, box-shadow .12s ease, border-color .12s ease, background .12s ease;
          }}
          .btn {{ background: var(--help-accent); color:#fff; border:1px solid rgba(255,255,255,.55);
                 box-shadow: 0 10px 24px rgba(2,6,23,.18); }}
          .btn:hover {{ transform: translateY(-1px); box-shadow: 0 14px 28px rgba(2,6,23,.22); }}
          .btn-ghost {{ background: rgba(255,255,255,.75); color: var(--help-ink); border:1px solid var(--help-line); }}
          .btn-ghost:hover {{ transform: translateY(-1px); border-color:#cbd5e1; box-shadow:0 8px 18px rgba(2,6,23,.10); }}
          /* Toggle ON */
          #help-pop:checked ~ .help-backdrop {{ opacity:1; pointer-events:auto; }}
          #help-pop:checked ~ .help-card {{ opacity:1; transform: translateY(0) scale(1); pointer-events:auto; }}
          /* Accessibility */
          .help-card a:focus {{ outline:none; box-shadow:0 0 0 3px rgba(59,130,246,.35); border-radius:12px; }}
          @media (prefers-reduced-motion: reduce) {{ .help-fab, .help-card {{ transition: none; }} }}
        </style>
        <div class="help-root" aria-live="polite">
          <input id="help-pop" class="help-toggle" type="checkbox" />
          <!-- Floating mini button -->
          <label class="help-fab" for="help-pop" title="Get help">
            <!-- Lifebuoy icon -->
            <svg viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
              <circle cx="12" cy="12" r="9"></circle>
              <circle cx="12" cy="12" r="3.5"></circle>
              <path d="M4.9 4.9l3.2 3.2M19.1 4.9l-3.2 3.2M4.9 19.1l3.2-3.2M19.1 19.1l-3.2-3.2"></path>
            </svg>
          </label>
          <!-- Backdrop for closing -->
          <label class="help-backdrop" for="help-pop" aria-hidden="true"></label>
          <!-- Glass popover -->
          <aside class="help-card" role="dialog" aria-modal="true" aria-labelledby="help-title">
            <div class="help-header">
              <h4 id="help-title" class="help-title">Problems with the Website?</h4>
              <label for="help-pop" class="help-close" title="Close">✕</label>
            </div>
            <div class="help-body">
              <div class="help-meta">
                <span class="help-strong">{_html.escape(name)}</span> — {_html.escape(role)}
              </div>
              <div class="row" style="margin-bottom:10px;">
                <a class="btn" href="{mailto}" target="_blank" rel="noopener" aria-label="Email">
                  <!-- mail icon -->
                  <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
                    <rect x="3" y="5" width="18" height="14" rx="2" ry="2"></rect>
                    <path d="M3 7l9 6 9-6"></path>
                  </svg>
                  Email
                </a>
                {"<a class='btn-ghost' href='"+_html.escape(tel_link)+"' aria-label='Call'><svg width='16' height='16' viewBox='0 0 24 24' fill='none' stroke='%230f172a' stroke-width='2' stroke-linecap='round' stroke-linejoin='round' aria-hidden='true'><path d='M22 16.92v3a2 2 0 0 1-2.18 2 19.79 19.79 0 0 1-8.63-3.07 19.5 19.5 0 0 1-6-6A19.79 19.79 0 0 1 1.09 4.18 2 2 0 0 1 3.1 2h3a2 2 0 0 1 2 1.72c.12.9.34 1.78.66 2.61a2 2 0 0 1-.45 2.11L7.09 9.91a16 16 0 0 0 6 6l1.47-1.22a2 2 0 0 1 2.11-.45c.83.32 1.71.54 2.61.66A2 2 0 0 1 22 16.92z'/></svg>Call</a>" if phone else ""}
                {"<a class='btn-ghost' href='"+_html.escape(wa_link)+"' target='_blank' rel='noopener' aria-label='WhatsApp'><svg width='16' height='16' viewBox='0 0 24 24' fill='none' stroke='%230f172a' stroke-width='2' stroke-linecap='round' stroke-linejoin='round' aria-hidden='true'><path d='M20.52 3.48A11.5 11.5 0 0 0 2.6 18.39L2 22l3.72-.58A11.5 11.5 0 1 0 20.52 3.48Z'/><path d='M7 10.5s1 3 5 5 3-2 3-2'/></svg>WhatsApp</a>" if phone else ""}
              </div>
              <a class="btn-ghost" href="{mailto}" target="_blank" rel="noopener">Report a problem</a>
            </div>
          </aside>
        </div>
        """,
        unsafe_allow_html=True,
    )







def show_lab_header( 
    image_relpath: str = "static/imgfront",
    address_lines: tuple[str, ...] = (
        "Laboratory of Molecular Pathology of Cancer",
        "Faculty of Health Sciences",
        "University of Brasília",
        "Brasília – DF – Brazil",
    ),
    logo_height_px: int = 180,
):
    from pathlib import Path
    p = Path(image_relpath)
    if not p.suffix:
        for ext in (".svg", ".png", ".jpg", ".jpeg", ".webp", ".gif"):
            cand = p.with_suffix(ext)
            if cand.exists():
                p = cand
                break

    img_tag = ""
    if p.exists():
        img_uri = data_uri_cached(str(p))
        img_tag = (
            f'<a href="?page=home" class="logo-link" aria-label="Home">'
            f'  <img class="logo" src="{img_uri}" alt="Lab logo" loading="eager" decoding="async">'
            f'</a>'
        )

    mobile_logo_h = max(72, logo_height_px - 48)

    import streamlit as st
    st.markdown(
        f"""
        <style>
          /* ===== Header card com fundo branco ===== */
          .brand-header {{
            background: #ffffff;                      /* <- branco puro */
            border: 1px solid rgba(2,6,23,.06);
            border-radius: 16px;
            box-shadow: 0 10px 26px rgba(2,6,23,.06);
            margin: 0 auto 12px;
          }}
          .brand-wrap {{
            max-width:1180px;
            margin:0 auto;
            padding: 24px 22px;
            display:grid;
            grid-template-columns:auto 1fr;          /* logo | texto */
            align-items:center;
            gap:34px;
          }}

          /* remove qualquer brilho atrás do logo */
          .brand-left {{ position:relative; }}
          .brand-left::before {{ display:none; }}

          /* Logo grande com escala responsiva */
          .brand-header .logo {{
            height: clamp({mobile_logo_h}px, 18vw, {logo_height_px}px);
            max-height: min(32vh, {logo_height_px}px);
            width:auto; display:block;
            filter: drop-shadow(0 1px 0 rgba(0,0,0,.04));
          }}
          .brand-header .logo-link {{ text-decoration:none; line-height:0; }}

          /* Tipografia do endereço */
          .brand-header .addr {{
            text-align:right;
            color: var(--TEXT_DARK, #0f172a);
            letter-spacing:.15px;
          }}
          .brand-header .addr .lab {{
            display:block;
            font-weight:800;
            font-size:clamp(24px, 3.0vw, 34px);
            line-height:1.12;
          }}
          .brand-header .addr .lab::after {{
            content:"";
            display:block;
            width:72px; height:3px; border-radius:2px;
            background: rgba(19,81,216,.45);
            margin:12px 0 10px auto;
          }}
          .brand-header .addr .muted {{
            opacity:.9;
            font-weight:700;
            font-size:clamp(14px, 1.15vw, 17px);
            line-height:1.28;
          }}

          /* ===== Mobile ===== */
          @media (max-width:900px){{
            .brand-wrap {{ grid-template-columns:1fr; gap:14px; padding:18px 14px; }}
            .brand-header .addr {{ text-align:center; }}
            .brand-header .logo {{ margin:0 auto; }}
            .brand-header .addr .lab::after {{ margin-left:auto; margin-right:auto; }}
          }}
        </style>
        <div class="brand-header" role="banner" aria-label="Laboratory heading">
          <div class="brand-wrap">
            <div class="brand-left">{img_tag}</div>
            <div class="addr">
              <span class="lab">{address_lines[0]}</span>
              <span class="muted">{'<br>'.join(address_lines[1:])}</span>
            </div>
          </div>
        </div>
        """,
        unsafe_allow_html=True
    )


def _get_param(key: str, default=None):
    try:
        params = st.query_params     # Streamlit >= 1.32
        val = params.get(key, default)
        if isinstance(val, (list, tuple)):
            return val[0]
        return val
    except Exception:
        params = st.experimental_get_query_params()  
        return params.get(key, [default])[0] if params.get(key) else default

def _set_param(**kwargs):
    try:
        st.query_params.update(kwargs)
    except Exception:
        st.experimental_set_query_params(**kwargs)
        
def show_front_strip(
    caption: str = "Welcome to the lab — advancing genomics & ctDNA",
    image_relpath: str = "static/imgfront.webp",   # auto-detect ext
):
    p = Path(image_relpath)
    if not p.suffix:
        for ext in (".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"):
            cand = p.with_suffix(ext)
            if cand.exists():
                p = cand
                break

    img_tag = ""
    if p.exists():
        img_uri = cached_img_uri(str(p))
        img_tag = f'<img class="front-img" src="{img_uri}" alt="front image" loading="eager" decoding="async">'

    st.markdown(
        f"""
        <style>
          .front-strip {{
            position: sticky; top: 0; left: 0; right: 0;
            z-index: 1000;
            background: var(--front-bg, #f3f9fd);
            border-bottom: 1px solid rgba(15,23,42,.06);
          }}
          .front-wrap {{
            max-width: 1180px; margin: 0 auto; padding: 8px 14px;
            display: grid; grid-template-columns: 1fr auto; align-items: center; gap: 10px;
            font: 600 14px/1.25 ui-sans-serif, system-ui, -apple-system, "Segoe UI", Roboto, Arial;
            color: var(--front-ink, #0f172a);
          }}
          .front-caption {{ letter-spacing: .2px; }}
          .front-img {{ display:block; height:30px; width:auto; }}
          /* diminuir o empurrão geral do app */
          .stApp {{ padding-top: 8px; }}
          @media (max-width:700px) {{
            .front-img {{ height:26px; }}
            .stApp {{ padding-top: 6px; }}
          }}
        </style>
        <div class="front-strip">
          <div class="front-wrap">
            <div class="front-caption">{caption}</div>
            {img_tag}
          </div>
        </div>
        """,
        unsafe_allow_html=True
    )



                       
def render_navbar(active: str = "home", unbimg: str | None = None):
    items = []
    for slug, label in PAGES_PUBLIC:
        is_active = (active == slug)
        cls = "active" if is_active else ""
        aria = ' aria-current="page"' if is_active else ""
        items.append(f'<a target="_self" class="{cls}" href="?page={slug}"{aria}>{label}</a>')
    menu_html = "".join(items)

    if unbimg is None:
        try:
            for pth, mime in [("static/unb.svg","image/svg+xml"),
                              ("static/unb.png","image/png"),
                              ("static/unb.jpg","image/jpeg")]:
                if Path(pth).exists():
                    unbimg = data_uri_cached(pth, mime=mime); break
        except Exception:
            unbimg = ""

    unb_html = (
        f'<a class="icon-img" href="https://www.unb.br/" target="_self" '
        f'rel="noopener" title="Universidade de Brasília">'
        f'  <img src="{unbimg or ""}" alt="UnB" />'
        f'</a>'
    )

    instagram_svg = (
        '<svg class="glyph" viewBox="0 0 24 24" aria-hidden="true">'
        '<path fill="currentColor" d="M7 2h10a5 5 0 015 5v10a5 5 0 01-5 5H7a5 5 0 01-5-5V7a5 5 0 015-5zm0 2a3 3 0 00-3 3v10a3 3 0 003 3h10a3 3 0 003-3V7a3 3 0 00-3-3H7zm5 3a5 5 0 110 10 5 5 0 010-10zm0 2a3 3 0 100 6 3 3 0 000-6zm6.5-2a1.5 1.5 0 100 3 1.5 1.5 0 000-3z"/>'
        '</svg>'
    )
    ig_html = (
        f'<a class="icon-btn ig" href="https://www.instagram.com/cancerlab_unb/" '
        f'target="_blank" rel="noopener" title="@cancerlab_unb">'
        f'{instagram_svg}<span class="label">Instagram</span></a>'
    )
    icons_html = unb_html + ig_html

    st.markdown(
        f"""
<style>
  :root {{
    --bar-bg: {PRIMARY_BLUE};
    --bar-shadow: 0 6px 18px rgba(2,6,23,.08);
    --ig-grad: linear-gradient(90deg,#F58529 0%,#FEDA77 22%,#DD2A7B 45%,#8134AF 70%,#515BD4 100%);
  }}

  header[data-testid="stHeader"] {{ height:0 !important; }}
  .main .block-container {{ padding-top:0 !important; }}
  div[data-testid="stDecoration"] {{ display:none; }}

  .topbar {{
    position:sticky; top:0; z-index:1000;
    background: linear-gradient(180deg, rgba(0,0,0,.05) 0%, rgba(0,0,0,0) 100%), var(--bar-bg);
    border-bottom:1px solid rgba(0,0,0,.08);
    backdrop-filter:saturate(140%) blur(8px);
    -webkit-backdrop-filter:saturate(140%) blur(8px);
    box-shadow:var(--bar-shadow);
    padding-top:max(env(safe-area-inset-top),0px);
  }}
  /* allow the left column to actually shrink without clipping */
  .topbar-inner {{
    max-width:1180px; margin:0 auto; padding:8px 14px;
    display:grid; grid-template-columns: minmax(0,1fr) auto;  /* <-- key fix */
    align-items:center; gap:12px;
  }}

  /* ===== Menu ===== */
  .menu {{
    display:flex; gap:16px; align-items:center;
    justify-content:flex-start;               /* <-- anchor to the left (no more cut “home”) */
    min-width:0;
    overflow:auto hidden;                     /* horizontal scroll only if really needed */
  }}
  .menu::-webkit-scrollbar {{ display:none; }}
  .menu a {{
    position:relative; text-decoration:none; color:#fff; text-transform:lowercase;
    padding:8px 12px; border-radius:10px; font-weight:700; letter-spacing:.2px;
    transition:background .2s, color .2s, transform .15s; white-space:nowrap; line-height:1;
    flex:0 0 auto;                            /* do not shrink individual buttons */
    font-size: clamp(12px, 1.05vw, 15px);     /* gentle shrink on tight widths */
  }}
  .menu a:hover {{ background:rgba(255,255,255,.14); transform:translateY(-1px); }}
  .menu a.active, .menu a[aria-current="page"] {{ background:rgba(255,255,255,.22); }}
  .menu a.active::after, .menu a[aria-current="page"]::after {{
    content:""; position:absolute; left:12px; right:12px; bottom:6px; height:2px;
    background:rgba(255,255,255,.9); border-radius:2px;
  }}

  /* when things get tight, allow a second line instead of clipping */
  @media (max-width: 1120px) {{
    .menu {{ flex-wrap: wrap; row-gap: 8px; }}
  }}

  /* ===== Right-side icons ===== */
  .icons {{
    display:flex; gap:6px; align-items:center; justify-content:flex-end;
    min-width:0; flex-wrap:nowrap;
  }}
  .icon-btn, .icon-img {{
    display:inline-flex; align-items:center; justify-content:center;
    background:rgba(255,255,255,.08);
    border:1px solid rgba(255,255,255,.45);
    color:#fff; border-radius:10px; padding:5px 8px;
    line-height:1; white-space:nowrap; height:32px;
    transition:background .2s, transform .15s, border-color .2s;
  }}
  .icon-btn:hover, .icon-img:hover {{ background:rgba(255,255,255,.16); transform:translateY(-1px); }}
  .icon-img img {{ height:18px; width:auto; display:block; filter:drop-shadow(0 1px 0 rgba(0,0,0,.05)); }}

  .icon-btn.ig {{ gap:6px; font-weight:800; }}
  .icon-btn.ig .glyph {{ width:18px; height:18px; flex:0 0 18px; color:#fff; }}
  .icon-btn.ig .label {{
    background:var(--ig-grad);
    -webkit-background-clip:text; background-clip:text; -webkit-text-fill-color:transparent;
  }}

  /* Mobile stays compact and proven-good */
  @media (max-width:760px) {{
    .topbar-inner {{ grid-template-columns:1fr; gap:8px; }}
    .icons {{ justify-content:flex-start; }}
    .icon-btn .label {{ display:none; }}
  }}

  @media (prefers-reduced-motion: reduce) {{
    .menu a, .icon-btn, .icon-img {{ transition:none; }}
  }}
</style>

<div class="topbar" role="navigation" aria-label="Primary">
  <div class="topbar-inner">
    <nav class="menu">{menu_html}</nav>
    <div class="icons">{icons_html}</div>
  </div>
</div>
""",
        unsafe_allow_html=True,
    )















def page_home():
    left_img = cached_img_uri("imghome1.jpg", mime="image/jpeg")
    right_img = cached_img_uri("imghome2.jpg", mime="image/jpeg")
    landmarks_img = cached_img_uri("landmarks.jpg", mime="image/jpeg")

    lat = -15.767937821610321
    lon = -47.86710221994476
    address_html = (
        "Campus Universitário Darcy Ribeiro<br>"
        "Brasília - DF - s/n<br>"
        "CEP 70910-000"
    )
    gmaps_view = f"https://www.google.com/maps?q={lat},{lon}"
    gmaps_dir  = f"https://www.google.com/maps?daddr={lat},{lon}"
    osm_view   = f"https://www.openstreetmap.org/?mlat={lat}&mlon={lon}#map=17/{lat}/{lon}"
    bbox_w = lon - 0.01; bbox_s = lat - 0.01; bbox_e = lon + 0.01; bbox_n = lat + 0.01
    osm_embed = (
        f"https://www.openstreetmap.org/export/embed.html?"
        f"bbox={bbox_w},{bbox_s},{bbox_e},{bbox_n}&layer=mapnik&marker={lat},{lon}"
    )
    st.markdown(
        f"""
        <style>
      /* --- Hero layout ------------------------------------------------ */
      .hero {{
        display:grid;
        gap: 26px;
        align-items:start;
        grid-template-columns: 1fr;
        grid-template-areas:
          "head"
          "left"
          "center"
          "right";
      }}
      .hero .col-head   {{ grid-area: head; }}
      .hero .col-left   {{ grid-area: left; }}
      .hero .col-center {{ grid-area: center; }}
      .hero .col-right  {{ grid-area: right; }}
      @media (min-width:700px) {{
        .hero {{
          grid-template-columns: 1fr 1fr;
          grid-template-areas:
            "head   head"
            "center center"
            "left   right";
          gap: 28px;
        }}
      }}
      @media (min-width:1100px) {{
        .hero {{
          grid-template-columns: 0.9fr 1.2fr 0.9fr;
          grid-template-areas:
            "head   head   head"
            "left   center right";
          gap: 32px;
        }}
        .col-left  {{ justify-self:start; }}
        .col-right {{ justify-self:end;   }}
        .col-left .imgframe, .col-right .imgframe {{ width:88%; }}
      }}
      .title {{
        font-size: clamp(32px, 5vw, 56px);
        font-weight: 800;
        color: var(--TEXT_DARK, #0f172a);
        margin: 6px 0 12px;
        line-height: 1.08;
        letter-spacing: -0.01em;
      }}
      .col-head {{ text-align:center; display:flex; flex-direction:column; align-items:center; }}
      .col-head .title {{ text-align:center; position:relative; }}
      .col-head .title::after {{
        content:"";
        display:block;
        height:3px;
        width:84px;
        margin:10px auto 0;
        border-radius:3px;
        background:linear-gradient(90deg, var(--primary, #1351d8), rgba(19,81,216,.25));
      }}
      .lead {{ color:#374151; line-height:1.75; font-size:17px; max-width:72ch; }}
      .hero .title + .lead {{
        color: var(--primary, #1351d8);
        font-weight: 700;
        font-size: clamp(18px, 2.1vw, 20px);
        line-height: 1.55;
        margin-top: 14px;
        text-align:center;
      }}
      .wrap a {{
        color: var(--primary, #1351d8);
        text-decoration: none;
        border-bottom: 1px dashed rgba(19,81,216,.35);
      }}
      .wrap a:hover {{ text-decoration: underline; }}
      .imgframe {{
        border:1px solid #e7e7ef;
        border-radius:16px;
        overflow:hidden;
        background:#fff;
        box-shadow: 0 12px 34px rgba(2,6,23,.10);
        transition: transform .25s ease, box-shadow .25s ease;
      }}
      .imgframe img {{ width:100%; height:auto; display:block; }}
      .imgframe:hover {{ transform: translateY(-2px); box-shadow:0 16px 42px rgba(2,6,23,.14); }}
      @media (prefers-reduced-motion: reduce) {{ .imgframe, .stack-card {{ transition:none; transform:none; }} }}
      .col-center .stack-cards {{ max-width:72ch; margin:10px auto 0; display:grid; grid-template-columns:1fr; gap:18px; }}
      .stack-card {{
        position:relative;
        padding:20px 24px;
        border-radius:16px;
        background:
          linear-gradient(#ffffff, #ffffff) padding-box,
          linear-gradient(180deg, rgba(19,81,216,.20), rgba(19,81,216,0)) border-box;
        border:1px solid #e6eaf4;
        box-shadow: 0 12px 28px rgba(2,6,23,.08);
        transition: box-shadow .25s ease, transform .25s ease;
      }}
      .stack-card:hover {{ transform: translateY(-2px); box-shadow:0 16px 38px rgba(2,6,23,.12); }}
      .stack-card h3 {{
        margin:0 0 8px;
        font-weight: 750;
        font-size: clamp(17px, 1.9vw, 20px);
        line-height:1.28;
        letter-spacing:-.005em;
        color: var(--TEXT_DARK, #0f172a);
      }}
      .stack-card h3::after {{
        content:"";
        display:block;
        height:2px;
        width:54px;
        margin-top:8px;
        border-radius:2px;
        background:linear-gradient(90deg, var(--primary, #1351d8), rgba(19,81,216,.20));
      }}
      .stack-card p {{ margin:10px 0 0; font-size:16.5px; line-height:1.75; color: var(--TEXT_DARK, #0f172a); }}
      .stack-card ul {{ margin:10px 0 0; padding-left: 22px; }}
      .stack-card li {{ margin:4px 0; line-height:1.7; color: var(--TEXT_DARK, #0f172a); }}
      .landmarks {{ margin: 42px auto 6px; }}
      .landmarks .lm-title {{
        text-align:center;
        font-weight:800;
        color: var(--TEXT_DARK, #0f172a);
        font-size: clamp(20px, 3vw, 28px);
        margin: 2px 0 14px;
      }}
      .landmarks .imgframe {{ border-radius:18px; }}
      .landmarks .caption {{
        text-align:center;
        color:#475569;
        font-size: 13.5px;
        margin-top:8px;
      }}
      /* === Map section ================================================= */
      .map-section {{ margin: 26px auto 8px; }}
      .map-title {{
        text-align:center; font-weight:800; color:#0b1220;
        font-size: clamp(20px, 3vw, 28px); margin: 6px 0 14px;
      }}
      .map-grid {{
        display:grid; grid-template-columns:1fr; gap:16px; align-items:stretch;
        max-width: 1180px; margin: 0 auto;
      }}
      @media (min-width: 980px){{
        .map-grid {{ grid-template-columns: 1.2fr .8fr; gap:18px; }}
      }}
      .map-card {{
        position:relative; border-radius:18px; overflow:hidden;
        border:1px solid #e6e8ee; background:#fff;
        box-shadow:0 16px 36px rgba(2,6,23,.12);
      }}
      .map-frame {{ position:relative; width:100%; aspect-ratio:16/9; background:#e5eefb; }}
      .map-frame iframe {{ position:absolute; inset:0; width:100%; height:100%; border:0; }}
      .map-frame::after {{
        content:""; position:absolute; left:50%; top:50%; transform:translate(-50%,-50%);
        width:64px; height:64px; border-radius:50%;
        box-shadow:0 0 0 2px rgba(19,81,216,.25), 0 0 0 12px rgba(19,81,216,.06);
        pointer-events:none;
      }}
      .map-badge {{
        position:absolute; left:12px; top:12px;
        background: rgba(255,255,255,.92);
        border:1px solid #e2e8f0; backdrop-filter: blur(6px);
        padding:4px 10px; border-radius:999px; font-weight:800; font-size:.8rem; color:#0b1220;
      }}
      .info-card {{
        border:1px solid #e6e8ee; border-radius:18px; background:#ffffff;
        box-shadow:0 12px 28px rgba(2,6,23,.08); padding:16px 18px;
      }}
      .info-card h3 {{ margin:0 0 8px; font-size:18px; font-weight:800; color:#0b1220; }}
      .addr {{ color:#111827; line-height:1.6; }}

      .btns {{ display:flex; flex-wrap:wrap; gap:10px; margin-top:12px; }}
      .btn {{
        display:inline-block; padding:10px 14px; border-radius:10px; text-decoration:none; font-weight:900;
        background:var(--primary,#1351d8); color:#ffffff !important; line-height:1;
        text-shadow: 0 1px 0 rgba(0,0,0,.25);
        border: none;
      }}
      .btn.alt {{
        background:#ffffff; color:#0b1220 !important; border:1px solid #dbe2f1;
        text-shadow:none;
      }}
      .wrap a.btn, .wrap a.btn.alt {{ border-bottom: none !important; }}
        </style>
        <div class="wrap">
          <div class="hero">
            <div class="col-head">
              <div class="title">CancerLab – Laboratory of Molecular Pathology of Cancer</div>
              <p class="lead">The Laboratory of Molecular Pathology of Cancer (CancerLab) at the University of Brasília is dedicated to advancing cancer research through the development of innovative liquid biopsy tools and the investigation of molecular targets both genetic and epigenetic that drive carcinogenesis. Our ultimate goal is to translate these discoveries into novel diagnostic and prognostic strategies that can improve cancer treatment and patient care</p>
            </div>
            <!-- Left image -->
            <div class="col-left">
              <div class="imgframe">
                <img src="{left_img}" alt="CancerLab – left illustrative image" />
              </div>
            </div>
            <div class="col-center">
              <div class="stack-cards">
                <article class="stack-card">
                  <h3>Our Mission</h3>
                  <p>We are committed to training highly qualified professionals in innovative techniques of gene editing, genomics, and liquid biopsy. Our work is dedicated to advancing prevention and comprehensive care for oncology patients, while also developing cutting edge products and services that strengthen scientific and technological innovation in health.</p>
                </article>
                <article class="stack-card">
                  <h3>Our Research</h3>
                  <p>Our research focuses on:</p>
                  <ul>
                    <li><strong>Liquid biopsy:</strong> studying circulating tumor DNA (ctDNA) for clinical applications.</li>
                    <li><strong>Epigenetics:</strong> investigating the role of methyltransferase enzymes in solid tumors and leukemias.</li>
                    <li><strong>Gene editing:</strong> developing CRISPR/Cas9-based strategies for selective knockout (KO) in cancer.</li>
                  </ul>
                </article>
                <article class="stack-card">
                  <h3>Our Team</h3>
                  <p>The Laboratory of Molecular Pathology of Cancer (CancerLab) is the operational unit of two research groups at the University of Brasília: <a href='http://dgp.cnpq.br/dgp/espelhogrupo/769498'>Patologia Molecular do Câncer</a> (created in 2008) and <a href='http://dgp.cnpq.br/dgp/espelhogrupo/775959'>Desenvolvimento e Aplicações em Biópsia Líquida</a> (created in 2022). Together, these initiatives bring together researchers dedicated to exploring the molecular mechanisms of human carcinogenesis and translating discoveries into clinical applications.</p>
                  <p>Our team integrates expertise across Clinical Pathology, Molecular Biology, Biochemistry, Immunology, and Genetics creating a truly multidisciplinary environment to drive high-impact research in cancer biology.</p>
                </article>
                <article class="stack-card">
                  <h3>Looking Ahead</h3>
                  <p>We aim to expand our collaborations with leading research centers across Brazil and abroad, fostering student and researcher exchanges that will strengthen our scientific network and amplify the global reach of our discoveries.</p>
                </article>
              </div>
            </div>
            <div class="col-right">
              <div class="imgframe">
                <img src="{right_img}" alt="CancerLab – right illustrative image" />
              </div>
            </div>
          </div>
          <section class="landmarks" aria-label="Landmarks timeline">
            <div class="lm-title">Landmarks</div>
            <div class="imgframe">
              <img src="{landmarks_img}" alt="CancerLab timeline of key milestones" />
            </div>
          </section>
          <!-- Map / Location -->
          <section class="map-section" aria-label="Location map">
            <div class="map-title">Where to find us</div>
            <div class="map-grid">
              <div class="map-card">
                <div class="map-frame">
                  <iframe
                    src="{osm_embed}"
                    loading="lazy"
                    referrerpolicy="no-referrer-when-downgrade"
                    title="CancerLab location map"
                  ></iframe>
                </div>
                <div class="map-badge">OpenStreetMap • Marker at {lat:.6f}, {lon:.6f}</div>
              </div>
              <div class="info-card">
                <h3>Laboratory location</h3>
                <div class="addr">{address_html}</div>
                <!-- removed Lat/Lon chips -->
                <div class="btns">
                  <a class="btn" href="{gmaps_view}" target="_blank" rel="noopener">View on Google Maps</a>
                  <a class="btn alt" href="{osm_view}" target="_blank" rel="noopener">OpenStreetMap</a>
                  <a class="btn" href="{gmaps_dir}" target="_blank" rel="noopener">Directions</a>
                </div>
              </div>
            </div>
          </section>
        </div>
        """,
        unsafe_allow_html=True,
    )
    show_help_widget(
    name="Victor Fleck",
    email="Victor.fleck906438@gmail.com",
    role="Undergraduate Pharmacy Student • Web Developer :) ",
    phone="+55 61 991281207",
    label="Problems with the website?"
    )
           
                       
                   



def page_research():
    st.markdown(
        """
        <style>
        /* ===========================================================
           NAVBAR HOTFIX (iOS/Safari + mobile) — sem alterar o HTML
           =========================================================== */

        /* evita zoom/auto-resize do texto e glitches de pintura */
        html { -webkit-text-size-adjust: 100%; }

        /* alvo genérico do seu header/nav; cobre nomes comuns */
        :where(header, .bar, .navbar, .site-nav, [data-nav]){
          position: sticky; top: 0; z-index: 999;
          overflow: clip;                 /* impede “vazar” conteúdo */
          will-change: transform;         /* estabiliza pintura no iOS */
        }
        /* Safari às vezes distorce texto com sticky/transform */
        @supports (-webkit-touch-callout: none){
          :where(header, .bar, .navbar, .site-nav, [data-nav]){
            -webkit-backface-visibility: hidden;
            -webkit-transform: translateZ(0);
          }
        }
        /* no mobile, a lista de links vira carrossel horizontal */
        @media (max-width: 768px){
          :where(header, .bar, .navbar, .site-nav, [data-nav]) :where(nav, .nav, .links, ul){
            display: flex;
            flex-wrap: nowrap;
            justify-content: flex-start;
            gap: 10px;
            overflow-x: auto;
            -webkit-overflow-scrolling: touch;
            scrollbar-width: none;
            padding: 8px 4px;
            white-space: nowrap;
            contain: content;  /* isola a pintura no iOS */
          }
          :where(header, .bar, .navbar, .site-nav, [data-nav]) :where(nav, .nav, .links, ul)::-webkit-scrollbar{ display:none; }

          /* cada item não encolhe nem se sobrepõe */
          :where(header, .bar, .navbar, .site-nav, [data-nav]) :where(li, a){
            flex: 0 0 auto;
            min-width: max-content;
            line-height: 1.2;
            white-space: nowrap;
          }
          /* aumenta alvo de toque, sem mexer no seu esquema de cores */
          :where(header, .bar, .navbar, .site-nav, [data-nav]) a{
            display: inline-flex; align-items: center;
            padding: 8px 12px;
            border-radius: 12px;
            text-decoration: none;
          }
        }
        /* ===========================================================
           MOBILE-SAFE RESET (não altera o desktop)
           =========================================================== */
        *, *::before, *::after { box-sizing: border-box; min-width:0; }
        html, body { width:100%; overflow-x:hidden; }
        /* previne estouro por palavras/comprimentos longos */
        .wrap, .section, .pill, .r-sub, .toc { overflow-wrap:anywhere; word-break:break-word; }
        /* imagens/SVG sempre fluídos */
        .wrap img, .wrap svg { max-width:100%; height:auto; display:block; }
        /* respeita notch/safe-area em iOS */
        :root {
          --safe-left: env(safe-area-inset-left, 0px);
          --safe-right: env(safe-area-inset-right, 0px);
          --safe-top: env(safe-area-inset-top, 0px);
        }


        /* container */
        .wrap { max-width:1180px; margin:32px auto; padding:0 16px; }
        /* melhora padding lateral em telas muito estreitas c/ notch */
        @media (max-width: 480px) {
          .wrap { padding-left: max(16px, var(--safe-left)); padding-right: max(16px, var(--safe-right)); }
        }
        /* ======= HERO (intro band) ======= */
        .r-hero {
          background: radial-gradient(1200px 300px at 50% -50px, rgba(19,81,216,.12), transparent 60%);
          padding: 14px 0 6px;
          text-align:center;
        }
        .r-title {
          font-size: clamp(28px, 4.8vw, 48px);
          font-weight: 800;
          color: var(--TEXT_DARK, #0f172a);
          letter-spacing:-.01em;
          margin: 6px 0 10px;
          line-height:1.08;
        }
        .r-sub {
          color: var(--primary, #1351d8);
          font-weight: 750;
          font-size: clamp(18px, 2.1vw, 20px);
          line-height:1.55;
          margin: 8px auto 18px;
          max-width: 80ch;
        }
        .topic-pills {
          display:grid;
          grid-template-columns: 1fr;
          gap:12px;
          max-width:980px;
          margin:0 auto 18px;
          text-align:left;
        }
        @media (min-width:720px){
          .topic-pills{ grid-template-columns: repeat(3, minmax(0,1fr)); }
        }
        .pill {
          position:relative;
          display:flex; gap:12px; align-items:flex-start;
          padding:14px 16px;
          border-radius:14px;
          background:#fff;
          border:1px solid #e7e7ef;
          box-shadow: 0 8px 22px rgba(2,6,23,.06);
          transition: transform .2s ease, box-shadow .2s ease;
        }
        .pill:hover{ transform: translateY(-2px); box-shadow: 0 14px 30px rgba(2,6,23,.10); }
        .pill-icon{
          flex:0 0 36px; height:36px; width:36px; border-radius:10px;
          display:grid; place-items:center;
          background: linear-gradient(180deg, rgba(19,81,216,.12), rgba(19,81,216,.0));
          border:1px solid #dfe3f2;
        }
        .pill h4{ margin:0 0 4px; font-size:16px; color: var(--TEXT_DARK, #0f172a); }
        .pill p{ margin:0; font-size:14.5px; line-height:1.55; color:#1f2937; }

        /* ======= MAIN LAYOUT: sticky side + content ======= */
        .r-body{
          display:grid; gap:24px; align-items:start;
          grid-template-columns: 1fr;
          margin-top: 12px;
        }
        @media (min-width:980px){
          .r-body{ grid-template-columns: .85fr 1.15fr; gap:28px; }
        }
        /* side (desktop sticky) */
        .r-side{
          position:sticky; top:88px;
          background:#fff; border:1px solid #e7e7ef; border-radius:14px;
          padding:14px; box-shadow:0 8px 22px rgba(2,6,23,.06);
        }
        .r-side h5{ margin:0 0 10px; font-size:14px; letter-spacing:.04em; text-transform:uppercase; color:#334155; }
        .toc{ list-style:none; margin:0; padding:0; }
        .toc li a{
          display:block; padding:8px 10px; border-radius:10px; text-decoration:none;
          color: var(--TEXT_DARK, #0f172a); border:1px dashed rgba(19,81,216,.25);
          margin:6px 0;
        }
        .toc li a:hover{ background:rgba(19,81,216,.06); }

        /* content cards */
        .r-content{ display:grid; gap:18px; }
        .section{
          background:#fff; border:1px solid #e7e7ef; border-radius:16px;
          padding:18px 20px; box-shadow:0 12px 28px rgba(2,6,23,.08);
          scroll-margin-top: clamp(64px, 10vh, 96px); /* ancora segura sob topbars */
        }
        .section h2{
          margin:0 0 8px; font-size: clamp(18px, 2.6vw, 24px); line-height:1.25;
          color: var(--TEXT_DARK, #0f172a);
        }
        .section .lede{ color:#0f172a; font-weight:600; }
        .section p{ margin:10px 0 0; font-size:16.2px; line-height:1.75; color:#1f2937; }
        .chips{ display:flex; flex-wrap:wrap; gap:8px; margin-top:10px; }
        .chip{
          font-size:12.5px; padding:4px 10px; border-radius:999px;
          border:1px solid #cfd6f3; background:#f7f9ff; color:#0f172a; font-weight:700;
        }
        /* links mantêm underline tracejada do site */
        .wrap a{ color: var(--primary, #1351d8); text-decoration:none; border-bottom:1px dashed rgba(19,81,216,.35); }
        .wrap a:hover{ text-decoration:underline; }
        /* ======= AJUSTES ESPECÍFICOS PARA MOBILE ======= */
        @media (max-width: 979px){
          /* desativa sticky, reduz sombra e troca TOC por chips roláveis */
          .r-side{
            position: static; top:auto;
            box-shadow:none; border:1px dashed rgba(19,81,216,.25); background:#f8fafc;
          }
          .r-side h5{ display:none; }
          .toc{
            display:flex; gap:8px; overflow-x:auto; padding:8px 2px 10px;
            -webkit-overflow-scrolling: touch;
          }
          .toc li{ flex:0 0 auto; }
          .toc li a{ margin:0; border-style: solid; background:#fff; white-space:nowrap; }

          /* menos sombra para rolagem suave */
          .section{ box-shadow:0 6px 16px rgba(2,6,23,.06); }

          /* respiros menores para caber melhor em 360–414px */
          .section { padding:16px 14px; }
          .pill { padding:12px 14px; gap:10px; }
        }

        /* reduz efeitos em usuários com motion reduzido */
        @media (prefers-reduced-motion: reduce) {
          .pill:hover { transform:none; box-shadow: 0 8px 22px rgba(2,6,23,.06); }
        }
        </style>
        <div class="wrap">
          <!-- ===== HERO ===== -->
          <section class="r-hero" aria-label="Research overview">
            <div class="r-title">Research</div>
            <p class="r-sub">Our Research is focused on 3 main topics:</p>
            <div class="topic-pills">
              <div class="pill">
                <div class="pill-icon">
                  <!-- droplet -->
                  <svg width="18" height="18" viewBox="0 0 24 24" fill="none" aria-hidden="true">
                    <path d="M12 3s6 6.4 6 10.2A6 6 0 1 1 6 13.2C6 9.4 12 3 12 3z" stroke="currentColor" stroke-width="1.6"/>
                  </svg>
                </div>
                <div>
                  <h4>Liquid biopsy tools</h4>
                  <p>The development and application of Liquid biopsy tools.</p>
                </div>
              </div>
              <div class="pill">
                <div class="pill-icon">
                  <!-- helix -->
                  <svg width="18" height="18" viewBox="0 0 24 24" fill="none" aria-hidden="true">
                    <path d="M7 3c4 0 4 6 10 6M7 21c4 0 4-6 10-6M7 9c4 0 4-6 10-6M7 15c4 0 4 6 10 6" stroke="currentColor" stroke-width="1.6"/>
                  </svg>
                </div>
                <div>
                  <h4>Cancer epigenetics</h4>
                  <p>The study of methyltransferase enzymes in solid tumors and leukemias.</p>
                </div>
              </div>
              <div class="pill">
                <div class="pill-icon">
                  <!-- edit gene -->
                  <svg width="18" height="18" viewBox="0 0 24 24" fill="none" aria-hidden="true">
                    <path d="M4 14c4 0 4-4 8-4m0 0c2 0 3 1 4 2m-4-2c-2 0-3-1-4-2M3 20l3-.6 9.4-9.4a2.1 2.1 0 0 0-3-3L3 16.4 3 20z" stroke="currentColor" stroke-width="1.6"/>
                  </svg>
                </div>
                <div>
                  <h4>CRISPR/Cas9 gene editing</h4>
                  <p>Selective knockout (KO) of cancer-related targets.</p>
                </div>
              </div>
            </div>
          </section>
          <!-- ===== BODY ===== -->
          <div class="r-body">
            <aside class="r-side" aria-label="On this page">
              <h5>On this page</h5>
              <ul class="toc">
                <li><a href="#liquid-biopsy">Liquid Biopsy Applications in Cancer</a></li>
                <li><a href="#epigenetics">Cancer Epigenetics</a></li>
                <li><a href="#gene-editing">CRISPR/Cas9 Gene Editing</a></li>
              </ul>
            </aside>
            <main class="r-content">
              <section id="liquid-biopsy" class="section" aria-labelledby="liquid-biopsy-h">
                <h2 id="liquid-biopsy-h">Liquid Biopsy Applications in Cancer</h2>
                <p class="lede">Our research line focuses on the development and clinical application of liquid biopsy as a non-invasive tool for early cancer detection, real-time monitoring of therapeutic response, and improved patient management.</p>
                <p>We investigate circulating tumor DNA (ctDNA) in blood samples, aiming to identify and quantify genetic alterations associated with oncogenesis and tumor progression. This includes the detection of frequent mutations, gene amplifications, and absolute copy numbers of carcinogenic nucleic acid fragments. Using advanced technologies such as next-generation sequencing (NGS) and digital PCR, we design mutation panels tailored to cancer-specific profiles.
                A key aspect of this work is the evaluation of ctDNA as a diagnostic screening method for early cancer detection, as well as a dynamic biomarker to monitor tumor burden in patients undergoing targeted therapies. This approach enables the real-time assessment of treatment effectiveness and supports more personalized therapeutic decisions.
                We are also applying these strategies to colorectal cancer (CRC), the second most common and third deadliest cancer in Brazil. By assessing the prognostic value of ctDNA in CRC patients, we aim to improve recurrence monitoring and clinical management after surgical resection—the current curative standard of care. This project is supported by CNPq/MS and FAPDF, reinforcing our commitment to translating liquid biopsy into practice for better cancer outcomes.</p>
                <p></p>
                <p></p>
                <div class="chips" aria-label="Funding and focus tags">
                  <span class="chip">NGS</span><span class="chip">digital PCR</span><span class="chip">ctDNA</span>
                  <span class="chip">CRC</span><span class="chip">Screening</span><span class="chip">Therapy monitoring</span>
                </div>
              </section>
              <section id="epigenetics" class="section" aria-labelledby="epigenetics-h">
                <h2 id="epigenetics-h">Cancer Epigenetics</h2>
                <p>Our laboratory investigates how epigenetic regulators, including protein methyltransferases and demethylases, influence the development and progression of solid tumors and leukemias. By studying their roles in gene regulation, chromatin organization, and genomic stability, we aim to uncover how their dysregulation drives cancer. We also seek to identify novel biomarkers and therapeutic targets, translating our findings into more precise and effective cancer treatments. We use a variety of approaches to investigate their association with carcinogenesis and to identify new prognostic markers and therapeutic targets. Our work highlights how the dysregulation of these enzymes are not just passive modifiers of the genome but active drivers of cancer biology, opening opportunities for more precise and effective treatments.</p>
                <div class="chips" aria-label="Funding and focus tags">
                  <span class="chip">Leukemias</span><span class="chip">Chromatin organization</span><span class="chip">Methyltransferases</span>
                </div>
              </section>
              <section id="gene-editing" class="section" aria-labelledby="gene-editing-h">
                <h2 id="gene-editing-h">CRISPR/Cas9 Gene Editing</h2>
                <p>CRISPR-Cas9 is a powerful technology that allows scientists to precisely modify DNA within living cells. It uses a guide RNA to locate a specific DNA sequence and the Cas9 enzyme to make a targeted cut at that location. This cut can then be repaired by the cell, allowing us to disable, modify, or insert genes. In this context, our group is actively to development CRISPR/Cas9-based Gene editing tools for selective knockout (KO) of cancer-related targets, with a primary focus on epigenetic modifiers. Although CRISPR is a relatively recent innovation in biotechnology, it has rapidly become an essential research tool. Nevertheless, optimizing and implementing this technology still requires significant effort, which can exceed the capacity of many research laboratories. Our laboratory aims to serve as a facilitator, making CRISPR accessible to other research groups through collaborative efforts or service provision. By bypassing the resource-intensive process of optimization, we provide practical, ready-to-use solutions for immediate applications, thereby contributing to the advancement of regional and national scientific research.</p>
                <div class="chips" aria-label="Funding and focus tags">
                  <span class="chip">KO</span><span class="chip">Cas9 nuclease</span><span class="chip">CRISPR/Cas9</span>
                  <span class="chip">Clonal isolation</span><span class="chip">Off-target analysis</span>
                </div>
              </section>
            </main>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )







def page_publications():
    """
    Publications + Patents (yearly sections, with sub-sections)
    - Journal/other publications and Patents are grouped by YEAR.
    - Inside each year, items are split into two blocks: “Publications” and “Patents”.
    - If a DOI is provided it’s used directly; otherwise we try Crossref; else fall back to Google Scholar.
    - New links you sent are included and open directly.
    """

    from collections import defaultdict
    from functools import lru_cache
    

    try:
        import requests  # keep optional; code still works without it
    except Exception:
        requests = None

    def _ext_icon_svg():
        return (
            '<span class="ext" aria-hidden="true"></span>'
        )

    @lru_cache(maxsize=256)
    def find_doi(title: str, journal: str = "", year: int | None = None) -> str | None:
        # Use provided DOI if we already cached it
        key = f"{journal.strip()}|{title.strip()}"
        if key in DOI_CACHE:
            return DOI_CACHE[key]
        if not requests:
            return None
        try:
            params = {"query.bibliographic": title, "rows": 1}
            if year:
                params["filter"] = f"from-pub-date:{year}-01-01,until-pub-date:{year}-12-31"
            r = requests.get("https://api.crossref.org/works", params=params, timeout=6)
            if r.ok:
                items = r.json().get("message", {}).get("items", [])
                if items:
                    doi = items[0].get("DOI")
                    if isinstance(doi, str) and "/" in doi:
                        return doi
        except Exception:
            return None
        return None

    def _scholar_query_link(title: str, journal: str) -> str:
        q = urllib.parse.quote_plus((title + " " + journal).strip())
        return f"https://scholar.google.com/scholar?q={q}"

    def _google_patents_link(number_or_title: str) -> str:
        q = urllib.parse.quote_plus(number_or_title)
        return f"https://patents.google.com/?q={q}"

    st.markdown(
        """
        <style>
          .pub-wrap { max-width:1180px; margin:28px auto 40px; padding:0 16px; }

          .pub-h1   { font-weight:900; font-size:clamp(26px,3.6vw,38px); margin:6px 0 0; color:var(--TEXT_DARK,#0f172a); }
          .pub-sub  { color:#475569; font-size:.98rem; margin: 4px 0 18px; }

          /* year index chips */
          .year-index { display:flex; flex-wrap:wrap; gap:8px; margin: 12px 0 18px; }
          .year-chip {
            display:inline-block; padding:6px 10px; border:1px solid #e6e8ee; border-radius:999px;
            text-decoration:none; font-weight:700; font-size:.92rem; color:#0f172a; background:#fff;
          }
          .year-chip:hover { border-color:#cdd3e0; }

          /* year section */
          .year-section { margin: 26px 0 24px; scroll-margin-top: 90px; }
          .year-title {
            font-weight:900; font-size:clamp(18px,2.6vw,24px); color:#0f172a; margin:0 0 10px;
            display:flex; align-items:center; gap:10px;
          }
          .year-count {
            font-weight:700; color:#475569; font-size:.88rem; background:#f1f5f9; border:1px solid #e2e8f0;
            padding:2px 8px; border-radius:999px;
          }

          /* subsection headers (Publications / Patents) */
          .subhead {
            display:flex; align-items:center; gap:10px; margin: 10px 0 8px;
            font-weight:900; color:#0f172a; font-size: clamp(14px,1.8vw,18px);
          }
          .subhead .bar {
            height:3px; border-radius:2px; background:rgba(63,94,181,.35); flex:1;
          }

          /* common list + item card */
          .item-list { list-style:none; padding:0; margin:0; display:grid; gap:10px; }
          .item {
            background:#fff; border:1px solid #e6e8ee; border-radius:14px;
            padding:12px 14px; display:grid; grid-template-columns: 1fr auto; gap:10px; align-items:start;
            position:relative; overflow:hidden;
            transition: box-shadow .15s ease, border-color .15s ease, transform .15s ease;
          }
          .item:hover { box-shadow:0 8px 18px rgba(2,6,23,.08); border-color:#d9dde3; transform: translateY(-2px); }
          .item .lead { font-weight:800; color:var(--primary, #1351d8); text-decoration:none; }
          .item .lead:hover { text-decoration:underline; }
          .item .meta { color:#475569; font-size:.92rem; margin-top:4px; }
          .aside { display:flex; gap:8px; align-items:center; }

          .doi-badge, .link-badge {
            display:inline-flex; align-items:center; gap:6px;
            font-size:.78rem; padding:3px 9px; border-radius:999px;
            border:1px solid #cbd5e1; color:#334155; background:#f8fafc; text-decoration:none;
          }
          .doi-badge:hover, .link-badge:hover { border-color:#aab7c7; }

          .ext {
            display:inline-block; width:12px; height:12px;
            background: conic-gradient(from 90deg, #64748b, #94a3b8);
            -webkit-mask: url('data:image/svg+xml;utf8,<svg viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg"><path fill="black" d="M14 3h7v7h-2V6.41l-9.29 9.3-1.42-1.42 9.3-9.29H14V3z"/><path fill="black" d="M5 5h5V3H3v7h2V5zm0 14v-5H3v7h7v-2H5z"/></svg>') center/contain no-repeat;
                    mask: url('data:image/svg+xml;utf8,<svg viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg"><path fill="black" d="M14 3h7v7h-2V6.41l-9.29 9.3-1.42-1.42 9.3-9.29H14V3z"/><path fill="black" d="M5 5h5V3H3v7h2V5zm0 14v-5H3v7h7v-2H5z"/></svg>') center/contain no-repeat;
          }

          /* patent flair */
          .item.patent::before {
            content:""; position:absolute; left:0; top:0; bottom:0; width:3px;
            background: linear-gradient(#fde68a,#fca5a5);
          }
          .patent .meta { color:#475569; }

          .back-top { display:inline-block; margin:10px 0 0; font-size:.88rem; color:#0f172a; text-decoration:none; }
          .back-top:hover { text-decoration:underline; }
        </style>
        """, unsafe_allow_html=True
    )

    PUBLICATIONS = [
        {"authors": "Lopes, M. A.; Cordeiro, M. E. R.; Barreto, F. A. T.; Moreno, L. S.; Silva, A. A. M.; Loyola, M. B.; Soares, M. V. A.; Sousa, J. B.; Pittella-Silva, F.",
         "title": "Assessment of cfDNA release dynamics during colorectal cancer surgery",
         "journal": "Oncotarget", "year": 2025},
        {"authors": "Hollanda, C. N.; Gualberto, A. C. M.; Motoyama, A. B.; Pittella-Silva, F.",
         "title": "Advancing Leukemia Management Through Liquid Biopsy: Insights into Biomarkers and Clinical Utility",
         "journal": "Cancers", "year": 2025},

        {"authors": "Olivera Santana, B. L.; De Loyola, M. B.; Gualberto, A. C. M.; Pittella-Silva, F.",
         "title": "Genetic Alterations of SMYD4 in Solid Tumors Using Integrative Multi-Platform Analysis",
         "journal": "International Journal of Molecular Sciences", "year": 2024},
        {"authors": "Gatto, C. C.; Cavalcante, C. de Q. O.; Lima, F. C.; Nascimento, É. C. M.; Martins, J. B. L.; Santana, B. L. O.; Gualberto, A. C. M.; Pittella-Silva, F.",
         "title": "Structural Design, Anticancer Evaluation, and Molecular Docking of Newly Synthesized Ni(II) Complexes with ONS-Donor Dithiocarbazate Ligands",
         "journal": "Molecules", "year": 2024},
        {"authors": "Silva-Carvalho, A. É.; Féliu-Braga, L. D. C.; Bogéa, G. M. R.; de Assis, A. J. B.; Pittella-Silva, F.; Saldanha-Araujo, F.",
         "title": "GLP and G9a histone methyltransferases as potential therapeutic targets for lymphoid neoplasms",
         "journal": "Cancer Cell International", "year": 2024},

        {"authors": "Zhu, Y.; Pittella-Silva, F.; Wang, Y.; Wang, T.; Long, F.",
         "title": "Editorial: Epigenetic Regulation and Therapy Resistance in Cancer",
         "journal": "Frontiers in Pharmacology", "year": 2023},
        {"authors": "Cavalcante, C. de Q. O.; da Mota, T. H. A.; de Oliveira, D. M.; Nascimento, É. C. M.; Martins, J. B. L.; Pittella-Silva, F.; Gatto, C. C.",
         "title": "Dithiocarbazate ligands and their Ni(II) complexes with potential biological activity: Structural, antitumor and molecular docking study",
         "journal": "Frontiers in Molecular Biosciences", "year": 2023},
        {"authors": "Mota, T. H. A.; Carmargo, R.; Biojone, E. R.; Guimarães, A. F. R.; Pittella-Silva, F.; Oliveira, D. M.",
         "title": "The Relevance of Telomerase and Telomere-Associated Proteins in B-Acute Lymphoblastic Leukemia",
         "journal": "Genes", "year": 2023},
        {"authors": "Machado, C. M. L.; Skubal, M.; Haedicke, K.; Silva, F. P.; Stater, E. P.; Silva, T. L. A. O.; Costa, É. T.; Masotti, C.; Otake, A. H.; Andrade, L. N. S.; et al.",
         "title": "Membrane-derived particles shed by PSMA-positive cells function as pro-angiogenic stimuli in tumors",
         "journal": "Journal of Controlled Release", "year": 2023},
        {"authors": "Assis, A. J.; Santana, B.; Gualberto, A. C. M.; Silva, F. P.; Pittella-Silva, F.",
         "title": "Therapeutic applications of CRISPR/Cas9 mediated targeted gene editing in acute lymphoblastic leukemia: current perspectives, future challenges, and clinical implications",
         "journal": "Frontiers in Pharmacology", "year": 2023},

        {"authors": "Pittella-Silva, F.",
         "title": "Targeting DNA Methylation as an Epigenetic Leukocyte Counting Tool",
         "journal": "Clinical Chemistry", "year": 2022},
        {"authors": "Watanabe, K.; Nakamura, T.; Kimura, Y.; Motoya, M.; Kojima, S.; Kuraya, T.; …; Pittella-Silva, F.; Nakamura, Y.; Low, A. S.",
         "title": "Tumor-informed approach improved ctDNA detection rate in resected pancreatic cancer",
         "journal": "International Journal of Molecular Sciences", "year": 2022},

        {"authors": "Silva, F. P.; Kimura, Y.; Low, S.-K.; Nakamura, Y.; Motoya, M.",
         "title": "Amplification of mutant in a patient with advanced metastatic pancreatic adenocarcinoma detected by liquid biopsy: A case report",
         "journal": "Molecular and Clinical Oncology", "year": 2021},

        {"authors": "Pittella-Silva, F.; Chin, Y. M.; Chan, H. T.; Nagayama, S.; Miyauchi, E.; Low, S.-K.; Nakamura, Y.",
         "title": "Plasma or Serum: Which Is Preferable for Mutation Detection in Liquid Biopsy?",
         "journal": "Clinical Chemistry", "year": 2020},

        {"authors": "Doralina do Amaral Rabello Carolina Amaro de Moura Rosangela Vieira de Andrade Andrea Barretto Motoyama Fabio Pittella Silva", "title": "Altered expression of MLL methyltransferase family genes in breast cancer)", "journal": "International Journal of Oncology",
         "year": 2013, "url": "https://www.spandidos-publications.com/10.3892/ijo.2013.1981",
         "doi": "10.3892/ijo.2013.1981"},
        
        {"authors": "Luis Henrique Toshihiro Sakamoto a c d, Rosangela Vieira de Andrade b, Maria Sueli Soares Felipe b c, Andrea Barretto Motoyama, Fabio Pittella Silva", "title": "SMYD2 is highly expressed in pediatric acute lymphoblastic leukemia and constitutes a bad prognostic factor", "journal": "",
         "year": 2014, "url": "https://www.sciencedirect.com/science/article/pii/S0145212614000319?via%3Dihub"},
        
        {"authors": "Fabiana Pirani Carneiro Maria Imaculada Muniz-Junqueira Fabio Pittella-Silva Marcos de Vasconcelos Carneiro Gustavo Henrique Soares Takano Leonora Maciel de Sousa Vianna Luciano Barbosa De Andrade Tercia Maria Mendes Lousa De Castro Isabela Peres Tatiana Karla Dos Santos Borges Vânia Moraes Ferreira Andrea Barretto Motoyama", "title": "A panel of markers for identification of malignant and non-malignant cells in culture from effusions", "journal": "Oncology Reports",
         "year": 2017, "url": "https://www.spandidos-publications.com/10.3892/or.2017.6022",
         "doi": "10.3892/or.2017.6022"},
        
        {"authors": "Fabiana Pirani Carneiro, Andersen Charles Darós, Adriana Cysneiro Milhomem Darós, Tércia Maria Mendes Lousa de Castro, Marcos de Vasconcelos Carneiro, Cecília Ramos Fidelis, Mariane Vieira Vilioni, Michelle Egídio da Costa Matsunaga, Jéssica Meneses Othon Sidou, Mariana Anaue Lozi Dias Chaves, Lívia Custódio Pereira, Ceres Nunes de Resende, ", "title": "Cervical Cytology of Samples with Ureaplasma urealyticum, Ureaplasma parvum, Chlamydia trachomatis, Trichomonas vaginalis, Mycoplasma hominis, and Neisseria gonorrhoeae Detected by Multiplex PCR", "journal": "Mediators of Inflammation",
         "year": 2020, "url": "https://onlinelibrary.wiley.com/doi/10.1155/2020/7045217",
         "doi": "10.1155/2020/7045217"},
      {
        "authors": "MOTA, TALES H. A.; GUIMARAES, A. F. R.; CARVALHO, A. E. S.; SALDANHA-ARAUJO, F.; LOPES, G. P. F.; PITTELLA-SILVA, F.; RAMOS, D. A.; OLIVEIRA, D. M.",
        "title": "Effects of in vitro short- and long-term treatment with telomerase inhibitor in U-251 glioma cells",
        "journal": "Tumor Biology",
        "year": 2021,
        "url": "https://doi.org/10.3233/TUB-211515",
        "doi": "10.3233/TUB-211515"
      },
      {
        "authors": "PITTELLA-SILVA, F.; MOTOYA, M.; KIMURA, Y.; NAKAMURA, T.; LOW, S.-K.; NAKAMURA, Y.",
        "title": "Assessment of mutation burden after clinical intervention in PDAC and BTC via profiling circulating tumor DNA (ctDNA)",
        "journal": "Journal of Clinical Oncology",
        "year": 2020,
        "url": "https://doi.org/10.1200/JCO.2020.38.15_suppl.e15529",
        "doi": "10.1200/JCO.2020.38.15_suppl.e15529"
      },
      {
        "authors": "CHEN, S.; WIEWIORA, R. P.; MENG, F.; BABAULT, N.; MA, A.; YU, W.; QIAN, K.; HU, H.; ZOU, H.; WANG, J.; FAN, S.; BLUM, G.; PITTELLA-SILVA, F.; BEAUCHAMP, K. A.; TEMPEL, W.; JIANG, H.; CHEN, K.; SKENE, R. J.; ZHENG, Y. G.; BROWN, P. J.; et al.",
        "title": "The dynamic conformational landscape of the protein methyltransferase SETD8",
        "journal": "eLife",
        "year": 2019,
        "url": "https://doi.org/10.7554/eLife.45403",
        "doi": "10.7554/eLife.45403"
      },
      {
        "authors": "MORAIS, K. S.; ARCANJO, D. S.; LOPES, G. P. F.; SILVA, G. G.; MOTA, T. H. A.; GABRIEL, T. R.; RABELLO RAMOS, D. A.; PITTELLA-SILVA, F.; OLIVEIRA, D. M.",
        "title": "Long-term in vitro treatment with telomerase inhibitor MST-312 induces resistance by selecting long telomeres cells",
        "journal": "Cell Biochemistry and Function",
        "year": 2019,
        "url": "https://doi.org/10.1002/cbf.3398",
        "doi": "10.1002/cbf.3398"
      },
      {
        "authors": "ALVES-SILVA, J. C.; DE CARVALHO, J. L.; RABELLO, D. A.; SEREJO, T. R. T.; REGO, E. M.; NEVES, F. A. R.; LUCENA-ARAUJO, A. R.; PITTELLA-SILVA, F.; SALDANHA-ARAUJO, F.",
        "title": "GLP overexpression is associated with poor prognosis in Chronic Lymphocytic Leukemia and its inhibition induces leukemic cell death",
        "journal": "Investigational New Drugs",
        "year": 2018,
        "url": "https://doi.org/10.1007/s10637-018-0613-x",
        "doi": "10.1007/s10637-018-0613-x"
      },
      {
        "authors": "RABELLO, D. A.; FERREIRA, V. D. S.; BERZOTI-COELHO, M. G.; BURIN, S. M.; MAGRO, C. L.; CACEMIRO, M. C.; SIMÕES, B. P.; SALDANHA-ARAUJO, F.; DE CASTRO, F. A.; PITTELLA-SILVA, F.",
        "title": "MLL2/KMT2D and MLL3/KMT2C expression correlates with disease progression and response to imatinib mesylate in chronic myeloid leukemia",
        "journal": "Cancer Cell International",
        "year": 2018,
        "url": "https://doi.org/10.1186/s12935-018-0523-1",
        "doi": "10.1186/s12935-018-0523-1"
      },
      {
        "authors": "CARVALHO ALVES-SILVA, J.; RABELLO, D. A.; BRAVO, M. O.; LUCENA-ARAUJO, A. R.; DE OLIVEIRA, D. M.; DE OLIVEIRA, F. M.; REGO, E. M.; PITTELLA-SILVA, F.; SALDANHA-ARAUJO, F.",
        "title": "Aberrant levels of SUV39H1 and SUV39H2 methyltransferase are associated with genomic instability in chronic lymphocytic leukemia",
        "journal": "Environmental and Molecular Mutagenesis",
        "year": 2017,
        "url": "https://doi.org/10.1002/em.22128",
        "doi": "10.1002/em.22128"
      },
      {
        "authors": "COLAÇO, C. S.; DE MATOS, A. R.; ESTRÊLA, M. S.; ROCHA-JÚNIOR, M. C.; OTAGUIRI, K. K.; RODRIGUES, E. S.; TAKAYANAGUI, O. M.; COVAS, D. T.; KASHIMA, S.; PITTELLA SILVA, F.; HADDAD, R.",
        "title": "Downregulation of histone methyltransferase EHMT2 in CD4+ T-cells may protect HTLV-1-infected individuals against HAM/TSP development",
        "journal": "Archives of Virology",
        "year": 2017,
        "url": "https://doi.org/10.1007/s00705-017-3428-8",
        "doi": "10.1007/s00705-017-3428-8"
      },
      {
        "authors": "MORAIS, K. S.; GUIMARÃES, A. F. R.; RAMOS, D. A. R.; PITTELLA-SILVA, F.; DE OLIVEIRA, D. M.",
        "title": "Long-term exposure to MST-312 leads to telomerase reverse transcriptase overexpression in MCF-7 breast cancer cells",
        "journal": "Anti-Cancer Drugs",
        "year":2017,
        "url": "https://doi.org/10.1097/CAD.0000000000000508",
        "doi": "10.1097/CAD.0000000000000508"
      },
      {
        "authors": "BUTLER, K. V.; MA, A.; YU, W.; LI, F.; TEMPEL, W.; BABAULT, N.; PITTELLA-SILVA, F.; SHAO, J.; WANG, J.; LUO, M.; VEDADI, M.; BROWN, P. J.; ARROWSMITH, C. H.; JIN, J.",
        "title": "Structure-Based Design of a Covalent Inhibitor of the SET Domain-Containing Protein 8 (SETD8) Lysine Methyltransferase",
        "journal": "Journal of Medicinal Chemistry",
        "year": 2016,
        "url": "https://doi.org/10.1021/acs.jmedchem.6b01244",
        "doi": "10.1021/acs.jmedchem.6b01244"
      },
      {
        "authors": "OLIVEIRA-SANTOS, W.; RABELLO, D. A.; LUCENA-ARAUJO, A. R.; OLIVEIRA, F. M.; REGO, E. M.; PITTELLA-SILVA, F.; SALDANHA-ARAUJO, F.",
        "title": "Residual expression of SMYD2 and SMYD3 is associated with the acquisition of complex karyotype in chronic lymphocytic leukemia",
        "journal": "Tumor Biology",
        "year": 2016,
        "url": "https://doi.org/10.1007/s13277-016-4846-z",
        "doi": "10.1007/s13277-016-4846-z"
      },
      {
        "authors": "DE OLIVEIRA, D. M.; LIMA, R. M. F.; CLARENCIO, J.; VELOZO, E. S.; DE AMORIM, I. A.; MOTA, T. H. A.; COSTA, S. L.; PITTELLA-SILVA, F.; EL-BACHÁ, R. D. S.",
        "title": "The classical photoactivated drug 8-methoxypsoralen and related compounds are effective without UV light irradiation against glioma cells",
        "journal": "Neurochemistry International",
        "year": 2016,
        "url": "https://doi.org/10.1016/j.neuint.2016.06.004",
        "doi": "10.1016/j.neuint.2016.06.004"
      },
      {
        "authors": "RABELLO, D. A.; LUCENA-ARAUJO, A. R.; ALVES-SILVA, J. C. R.; DA EIRA, V. B. A. S.; DE VASCONCELLOS, M. C. C.; OLIVEIRA, F. M.; REGO, E. M.; SALDANHA-ARAUJO, F.; PITTELLA SILVA, F.",
        "title": "Overexpression of EZH2 associates with a poor prognosis in chronic lymphocytic leukemia",
        "journal": "Blood Cells, Molecules and Diseases",
        "year": 2015,
        "url": "https://doi.org/10.1016/j.bcmd.2014.07.013",
        "doi": "10.1016/j.bcmd.2014.07.013"
      },
      {
        "authors": "FIGUEIREDO, D.; RABELLO, D.; ZANETTE, D.; SAGGIORO, F.; MAMEDE, R.; ZAGO, M.; DA SILVA, W.; SALDANHA-ARAÚJO, F.; PITTELLA SILVA, F.",
        "title": "Assessment of MLL methyltransferase gene expression in larynx carcinoma",
        "journal": "Oncology Reports",
        "year": 2015,
        "url": "https://doi.org/10.3892/or.2015.3756",
        "doi": "10.3892/or.2015.3756"
      },
      {
        "authors": "LAFOREST, A.; APARICIO, T.; ZAANAN, A.; PITTELLA-SILVA, F.; DESREBEAUX, A.; LE CORRE, D.; PALLIER, K.; AUST, D.; BLONS, H.; SVRCEK, M.; LAURENT-PUIG, P.",
        "title": "ERBB2 gene as a potential therapeutic target in small bowel adenocarcinoma",
        "journal": "European Journal of Cancer",
        "year": 2014,
        "url": "https://doi.org/10.1016/j.ejca.2014.04.007",
        "doi": "10.1016/j.ejca.2014.04.007"
      },
      {
        "authors": "FARIA, J. A. Q. A.; CORREA, N. C. R.; ANDRADE, C.; CAMPOS, A. C. A.; ALMEIDA, R. S. S.; GOES, A. M.; GOMES, D. A.; PITTELLA-SILVA, F.",
        "title": "SETD4 is a Newly Identified Cytosolic and Nuclear Lysine Methyltransferase involved in Breast Cancer Cell Proliferation",
        "journal": "Journal of Cancer Science & Therapy",
        "year": 2013,
        "url": "https://doi.org/10.4172/1948-5956.1000185",
        "doi": "10.4172/1948-5956.1000185"
      },
      {
        "authors": "Davi Trombini Aleixoa, Estael L. C. Cruz-Cazarim, Kezia C. B. Ferreira, Lívia N. Grossi, Wilson R. Braz, Natália P. Silva, Fábio Pittella-Silva, Marina F. Dias, Silvia L. Fialho, Guilherme D. Tavares, Frederico Pittella",
        "title": "Breakthroughs in the nanoparticle-mediated delivery of siRNA for breast cancer treatment",
        "journal": "Journal of Cancer Science & Therapy",
        "year": 2025,
        "url": "https://doi.org/10.1080/17435889.2025.2567842",
        "doi": "10.1080/17435889.2025.2567842"
      },
      {
        "authors": "MORENO, LARA DE SOUZA ; DA SILVA, TONY ALEXANDRE MEDEIROS ; SOARES, MAYRA VELOSO AYRIMORAES ; DE CARVALHO, JOÃO LUIZ AZEVEDO ; PITTELLA-SILVA, FABIO.",
        "title": "Machine Learning Radiomics in Computed Tomography for Prediction of Tumor and Nodal Stages in Colorectal Cancer",
        "journal": "Cancers",
        "year": 2026,
        "url": "https://doi.org/10.3390/cancers18030377",
        "doi": "10.3390/cancers18030377"
      }
        

    ]

    DOI_CACHE = {
        "International Journal of Oncology|Article (Spandidos – IJO)": "10.3892/ijo.2013.1981",
        "Oncology Reports|Article (Spandidos – Oncol Rep.)": "10.3892/or.2017.6022",
        "Mediators of Inflammation|Article (Wiley – Mediators of Inflammation)": "10.1155/2020/7045217",
    }

    PATENTS = [
        {
            "year": 2013,
            "inventors": "SILVA, F. P. PITTELLA-SILVA, F.; Gomes, A. D.; Lunardi, C. N.; Gomes, D. A.",
            "title": "Composição antitumoral à base de micro e nanopartículas poliméricas, seu processo de obtenção e suas aplicações",
            "jurisdiction": "Brazil (INPI)",
            "number": "BR102013017203",
            "deposited": "2013-07-04",
            "url": _google_patents_link("BR102013017203"),
        },
        {
            "year": 2015,
            "inventors": "LUO, M.; BLUM, G.; SANCHEZ, G. I.; YANG, L.; SILVA, F. P. PITTELLA-SILVA, F.",
            "title": "Naphthaquinone Methyltransferase Inhibitors and Uses Thereof",
            "jurisdiction": "WIPO (PCT)",
            "number": "WO/2015/172076",
            "deposited": "2015-08-05",
            "granted": "2015-12-11",
            "url": _google_patents_link("WO2015172076"),
        },
        {
            "year": 2015,
            "inventors": "LUO, M.; BLUM, G.; SANCHEZ, G. I.; SILVA, F. P. PITTELLA-SILVA, F.",
            "title": "Naphthaquinone Methyltransferase Inhibitors and Uses Thereof",
            "jurisdiction": "Canada (CIPO)",
            "number": "CA2946609",
            "deposited": "2015-12-11",
            "url": _google_patents_link("CA2946609"),
        },
        {
            "year": 2016,
            "inventors": "LUO, M.; BLUM, G.; SANCHEZ, G. I.; SILVA, F. P. PITTELLA-SILVA, F.",
            "title": "Naphthaquinone Methyltransferase Inhibitors and Uses Thereof",
            "jurisdiction": "Australia (IP Australia)",
            "number": "AU2015255692",
            "deposited": "2016-11-03",
            "url": _google_patents_link("AU2015255692"),
        },
        {
            "year": 2017,
            "inventors": "LUO, M.; BLUM, G.; SANCHEZ, G. I.; SILVA, F. P. PITTELLA-SILVA, F.",
            "title": "Naphthaquinone Methyltransferase Inhibitors and Uses Thereof",
            "jurisdiction": "Japan (JPO)",
            "number": "JP2017511562",
            "deposited": "2017-06-15",
            "url": _google_patents_link("JP2017511562"),
        },
        {
            "year": 2016,
            "inventors": "LUO, M.; WANG, J.; SILVA, F. P. PITTELLA-SILVA, F.",
            "title": "SET8 INHIBITORS AND USES THEREOF",
            "jurisdiction": "USA (USPTO)",
            "number": "US 62/287,263 (provisional)",
            "deposited": "2016-01-26",
            "url": _google_patents_link("62/287,263 SET8 inhibitors"),
        },
    ]

    pubs_by_year: dict[int, list[dict]] = defaultdict(list)
    pats_by_year: dict[int, list[dict]] = defaultdict(list)

    for p in PUBLICATIONS:
        pubs_by_year[int(p["year"])].append(p)
    for pt in PATENTS:
        pats_by_year[int(pt["year"])].append(pt)

    years = sorted(set(pubs_by_year.keys()) | set(pats_by_year.keys()), reverse=True)

    st.markdown('<div class="pub-wrap" id="top">', unsafe_allow_html=True)
    st.markdown('<div class="pub-h1">Publications</div>', unsafe_allow_html=True)
    st.markdown('<div class="pub-sub">Peer-reviewed articles and patents, grouped by year.</div>', unsafe_allow_html=True)

    chips = [f'<a class="year-chip" href="#year-{y}">{y}</a>' for y in years]
    st.markdown(f'<nav class="year-index">{"".join(chips)}</nav>', unsafe_allow_html=True)

    for y in years:
        pub_items = pubs_by_year.get(y, [])
        pat_items = pats_by_year.get(y, [])

        st.markdown(f'<section class="year-section" id="year-{y}">', unsafe_allow_html=True)
        st.markdown(f'<div class="year-title">{y} '
                    f'<span class="year-count">{len(pub_items) + len(pat_items)}</span></div>', unsafe_allow_html=True)

        # ---- Publications
        if pub_items:
            st.markdown('<div class="subhead">Publications <span class="bar"></span></div>', unsafe_allow_html=True)
            st.markdown('<ul class="item-list">', unsafe_allow_html=True)
            for p in pub_items:
                title = (p.get("title") or "").strip()
                journal = (p.get("journal") or "").strip()
                authors = (p.get("authors") or "").strip()
                doi = (p.get("doi") or None)
                url = (p.get("url") or None)

                href = None
                badge_html = ""
                if doi:
                    href = f"https://doi.org/{urllib.parse.quote(doi, safe='/')}"
                    badge_html = f'<a class="doi-badge" href="{href}" target="_blank" rel="noopener">DOI {_ext_icon_svg()}</a>'
                elif url:
                    href = url
                    badge_html = f'<a class="link-badge" href="{html.escape(url)}" target="_blank" rel="noopener">Open {_ext_icon_svg()}</a>'
                else:
                    guess = find_doi(title, journal, y)
                    if guess:
                        href = f"https://doi.org/{urllib.parse.quote(guess, safe='/')}"
                        badge_html = f'<a class="doi-badge" href="{href}" target="_blank" rel="noopener">DOI {_ext_icon_svg()}</a>'
                    else:
                        href = _scholar_query_link(title, journal)
                        badge_html = f'<a class="link-badge" href="{href}" target="_blank" rel="noopener">Search {_ext_icon_svg()}</a>'

                meta_bits = []
                if authors: meta_bits.append(html.escape(authors))
                if journal: meta_bits.append(f"<i>{html.escape(journal)}</i>")
                meta = " · ".join(meta_bits)

                st.markdown(
                    f"""
                    <li class="item">
                      <div class="pub-main">
                        <a class="lead" href="{html.escape(href)}" target="_blank" rel="noopener">{html.escape(title)}</a>
                        {'<div class="meta">'+meta+'</div>' if meta else ''}
                      </div>
                      <div class="aside">{badge_html}</div>
                    </li>
                    """,
                    unsafe_allow_html=True,
                )
            st.markdown('</ul>', unsafe_allow_html=True)

        if pat_items:
            st.markdown('<div class="subhead">Patents <span class="bar"></span></div>', unsafe_allow_html=True)
            st.markdown('<ul class="item-list">', unsafe_allow_html=True)
            for pt in pat_items:
                title = pt["title"]
                num = pt.get("number", "")
                jur = pt.get("jurisdiction", "")
                inv = pt.get("inventors", "")
                dep = pt.get("deposited", "")
                grant = pt.get("granted", "")
                url = pt.get("url") or _google_patents_link(f'{num} {title}')

                meta_bits = []
                if inv:  meta_bits.append(html.escape(inv))
                if jur:  meta_bits.append(html.escape(jur))
                if dep:  meta_bits.append(f"Deposit: {html.escape(dep)}")
                if grant: meta_bits.append(f"Grant: {html.escape(grant)}")
                meta = " · ".join(meta_bits)

                st.markdown(
                    f"""
                    <li class="item patent">
                      <div class="pub-main">
                        <a class="lead" href="{html.escape(url)}" target="_blank" rel="noopener">
                          {html.escape(title)} ({html.escape(num)})
                        </a>
                        {'<div class="meta">'+meta+'</div>' if meta else ''}
                      </div>
                      <div class="aside">
                        <a class="link-badge" href="{html.escape(url)}" target="_blank" rel="noopener">Open {_ext_icon_svg()}</a>
                      </div>
                    </li>
                    """,
                    unsafe_allow_html=True,
                )
            st.markdown('</ul>', unsafe_allow_html=True)

        st.markdown('<a class="back-top" href="#top">Back to top ↑</a>', unsafe_allow_html=True)
        st.markdown('</section>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

def page_equipment():
    """
    Equipment page (no search bar)
    - Responsive card grid
    - Full-card click -> PNIPE equipment page in a new tab
    """


    EQUIPMENT = [
        {"name": "Capela de Fluxo Laminar NB2", "url": "https://pnipe.mcti.gov.br/equipment/10574",
         "function": "Class II laminar-flow biosafety cabinet for aseptic sample handling."},
        {"name": "Countless 3", "url": "https://pnipe.mcti.gov.br/equipment/59150",
         "function": "Automated cell counter for rapid cell concentration and viability."},
        {"name": "Centrífuga de bancada", "url": "https://pnipe.mcti.gov.br/equipment/10578",
         "function": "Routine bench centrifugation of tubes and microplates."},
        {"name": "digital PCR", "url": "https://pnipe.mcti.gov.br/equipment/59231",
         "function": "Absolute nucleic-acid quantification by partitioned PCR."},
        {"name": "DNA Workstation", "url": "https://pnipe.mcti.gov.br/equipment/10570",
         "function": "Clean workstation dedicated to DNA/PCR setup to prevent contamination."},
        {"name": "Fluorímetro Qubit 3", "url": "https://pnipe.mcti.gov.br/equipment/59230",
         "function": "Sensitive fluorometric quantification of DNA, RNA, and protein."},
        {"name": "Incubadora de Células com CO₂ (PHP)", "url": "https://pnipe.mcti.gov.br/equipment/59148",
         "function": "CO₂ incubator for controlled-temperature, humidified cell culture."},
        {"name": "Incubadora de Células com CO₂ (Sanyo)", "url": "https://pnipe.mcti.gov.br/equipment/10575",
         "function": "CO₂ incubator for mammalian cell growth under stable conditions."},
        {"name": "Incubadora Shaker", "url": "https://pnipe.mcti.gov.br/equipment/10576",
         "function": "Temperature-controlled orbital shaking for cultures and reactions."},
        {"name": "Ion Chef", "url": "https://pnipe.mcti.gov.br/equipment/10565",
         "function": "Automated template preparation and chip loading for Ion Torrent NGS."},
        {"name": "NanoDrop Plus", "url": "https://pnipe.mcti.gov.br/equipment/59149",
         "function": "Micro-volume UV-Vis spectrophotometer for nucleic acids and protein."},
        {"name": "NanoVue – Espectrofotômetro DNA", "url": "https://pnipe.mcti.gov.br/equipment/10572",
         "function": "Micro-volume spectrophotometry for DNA/RNA quantification."},
        {"name": "Pala Cell Sorter and Single Cell Dispenser", "url": "https://pnipe.mcti.gov.br/equipment/59229",
         "function": "Sorting and dispensing of single cells into plates for downstream assays."},
        {"name": "QuantStudio 3D – PCR Digital", "url": "https://pnipe.mcti.gov.br/equipment/10564",
         "function": "Chip-based digital PCR for precise copy-number and rare variant detection."},
        {"name": "SeptOne Plus – Real-Time PCR", "url": "https://pnipe.mcti.gov.br/equipment/10568",
         "function": "qPCR amplification with real-time fluorescence detection."},
        {"name": "Sequenciador NGS – Ion S5 Plus", "url": "https://pnipe.mcti.gov.br/equipment/10563",
         "function": "Next-generation sequencing for targeted panels and small genomes."},
        {"name": "Sistema de Eletroforese Automatizado – Experion", "url": "https://pnipe.mcti.gov.br/equipment/10566",
         "function": "Automated electrophoresis for RNA/DNA/protein sizing and quality."},
        {"name": "Tanque de Criopreservação", "url": "https://pnipe.mcti.gov.br/equipment/10579",
         "function": "Liquid-nitrogen storage for long-term cryopreservation."},
        {"name": "TapeStation 4150", "url": "https://pnipe.mcti.gov.br/equipment/17245",
         "function": "Automated electrophoresis for DNA/RNA QC using ScreenTape."},
        {"name": "Termociclador PCR", "url": "https://pnipe.mcti.gov.br/equipment/10571",
         "function": "Conventional thermal cycler for endpoint PCR."},
        {"name": "Transiluminador UV", "url": "https://pnipe.mcti.gov.br/equipment/10577",
         "function": "UV visualization of nucleic acids and gels."},
        {"name": "Ultrafreezer −80 °C", "url": "https://pnipe.mcti.gov.br/equipment/10573",
         "function": "Ultra-low temperature storage of biological samples (−80 °C)."},
    ]

    st.markdown(dedent("""
    <style>
      :root{
        --eq-ink: var(--TEXT_DARK, #0f172a);
        --eq-muted:#55637b;
        --eq-line:#e6e8ee;
        --eq-accent: var(--bar-bg, #3f5eb5);
        --eq-card:#fff;
      }
      .eq-wrap{ max-width:1180px; margin:0 auto; padding: 8px 16px 28px; }
      .eq-h1{ font-weight:900; font-size:clamp(24px,3.4vw,34px); color:var(--eq-ink); margin:18px 0 6px; }
      .eq-sub{ color:#4b5563; margin-bottom:14px; }

      .eq-grid{ display:grid; grid-template-columns:repeat(auto-fit,minmax(280px,1fr)); gap:14px; }
      .eq-card{
        display:block; text-decoration:none; color:inherit;
        background:var(--eq-card);
        border:1px solid var(--eq-line); border-radius:14px; padding:14px;
        transition:transform .12s ease, box-shadow .12s ease, border-color .12s ease, background .12s ease;
        position:relative;
      }
      .eq-card::before{
        content:""; position:absolute; left:0; top:0; bottom:0; width:3px;
        background: linear-gradient(#bfdbfe,#93c5fd); border-radius:3px 0 0 3px; opacity:.8;
      }
      .eq-card:hover{
        transform:translateY(-2px); box-shadow:0 10px 18px rgba(2,6,23,.08); border-color:#d9dde3; background:#ffffff;
      }
      .eq-title{ font-weight:900; color:var(--eq-ink); line-height:1.18; }
      .eq-func{ color:var(--eq-muted); font-size:.92rem; margin-top:4px; }

      .eq-open{
        display:inline-flex; gap:6px; align-items:center; margin-top:10px;
        font-size:.86rem; color:var(--eq-accent); font-weight:800;
      }
      .eq-open .ext{
        width:12px; height:12px; display:inline-block; background: currentColor;
        -webkit-mask:url('data:image/svg+xml;utf8,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24"><path fill="black" d="M14 3h7v7h-2V6.41l-9.29 9.3-1.42-1.42 9.3-9.29H14V3z"/><path fill="black" d="M5 5h5V3H3v7h2V5zm0 14v-5H3v7h7v-2H5z"/></svg>') center/contain no-repeat;
                mask:url('data:image/svg+xml;utf8,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24"><path fill="black" d="M14 3h7v7h-2V6.41l-9.29 9.3-1.42-1.42 9.3-9.29H14V3z"/><path fill="black" d="M5 5h5V3H3v7h2V5zm0 14v-5H3v7h7v-2H5z"/></svg>') center/contain no-repeat;
      }
      .eq-card:focus{ outline:none; box-shadow:0 0 0 3px rgba(59,130,246,.35); }
    </style>
    """), unsafe_allow_html=True)

    st.markdown('<div class="eq-wrap">', unsafe_allow_html=True)
    st.markdown('<div class="eq-h1">Equipment</div>', unsafe_allow_html=True)
    st.markdown('<div class="eq-sub">Click any card to open the PNIPE equipment page.</div>', unsafe_allow_html=True)

    cards = ['<div class="eq-grid">']
    for it in EQUIPMENT:
        name = html.escape(it["name"])
        fn   = html.escape(it.get("function",""))
        url  = html.escape(it["url"])
        cards.append(dedent(f"""
        <a class="eq-card" href="{url}" target="_blank" rel="noopener noreferrer" tabindex="0" aria-label="{name}">
          <div class="eq-title">{name}</div>
          {f'<div class="eq-func">{fn}</div>' if fn else ''}
          <div class="eq-open"><span class="ext" aria-hidden="true"></span> Open on PNIPE</div>
        </a>
        """).strip())
    cards.append("</div>")
    st.markdown("".join(cards), unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

    show_help_widget(
    name="Victor Fleck",
    email="Victor.fleck906438@gmail.com",
    role="Undergraduate Pharmacy Student • Web Developer :) ",
    phone="+55 61 991281207",
    label="Problems with the website?"
    )

def page_people():


    BASE_PHOTO_DIR = Path(__file__).parent / "static" / "people"

    PEOPLE = [
        {"name": "Fábio Pittella Silva, PhD", "role": "Principal Investigator - Labhead",
         "photo": "Pittella.jpeg",
         "bio": "Head of the Laboratory of Molecular Pathology of Cancer at the University of Brasília. Ph.D in Molecular Pathology from the Human Genome Center, University of Tokyo (2008); M.Sc. in Medical Sciences from the University of Tokyo (2004); B.Sc. in Pharmacy and Biochemistry (Clinical Analysis) from the Federal University of Juiz de Fora (2001). At the University of Tokyo’s Human Genome Center, identified and described the genes WDRPUH, TIPUH1, and SMYD3—the first methyltransferase implicated in human carcinogenesis. Professor at the University of Brasília (since 2008), supervising master’s, doctoral, and postdoctoral trainees. Associate researcher in liquid biopsy at the Cancer Precision Medicine Center, Japanese Foundation for Cancer Research (JFCR), Tokyo (since 2018). Visiting professor at Memorial Sloan Kettering Cancer Center, New York (2014–2018), working on methyltransferase inhibitors. Expertise in molecular genetics, with focus on epigenetics, protein methyltransferases, liquid biopsy, and molecular targets for cancer diagnosis/therapy. Ad-hoc consultant to DECIT/SCTIE/MS, CNPq, FAPITEC/SE, and FAPDF; co-founder of the innovation company i.rad.Particles",
         "degrees": [
             "B.Sc. in Pharmacy & Biochemistry, Federal University of Juiz de Fora (1995–2001)",
             "Complementary Studies in Medical Sciences, University of Tokyo (2001–2002)",
             "M.Sc. in Medical Sciences, University of Tokyo (2002–2004)",
             "Ph.D. in Molecular Pathology, University of Tokyo (2004–2008)",
             "Advanced Training: Semiconductor Sequencing & Digital PCR in Cancer, Université Paris Descartes – Paris V (2012)",
             "Postdoctoral Research in Molecular Pathology, Memorial Sloan Kettering Cancer Center (2014–2015)"
         ],
         "email": "pittella@unb.br", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/4961976945101495", "orcid": "https://orcid.org/0000-0002-9644-7098"},
        {"name": "Andrea Barretto Motoyama, PhD", "role": "Principal Investigator - Labhead",
         "photo": "andrea.jpg",
         "bio": "Head of the Laboratory of Molecular Pathology of Cancer at the University of Brasília. Associate Professor IV in the Department of Pharmacy at the University of Brasília. She earned a BSc in Biological Sciences (UnB, 1997) and a PhD in Biochemistry (University of Basel, 2002). Her PhD explored signal-transduction pathways that drive cellular resistance to targeted anticancer drugs such as trastuzumab. She completed postdoctoral training at The Scripps Research Institute (RNA stability) and the Burnham Institute for Medical Research (cell-adhesion proteins and Rho-family GTPases). Her work spans cell biology, biochemistry, molecular biology, oncology, and pharmacology",
         "degrees": ["BSc in Biological Sciences, University of Brasília (1997)",
                     "PhD in Biochemistry, University of Basel (2002)",
                     "Postdoctoral Fellow in RNA Stability, The Scripps Research Institute, La Jolla",
                     "Postdoctoral Fellow in Adhesion Proteins & Rho-family GTPases, Burnham Institute for Medical Research, La Jolla"],
         "email": "andreabm@unb.br", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/6229434599232057", "orcid": "https://orcid.org/0000-0002-9280-2638"},
        
        {"name": "Ana Cristina Moura Gualberto, PhD", "role": "Postdoctoral Fellow",
         "photo": "anagualberto.jpg", "bio": "BSc and teaching licensure in Biological Sciences from UFJF, followed by an MSc and PhD in Immunology investigating matrix metalloproteinases in breast cancer, including RNAi-based silencing. She completed a postdoc at UFJF on adipose-tissue exosomes in tumor progression and is currently a postdoctoral fellow at the University of Brasília working with CRISPR/Cas9 genome editing in cancer. Her experience includes immune response, breast-cancer animal models, cell culture, molecular biology, confocal microscopy, histology and immunohistochemistry, and functional assays of migration, proliferation, and cell viability.",
         "degrees": ["BSc and Teaching Licensure in Biological Sciences, Federal University of Juiz de Fora (2008–2012)",
                     "MSc in Biological Sciences (Immunology), Federal University of Juiz de Fora — focus on matrix metalloproteinases as markers of breast cancer progression (2013–2015)",
                     "PhD in Biological Sciences (Immunology), Federal University of Juiz de Fora — RNAi silencing of matrix metalloproteinases in breast cancer progression (2015–2019)",
                     "Postdoctoral Fellow, Federal University of Juiz de Fora — adipose tissue exosomes in breast cancer",
                     "Postdoctoral Fellow, University of Brasília, CRISPR/Cas9 genome editing in cancer (current)"], "email": "", "institution": "Federal University of Juiz de Fora",
         "lattes": "http://lattes.cnpq.br/6338541953564765", "orcid": ""},
        
        {"name": "Luis Janssen Maia, PhD", "role": "Postdoctoral Fellow",
         "photo": "luisJ.jpg", "bio": "Experience in microbiology and molecular biology. Work focuses on high-throughput nucleic-acid sequencing (NGS) using both in-house data and public datasets—to study the diversity and molecular evolution of viruses and bacteria, with applications to molecular studies and liquid biopsy in tumors",
         "degrees": ["BSc in Biological Sciences, University of Brasília",
                     "MSc in Molecular Biology, University of Brasília",
                     "PhD in Molecular Biology, University of Brasília"], "email": "", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/7167611866112740", "orcid": ""},
        
        {"name": "Ekaly Ivoneth Porto Apangha, B.Pharm", "role": "Technical assistance",
         "photo": "ekaly.jpg", "bio": "Pharmacist graduated from the University of Brasília (UnB); Specialist in Cancer Care from the Escola Superior de Ciências da Saúde (ESCS). Experienced in chemotherapy compounding (antineoplastic handling), Clinical Pharmacy, Oncologic Clinical Pharmacy, Hospital Pharmacy, and Community Pharmacy",
         "degrees": ["Specialization – Multiprofessional Residency in Cancer Care, ESCS/FEPECS (2018–2020)",
                     "B.Pharm. (Pharmacy), University of Brasília (2011–2016)"], "email": "ekaly.apangha@unb.br", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/0622839652506358", "orcid": ""},
        
        {"name": "André Araújo de Medeiros Silva, MD", "role": "Coloproctologist & General Surgeon • PhD Student",
         "photo": "Andre.jpg", "bio": "Master’s graduate and current PhD candidate in Medical Sciences at the University of Brasília, focused on the molecular pathology of colorectal cancer. Holds a Biomedicine degree (UniCEUB, 2017) and a specialization in Imaging from Hospital Israelita Albert Einstein (2019). Qualified in clinical analysis, molecular biology, and imaging. Currently works in the Radiology Department (CT unit) at Hospital Sírio-Libanês in Brasília",
         "degrees": ["PhD Candidate in Medical Sciences, University of Brasília (2024–present)",
                     "MSc in Medical Sciences, University of Brasília (2021–2023)",
                     "Medical Residency in Coloproctology, Hospital São Rafael (2008–2010)",
                     "Medical Residency in Videolaparoscopy (R3 General Surgery), Hospital Geral Roberto Santos (2007–2008)",
                     "Medical Residency in General Surgery, Hospital Geral Clériston Andrade (2005–2007)",
                     "MD, Escola Bahiana de Medicina e Saúde Pública (1999–2004)"
                     "PhD Candidate in Medical Sciences (Molecular Pathology of Colorectal Cancer), University of Brasília"], "email": "", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/6330213791982657", "orcid": "https://orcid.org/0000-0002-3691-3739"},
        {"name": "Flávio de Alencar Teles Barreto, MSc", "role": "PhD Student",
         "photo": "flavio.jpg", "bio": "Master’s graduate and current PhD candidate in Medical Sciences at the University of Brasília, focused on the molecular pathology of colorectal cancer. Holds a Biomedicine degree (UniCEUB, 2017) and a specialization in Imaging from Hospital Israelita Albert Einstein (2019). Qualified in clinical analysis, molecular biology, and imaging. Currently works in the Radiology Department (CT unit) at Hospital Sírio-Libanês in Brasília",
         "degrees": ["BSc in Biomedicine, Centro Universitário de Brasília (2017)",
                     "Postgraduate Specialization in Imaging (Imagenologia), Hospital Israelita Albert Einstein (2019)",
                     "MSc in Medical Sciences (Molecular Pathology of Colorectal Cancer), University of Brasília",
                     "PhD Candidate in Medical Sciences (Molecular Pathology of Colorectal Cancer), University of Brasília"], "email": "flavioalencarbarreto92@gmail.com", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/7778492502300819", "orcid": ""},
        {"name": "Brunna Letícia Oliveira Santana, MSc", "role": "PhD Student",
         "photo": "bru.jpg", "bio": "Biologist with an MSc in Medical Sciences focused on cancer epigenetics, now a PhD candidate at the University of Brasília investigating lysine methyltransferase knockouts to identify molecular and therapeutic biomarkers for acute lymphoblastic leukemia",
         "degrees": ["BSc in Biological Sciences (Bachelor), Universidade Paulista – Brasília (2019)",
                     "MSc in Medical Sciences (Cancer Epigenetics), University of Brasília (2022–2024)",
                     "PhD Candidate in Medical Sciences, University of Brasília — Lysine methyltransferase gene knockouts as potential molecular and therapeutic biomarkers for acute lymphoblastic leukemia"], "email": "brunna.los@hotmail.com", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/5131211187661647", "orcid": "https://orcid.org/0000-0003-4402-8358"},
        {"name": "Matheus de Oliveira Andrade, MD", "role": "Clinical Oncologist • PhD Student",
         "photo": "matheusmd.jpg", "bio": "MD graduated from the University of Brasília (UnB), with an exchange program in Biomedical Sciences at Queen Mary University of London (QMUL). Fellowship in Internal Medicine at HCFMRP-USP and in Medical Oncology at ICESP-FMUSP. Currently, he is a clinical oncologist at Rede Américas in Brasília (Hospital Brasília Lago Sul) and a PhD candidate in Medical Sciences at the University of Brasília, focusing on liquid biopsy in colorectal cancer.",
         "degrees": ["PhD Candidate in Medical Sciences, University of Brasília (2025–present)",
                     "Medical Residency in Medical Oncology, ICESP – FMUSP (2022–2025)",
                     "Medical Residency in Internal Medicine, HCFMRP–USP (2020–2022)",
                     "Exchange/Advanced Training in Biomedical Sciences, Queen Mary University of London (2014–2015)",
                     "MD, University of Brasília (2011–2018)"], "email": "matheusandradeoncologia@gmail.com", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/3451033370311538", "orcid": "https://orcid.org/0000-0001-8922-2715"},
        {"name": "Mayra Veloso Ayrimoraes Soares, MD", "role": "Radiologist • PhD Student",
         "photo": "mayramd.jpg", "bio": "MD graduated from the University of Brasília (UnB, 1999), Physician–radiologist with residency training in Radiology and Diagnostic Imaging at Hospital de Base do Distrito Federal (2001–2003). Supervisor of the Radiology and Diagnostic Imaging Residency Program at the University of Brasília. Radiologist at Hospital Sírio-Libanês (Brasília-DF) and at Exame/DASA-DF. Holds an MBA in Management of Clinics, Hospitals and Health Services (FGV, Brasília) and is currently a PhD candidate in Medical Sciences at the University of Brasília",
         "degrees": ["PhD Candidate in Medical Sciences, University of Brasília (2022–present)",
                     "Medical Residency in Radiology and Diagnostic Imaging (R3 optional year), Hospital de Base do Distrito Federal – HBDF (2003–2004)",
                     "Medical Residency in Radiology and Diagnostic Imaging, Hospital de Base do Distrito Federal – HBDF (2001–2003)",
                     "Lato Sensu Specialization in Medical Residency Preceptorship (SUS), Hospital Sírio-Libanês (2015–2016)",
                     "MD, University of Brasília (1993–1999)"], "email": "mayra.veloso@unb.br", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/4358326060572502", "orcid": "https://orcid.org/0000-0003-0796-5123"},
        {"name": "Larissa Fernanda Campos Moreira da Silva, MD", "role": "Clinical Oncologist • PhD Student",
            "photo": "larissa.jpg", "bio": "Graduated in Medicine from the Federal University of the State of Rio de Janeiro (2003), with residencies in Internal Medicine and Medical Oncology. Currently heads the Oncology Department at the Brasília Military Area Hospital, with managerial and attending responsibilities. Leads the lung-cancer screening project across military hospitals in the Federal District and serves as regional lead for Clinical Oncology in military hospitals of Brazil’s Midwest. Serves on the Brazilian Army’s Technical Chamber of Oncology, with responsibility for breast, skin, and tumor-agnostic indications. Practices as a medical oncologist at Oncologia Alvorada (Américas group). From 2020 to June 2025, served at Oncologia Santa Marta as medical coordinator, leading care pathways and the lung-cancer screening protocol. Pursuing a PhD in Medical Sciences (oncology and molecular biology applied to medicine) at the University of Brasília. Thesis: “Plasma expression of miRNAs: miRNA-145-5p, miRNA-145-3p2, miRNA-155, miR-21, miR-10b, and let-7 in patients with breast cancer. Established three infusion centers: two military (Military Area Hospital of Recife and Military Area Hospital of Brasília) and one civilian intra-hospital center at Hospital Santa Marta (Oncologia Santa Marta). Previous experience includes attending physician at Hospital Rios D’or, preceptor for the Internal Medicine residency at Santa Casa de Misericórdia de Campo Grande, faculty in Medical Skills at the University for the Development of the State and the Pantanal Region, and clinical roles at the Brazilian Army Central Hospital and other oncology reference institutions. Extensive background in service management, creation of infusion centers, clinical research, implementation of care pathways, and delivery of both outpatient and inpatient oncology care.",
            "degrees": ["PhD Candidate in Medical Sciences, University of Brasília (2022–present)",
                        "Master's Candidate in Medical Sciences, University of Brasília (2022–present)",
                        "Medical Residency in Internal Medicine, UNIRIO (2004)",
                        "Specialization in Military Sciences (Officer Improvement Course), ESAO (2020–2021)",
                        "Postgraduate Specialization in Clinical Oncology, Hospital Central do Exército (2013–2016)",
                        "Lato Sensu Postgraduate in Military Sciences, Escola de Saúde do Exército (2011)",
                        "MD (Medicine), UNIRIO (1997–2003) "], "email": "", "institution": "University of Brasília",
        "lattes": "http://lattes.cnpq.br/2380982097052834", "orcid": "https://orcid.org/0000-0001-5896-7204"},
           {"name": "Ísis de Araujo Oliveira Nakashoji, MSc", "role": "PhD Student",
            "photo": "isi.jpg", "bio": "Holds a B.Sc. in Pharmacy (University of Brasília, 2018), a Specialization in Clinical Analysis and an M.Sc. in Medical Sciences (University of Brasília, 2021). Currently a PhD candidate in Medical Sciences at the University of Brasília, evaluating the impact of HPV vaccination among women treated in Brazil’s public health system (SUS). Works as a biochemical pharmacist with the Federal District Health Department, in hospital pharmacy. Experience spans Molecular Biology, Cytopathology, Cancer Cells, Immunohistochemistry, and Hospital Pharmacy.",
            "degrees": ["PhD Candidate in Medical Sciences, University of Brasília (2021–present)",
                        "M.Sc. in Medical Sciences, University of Brasília (2019–2021)",
                        "Lato Sensu Postgraduate Specialization in Traditional Chinese Medicine, Curante Cursos Educacional (2022–2024)",
                        "Specialization in Clinical Analysis, Universidade Paulista – UNIP (2019)",
                        "B.Sc. in Pharmacy, University of Brasília (2014–2018)"], "email": "", "institution": "University of Brasília",
            "lattes": "http://lattes.cnpq.br/1743833485792512", "orcid": "https://orcid.org/0000-0002-9303-1169"},
           {"name": "Lara De Souza Moreno, MD", "role": "Radiologist • MSc Student",
            "photo": "lara.jpg", "bio": "Holds a medical degree from the Catholic University of Brasília (2020). Residency in Radiology and Diagnostic Imaging at the University Hospital of Brasília (2025). Master’s in Medical Sciences from the University of Brasília (2024). Internal Medicine fellow at the DFSTAR Radiology Service – Rede D’Or.",
            "degrees": ["Master’s Candidate in Medical Sciences, University of Brasília (2024–present)",
                        "Medical Residency in Radiology and Diagnostic Imaging, University Hospital of Brasília / Ebserh (2022–2025)",
                        "Fellowship in Radiology and Diagnostic Imaging (Internal Medicine), Hospital DF Star – Rede D’Or (2025–present)",
                        "M.D., Catholic University of Brasília (2014–2020)"], "email": "larasmoreno@gmail.com", "institution": "University of Brasília",
            "lattes": "http://lattes.cnpq.br/4568437840946460", "orcid": "https://orcid.org/0009-0002-3390-4551"},
           {"name": "Debora Rodrigues Silveira, B.Pharm", "role": "MSc Student",
            "photo": "dbr.jpg", "bio": "Bachelor in Pharmaceutical Sciences from the University of Brasília (UnB), with experience in clinical research and cancer molecular genetics, plus internships across multiple levels of healthcare. Professional experience at ANVISA and participation in outreach projects and academic events, developing skills in health surveillance, health education, and scientific organization. Training is complemented by courses in pharmacotherapy, public policy, and public health. Interests include research, innovation, and qualified practice in biomedical sciences, genetics, and molecular biology.",
            "degrees": ["M.Sc. Candidate in Medical Sciences, University of Brasília (2025–present)",
                "B.Sc. in Pharmaceutical Sciences, University of Brasília (2018–2024)"], "email": "", "institution": "University of Brasília",
            "lattes": "http://lattes.cnpq.br/5490962025847708", "orcid": ""},
        
        {"name": "Mariana Braccialli de Loyola, B.Pharm", "role": "MSc Student",
         "photo": "Mari.jpg", "bio": "Pharmacist from the University of Brasília and current Master’s student in the Graduate Program in Medical Sciences at the University of Brasília. Works at the Cancer Molecular Pathology Laboratory, researching lysine methyltransferases in the context of acute lymphoblastic leukemia development.",
         "degrees": ["Master’s Candidate in Medical Sciences, University of Brasília (2024–present)",
                     "B.Pharm. (Pharmacy), University of Brasília (2017–2023)"], "email": "maribraccialli@gmail.com", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/4480047944308175", "orcid": "https://orcid.org/0000-0002-7816-764X"},
        
        {"name": "Israel James Cutrim de Oliveira Filho, MD", "role": "MSc Student",
         "photo": "israel.jpg", "bio": "MD graduated from the Federal University of Maranhão (UFMA), Pinheiro campus (2022). Former Dentistry student (UFMA; program discontinued in 2016). Worked in urgent and emergency care and served in the Mais Médicos program for two years in Eldorado do Sul, RS (2023–2024). Currently a Radiology and Diagnostic Imaging resident at the University Hospital of Brasília (2025–present).",
         "degrees": ["Medical Residency in Radiology and Diagnostic Imaging, University Hospital of Brasília / Ebserh (2025–present)",
                     "M.D., Federal University of Maranhão (2017–2022)"], "email": "", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/0956170261692464", "orcid": ""},
        
        {"name": "Daniel dos Santos Ramos, B.Pharm", "role": "MSc Student",
         "photo": "daniel.jpg", "bio": "Holds a B.Pharm. from the Euro-American Institute of Education, Science and Technology (2013) and a postgraduate specialization in Forensic Sciences from UniLS. Currently a Master’s student in the Graduate Program in Nanoscience and Nanobiotechnology at the University of Brasília (UnB). Served as a Public Policy and Educational Management Analyst at the Federal District Department of Education (2010–2022). Now works as a pharmacist and Responsible Technical Officer of the Green Nanotechnology Laboratory (NAVE), affiliated with the School of Health Sciences and Technologies at the University of Brasília, Ceilândia campus (FCTS/UnB)",
         "degrees": ["Master’s Candidate in Nanoscience and Nanobiotechnology, University of Brasília (2025–present)",
                     "Specialization in Forensic Sciences, UniLS (2016–2017)",
                     "B.Pharm. (Pharmacy), Euro-American Institute of Education, Science and Technology – EUROAM (2009–2013)"], "email": "daniel.ramos@unb.br", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/4153360465428627", "orcid": ""},
        
        {"name": "Sophia Alves da Costa Pereira", "role": "Undergraduate Student",
         "photo": "sofia1.jpg", "bio": "Bachelor’s student in Pharmacy at the University of Brasília (UnB). Graduated in Pre-Pharmacy – Associate in Science from Kansas City Kansas Community College, USA (2024)",
         "degrees": ["B.Sc. Candidate in Pharmacy, University of Brasília (2019–present)",
                     "Associate in Science – Pre-Pharmacy, Kansas City Kansas Community College (2024)"], "email": "", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/7674938627453757", "orcid": ""},
        
        {"name": "Yasmin Albuquerque", "role": "Undergraduate Student",
         "photo": "yas.jpg", "bio": "Undergraduate student in Biological Sciences (B.Sc.) at the University of Brasília (since 2022). Has experience in Molecular Biology with a focus on diagnostics, including RNA extraction, RT-PCR, and PCR for Plant Virology. Also works in human and medical genetics, specializing in the extraction and purification of circulating free DNA (cfDNA) from liquid biopsy and in the development of an alternative in-house cfDNA extraction and purification kit.",
         "degrees": ["B.Sc. Candidate in Biological Sciences, University of Brasília (2022–present)"], "email": "", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/2305511706458935", "orcid": ""},
        
        {"name": "Sofia Arrais Haidar", "role": "Undergraduate Student",
         "photo": "sofia2.jpg", "bio": "Fourth-year medical student at the University of Brasília (UnB), currently seeking experience and training in healthcare, especially in scientific research. Completed high school in 2020 at Colégio Educacional Sigma. Has academic and clinical practice experience in outpatient care, inpatient wards, and operating rooms, with activities at the University Hospital of Brasília (HUB) and Primary Health Units (SES-DF). Actively participates in regional and national congresses, symposia, and courses, as well as academic leagues: Family and Community Medicine (LAMEF-UnB), Cardiology (LACor-UnB), Surgery (LAC-UnB), Hematology (LAHem-UnB), and Neurology (Neuroliga-UnB)",
         "degrees": ["M.D. student, University of Brasília (2021–present)"], "email": "", "institution": "University of Brasília",
         "lattes": " http://lattes.cnpq.br/2664347077103484", "orcid": ""},
        
        {"name": "Priscila de Medeiros Vergara", "role": "Undergraduate Student",
         "photo": "pri.jpg", "bio": "Pharmacy student at the University of Brasília (UnB) (2022–present). Conducts research in microbiology at the Enzymology Laboratory of UnB, working with microorganism cultivation, DNA extraction, and analysis of experimental results.",
         "degrees": ["B.Sc. student in Pharmacy, University of Brasília (2022–present)"], "email": "", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/9595288580014037", "orcid": ""},
        
        {"name": "Matheus Lopes da Silva", "role": "Undergraduate Student",
         "photo": "mtl.jpg", "bio": "Undergraduate student in Pharmacy at the Faculty of Health Sciences, University of Brasília (FCS-UnB), Brazil. Completed high school at Colégio Alub – Associação Lecionar Unificada de Brasília in 2013. Intermediate English. Currently serves as a teaching assistant in the subject of Sanitary Surveillance Applied to Pharmacy, taught by Prof. Helaine Carneiro Capucho (FCS-UnB) and ormer Organizational President of the junior enterprise Terapêutica Jr. (2024) and active member of the ValueHealthLab – Laboratory of Studies for Improving Quality, Patient Safety, and Value in Health, coordinated by Prof. Helaine Carneiro Capucho.",
         "degrees": ["B.Sc. student in Pharmacy, University of Brasília (2019–present)"], "email": "", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/1126659468006228", "orcid": ""},
           
           {"name": "Augusto Silva Alves", "role": "Undergraduate Student",
         "photo": "ab.jpg", "bio": "Medical student at the University of Brasília (UnB), Brazil. Scientific initiation scholarship holder at UnB.",
         "degrees": ["MD Candidate (Medicine), University of Brasília (2022–present)"], "email": "", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/1220931705760601", "orcid": "https://orcid.org/0009-0005-6099-3889"},
           
           {"name": "Ângelo Antônio Silva Lima", "role": "Undergraduate Student",
         "photo": "as.jpg", "bio": "Medical student at the University of Brasília (UnB), Brazil.",
         "degrees": ["MD Candidate (Medicine), University of Brasília (2022–present)"], "email": "", "institution": "University of Brasília",
         "lattes": "http://lattes.cnpq.br/8940572165307881", "orcid": ""},
    
    
    ]

    st.markdown("""
    <style>
      .people-wrap {max-width:1180px;margin:0 auto;padding:0 16px;}
      .person-name {font-weight:700;font-size:1.05rem;line-height:1.2;}
      .person-role {color:#374151;font-size:0.92rem;margin-top:2px;}
      /* style the st.image inside People */
      .people-wrap [data-testid="stImage"] img {
        width:90px;height:90px;object-fit:cover;border-radius:10px;
        border:1px solid #d1d5db;background:#fff;
      }
      .ph-placeholder {
        width:90px;height:90px;border-radius:10px;border:1px dashed #cbd5e1;
        display:flex;align-items:center;justify-content:center;color:#9ca3af;background:#f8fafc;
      }
      /* remove the gray rails/dividers */
      .people-wrap div[data-testid="stDivider"],
      .people-wrap hr,
      .people-wrap [role="separator"] { display:none !important; height:0 !important; border:0 !important; }
      .people-wrap div[data-testid="stExpander"] { border:none !important; box-shadow:none !important; background:transparent !important; }
      .people-wrap div[data-testid="stExpander"] summary { border:none !important; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("## People")
    st.markdown('<div class="people-wrap">', unsafe_allow_html=True)

    for p in PEOPLE:
        left, right = st.columns([0.18, 0.82], gap="large")
        with left:
            photo = p.get("photo") or ""
            if photo.startswith("http://") or photo.startswith("https://"):
                st.image(photo, width=80)
            else:
                path = (BASE_PHOTO_DIR / photo) if photo else None
                if path and path.exists():
                    st.image(str(path), width=120)
                else:
                    st.markdown('<div class="ph-placeholder">?</div>', unsafe_allow_html=True)

        with right:
            st.markdown(f'<div class="person-name">{p["name"]}</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="person-role">{p.get("role","")}</div>', unsafe_allow_html=True)
            with st.expander("See more"):
                if p.get("bio"):
                    st.write(p["bio"])
                if p.get("degrees"):
                    st.markdown("**Degrees**")
                    st.write("\n".join([f"- {d}" for d in p["degrees"]]))
                links = []
                if p.get("lattes"):
                    links.append(f"[Lattes]({p['lattes']})")
                if p.get("orcid"):
                    links.append(f"[ORCID]({p['orcid']})")
                if links:
                    st.markdown("**Links:** " + " · ".join(links))
                contact = []
                if p.get("email"):
                    contact.append(f"📧 {p['email']}")
                if p.get("institution"):
                    contact.append(f"🏛️ {p['institution']}")
                if contact:
                    st.markdown("**Contact:** " + " • ".join(contact))

    st.markdown('</div>', unsafe_allow_html=True)



    st.markdown("""
    <style>
      .alumini-title { color:#040359 !important; font-weight:800; margin: 20px 0 20px; }
      .alumini-item { margin: 6px 0; }
      .alumini-item a { font-weight:700; text-decoration:none; }
      .alumini-meta { color:#040359; margin-left:8px; }
    </style>
    """, unsafe_allow_html=True)

    ALUMINI = [
         {"name": "Athos Silva Oliveira", "lattes": "http://lattes.cnpq.br/5665031794527967", "year": "2023", "activity": "Postdoctoral Fellows"},
         {"name": "Vívian D'Afonseca da Silva Ferreira", "lattes": "http://lattes.cnpq.br/6228921073127080", "year": "2015", "activity": "Postdoctoral Fellows"},
         {"name": "Doralina do Amaral Rabello Ramos", "lattes": "http://lattes.cnpq.br/4636606799429977", "year": "2011", "activity": "Postdoctoral Fellows"},
         {"name": "Luís Henrique Toshihiro Sakamoto", "lattes": "http://lattes.cnpq.br/9517030460820076", "year": "2014", "activity": "Ph.D."},
         {"name": "João Nunes de Mattos Neto", "lattes": "http://lattes.cnpq.br/8975088486815463", "year": "2015", "activity": "Ph.D."},
         {"name": "Alan Jhones Barbosa de Assis", "lattes": "http://lattes.cnpq.br/4200078779107622", "year": "2023", "activity": "M.Sc."},
         {"name": "Mailson Alves Lopes", "lattes": "http://lattes.cnpq.br/1366042871721219", "year": "2024", "activity": "M.Sc."},
         {"name": "Brunna Letícia Oliveira Santana", "lattes": "http://lattes.cnpq.br/5131211187661647", "year": "2024", "activity": "M.Sc."},
         {"name": "Flávio de Alencar Teles Barreto", "lattes": "http://lattes.cnpq.br/7778492502300819", "year": "2024", "activity": "M.Sc."},
         {"name": "Maíra de Azevedo Feitosa Araujo", "lattes": "http://lattes.cnpq.br/4044380416858498", "year": "2009", "activity": "M.Sc."},
         {"name": "Carolina Amaro de Moura", "lattes": "http://lattes.cnpq.br/7936117807018383", "year": "2009", "activity": "M.Sc."},
         {"name": "Martha Silva Estrêla", "lattes": "http://lattes.cnpq.br/2003119441638169", "year": "2014", "activity": "M.Sc."},
         {"name": "Rubens dos Santos Samuel de Almeida", "lattes": "http://buscatextual.cnpq.br/buscatextual/visualizacv.jsp?id=K4465112U8", "year": "2014", "activity": "M.Sc."},
         {"name": "Flavia Zattar Piazera", "lattes": "http://lattes.cnpq.br/2024630817705719", "year": "2015", "activity": "M.Sc."},
         {"name": "Luís Augusto Muniz Telles", "lattes": "http://lattes.cnpq.br/1987858702905034", "year": "2016", "activity": "M.Sc."},
         {"name": "Erika Soares Vidigal", "lattes": "http://lattes.cnpq.br/8655167478334942", "year": "", "activity": "PIBIC"},
         {"name": "Carolina Queiroga de Azevedo", "lattes": "http://lattes.cnpq.br/3864899787247206", "year": "", "activity": "PIBIC"},
         {"name": "Guilherme da Rocha Ferreira", "lattes": "http://lattes.cnpq.br/8590158442391134", "year": "", "activity": "PIBIC"},
         {"name": "Larissa Carvalho Meireles", "lattes": "http://lattes.cnpq.br/0859422500973304", "year": "", "activity": "PIBIC"},
         {"name": "Clara Laís Vieira de Almeida", "lattes": "http://buscatextual.cnpq.br/buscatextual/visualizacv.jsp?id=K4465889J8", "year": "", "activity": "PIBIC"},
         {"name": "José Ranclenisson Lopes Moreira", "lattes": "http://buscatextual.cnpq.br/buscatextual/visualizacv.jsp?id=K4465881H6", "year": "", "activity": "PIBIC"},
         {"name": "Rubens dos Santos Samuel de Almeida", "lattes": "http://buscatextual.cnpq.br/buscatextual/visualizacv.jsp?id=K4465112U8", "year": "", "activity": "PIBIC"},
         {"name": "Nathalia Moraes de Vasconcelos", "lattes": "http://lattes.cnpq.br/3874243004616213", "year": "", "activity": "PIBIC"},
         {"name": "Igor Ribeiro Fernandes", "lattes": "http://lattes.cnpq.br/7452149888725969", "year": "", "activity": "PIBIC"},
         {"name": "Victor Henrique Fragoso de Mendonça Santiago Paula", "lattes": "http://lattes.cnpq.br/7903416721142160", "year": "", "activity": "PIBIC"},
         {"name": "Julia Silva Valerio Diniz", "lattes": "http://lattes.cnpq.br/8609581256832375", "year": "", "activity": "PIBIC"},
         {"name": "Anne Caroline Duarte Moreira", "lattes": "http://buscatextual.cnpq.br/buscatextual/visualizacv.jsp?id=K4465126E4", "year": "", "activity": "PIBIC"},
         {"name": "Mariana Carneiro da Cunha", "lattes": "http://lattes.cnpq.br/6954304874399333", "year": "", "activity": "PIBIC"},
         {"name": "Luís Augusto Muniz Telles", "lattes": "http://lattes.cnpq.br/1987858702905034", "year": "", "activity": "PIBIC"},
         {"name": "Brenno Vinícius Martins Henrique", "lattes": "http://buscatextual.cnpq.br/buscatextual/visualizacv.jsp?id=K4465807Y2", "year": "", "activity": "PIBIC"},
         {"name": "Isabel Santos Cardoso", "lattes": "http://lattes.cnpq.br/2615729486438083", "year": "", "activity": "PIBIC"},
         {"name": "Hugo Paceli Souza de Oliveira", "lattes": "http://lattes.cnpq.br/9505651806604331", "year": "", "activity": "PIBIC"},
         {"name": "Adriano Drummond de Abreu Barreto", "lattes": "http://lattes.cnpq.br/7298891467901069", "year": "", "activity": "PIBIC"},
         {"name": "Raul Lima Barbosa Sousa", "lattes": "http://lattes.cnpq.br/3374006283084914", "year": "", "activity": "PIBIC"},
         {"name": "Martha Silva Estrela", "lattes": "http://lattes.cnpq.br/2003119441638169", "year": "", "activity": "PIBIC"},
         {"name": "Luana Lima da Cunha", "lattes": "http://lattes.cnpq.br/6116126642316896", "year": "", "activity": "PIBIC"},
         {"name": "Lisandra Campos Cheberle", "lattes": "http://lattes.cnpq.br/5860655837290521", "year": "", "activity": "PIBIC"},
         {"name": "Luiza Akemi Takenaka Pereira", "lattes": "http://lattes.cnpq.br/3317558172750868", "year": "", "activity": "PIBIC"},
         {"name": "Talita Fernandes Nunes", "lattes": "http://lattes.cnpq.br/4910914982411340", "year": "", "activity": "PIBIC"},
         {"name": "Camila De Andrade Paula Firmino", "lattes": "http://lattes.cnpq.br/1723913807773255", "year": "", "activity": "PIBIC"},
         {"name": "Melissa Platiau", "lattes": "http://lattes.cnpq.br/4093458083592208", "year": "", "activity": "PIBIC"},
         {"name": "Mariana Braccialli de Loyola", "lattes": "http://lattes.cnpq.br/4480047944308175", "year": "", "activity": "PIBIC"},
         {"name": "Amanda Kallyne Pereira Silva", "lattes": "http://lattes.cnpq.br/6883360833820914", "year": "", "activity": "PIBIC"},
         {"name": "Ana Luísa Nunes Cantuário", "lattes": "http://lattes.cnpq.br/9498096315615965", "year": "", "activity": "PIBIC"},
         {"name": "Cíntia Nogueira Hollanda", "lattes": "http://lattes.cnpq.br/4163677223230109", "year": "", "activity": "PIBIC"}
    ]

        
    st.markdown("""
    <style>
      .alumini-title{ color:#040359 !important; font-weight:800; margin:20px 0 16px; }
      .alumini-grid{ display:grid; grid-template-columns:repeat(3,minmax(220px,1fr)); gap:8px 22px; align-items:start; }
      @media (max-width:1000px){ .alumini-grid{ grid-template-columns:repeat(2,minmax(220px,1fr)); } }
      @media (max-width:640px){ .alumini-grid{ grid-template-columns:1fr; } }
    
      .alumini-card{ padding:4px 0; }
      .alumini-card a{ font-weight:700; text-decoration:none; }
      .alumini-card a:hover{ text-decoration:underline; }
      .alumini-meta{ font-size:13px; color:#040359; opacity:.9; margin-top:2px; }
    </style>
    """, unsafe_allow_html=True)
    
    html_rows = ['<h3 class="alumini-title"></h3>', '<div class="alumini-grid">']
    
    order = ["Postdoctoral Fellows","Ph.D.","M.Sc.","PIBIC",""]
    groups = {k:[a for a in ALUMINI if (a.get("activity") or "") == k] for k in order}
    
    html = ['<h3 class="alumini-title">Lab Allumini</h3>']
    for label, items in groups.items():
        if not items: continue
        section_title = label or "Other"
        html.append(f'<h4 style="margin:14px 0 8px;color:#040359;font-weight:700;">{escape(section_title)}</h4>')
        html.append('<div class="alumini-grid">')
        for a in items:
            name = escape(a["name"])
            lattes = (a.get("lattes") or "").strip()
            link_html = f'<a href="{escape(lattes)}" target="_blank" rel="noopener">{name}</a>' if lattes else name
            year = (a.get("year") or "").strip()
            meta_html = f'<div class="alumini-meta">{escape(year) if year else ""}</div>' if year else ""
            html.append(f'<div class="alumini-card">{link_html}{meta_html}</div>')
        html.append('</div>')
    
    st.markdown("".join(html), unsafe_allow_html=True)

    
    html_rows.append('</div>')
    st.markdown("".join(html_rows), unsafe_allow_html=True)
    show_help_widget(
    name="Victor Fleck",
    email="Victor.fleck906438@gmail.com",
    role="Undergraduate Pharmacy Student • Web Developer :) ",
    phone="+55 61 991281207",
    label="Problems with the website?"
    )

    



def page_partners():
    """
    Partners (labs/hospitals) + Funding (logo wall)
    - Put partner logos in:   static/partners/
    - Put funding logos in:   static/funding/
    - Partners show name, link, summary, country
    - Funding shows ONLY images (clickable if you add links in FUNDING_LINKS)
    """

    WRAP_MAX = 1180
    BASE_DIR = Path(__file__).parent / "static"
    PARTNER_DIR = BASE_DIR / "partners"
    FUNDING_DIR = BASE_DIR / "funding"

    def _data_uri(p: Path) -> str:
        mime = mimetypes.guess_type(str(p))[0] or "image/png"
        return f"data:{mime};base64,{base64.b64encode(p.read_bytes()).decode()}"

    def _img_src(folder: Path, filename: str | None) -> str | None:
        if not filename:
            return None
        p = folder / filename
        return _data_uri(p) if p.exists() else None

    PARTNERS = [
        {
            "name": "Cancer Institute of the State of São Paulo (ICESP)",
            "website": "https://icesp.org.br/cto/",
            "logo": "icesp.jpg",
            "country": "Brazil",
            "summary": "Translational oncology hub with advanced core facilities, biobank and research groups",
        },
        {
            "name": "Green Nanotechnology Laboratory (NaVe)",
            "website": "https://nanotecnologiaverde.complexonano.com/sobre-nos/",
            "logo": "nave.jpg",
            "country": "UnB – Brazil",
            "summary": "Multi-user lab focused on nanomaterials and nanosystems for health and environment",
        },
        {
            "name": "i.rad.Particles",
            "website": "https://i-rad-particles.com/#home",
            "logo": "rdp.jpg",
            "country": "Brazil",
            "summary": "Deep-tech company in energy and medical biotech; nanoparticle diagnostics and therapies",
        },
        {
            "name": "Laboratory for the Development of Nanostructured Systems (LDNano)",
            "website": "https://www2.ufjf.br/ldnano/",
            "logo": "ldnano.jpg",
            "country": "UFJF – Brazil",
            "summary": "Nanotechnology-based drug delivery to tune pharmacological properties",
        },
        {
            "name": "Biometals Lab Metal Biochemistry, Chemical Biology & Oxidative Stress",
            "website": "https://pnipe.mcti.gov.br/laboratory/23798",
            "logo": "biom.jpg",
            "country": "UFABC – Brazil",
            "summary": "Biochemistry and bioinorganic chemistry of trace metals in biology",
        },
        {
            "name": "Jan Grimm Lab",
            "website": "https://www.mskcc.org/research/ski/labs/jan-grimm",
            "logo": "msk.jpg",
            "country": "MSKCC – United States",
            "summary": "Imaging-driven nanomedicine for cancer: nanoparticles for therapy, Cerenkov/optoacoustic imaging, and PSMA biology",
        },
        {
            "name": "Children's Hospital of Brasília José Alencar",
            "website": "https://www.hcb.org.br/",
            "logo": "hcb.jpg",
            "country": "Brazil",
            "summary": "Public pediatric hospital providing medium- and high-complexity care (oncology, transplants, imaging), integrating care, teaching, and research",
        },
        {
            "name": "Center for Molecular Biotechnology (C-BIOTECH)",
            "website": "https://cbiotech.unb.br/",
            "logo": "cbio.jpg",
            "country": "UnB - Brazil",
            "summary": "The Center provides shared labs for startups and applies molecular biology to develop biopharmaceuticals, chemicals, biofuels, industrial enzymes, and biomaterials, and to valorize biomass.",
        }
        

    ]


    FUNDING_LINKS = {
         "cnpq.jpg": "https://www.gov.br/cnpq/pt-br",
         "fapdf.jpg": "https://www.fap.df.gov.br/",
         "decit.jpg": "https://www.gov.br/saude/pt-br/composicao/sectics/decit",
         "capes.svg": "https://www.gov.br/capes/pt-br",
    }

    # If the folder has files, show what’s in it. Otherwise show nothing.
    def _discover_funding():
        if not FUNDING_DIR.exists():
            return []
        logos = []
        for p in sorted(FUNDING_DIR.iterdir()):
            if p.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp", ".svg"}:
                logos.append({"file": p.name, "src": _data_uri(p), "href": FUNDING_LINKS.get(p.name)})
        return logos

    FUNDERS = _discover_funding()

    # ---------- styles ----------
    st.markdown(
        f"""
        <style>
          :root {{
            --ink: var(--TEXT_DARK, #0f172a);
            --muted:#4b5563;
            --line:#e6e8ee;
            --card:#ffffff;
          }}
          .partners-wrap {{ max-width:{WRAP_MAX}px; margin:0 auto; padding:0 16px 32px; }}
          .page-title {{
            font-size: clamp(26px,3.5vw,38px); font-weight:900; color:var(--ink);
            margin: 18px 0 6px; text-align:left;
          }}
          .section {{ margin: 18px 0 28px; scroll-margin-top:90px; }}
          .section-title {{
            font-size: clamp(18px,2.6vw,24px); font-weight:900; color:var(--ink);
            margin: 4px 0 14px; display:flex; align-items:center; gap:10px;
          }}
          .section-title::after {{
            content:""; flex:1; height:3px; border-radius:2px; background:rgba(63,94,181,.35);
          }}
          /* partner cards */
          .grid-partners {{
            display:grid; grid-template-columns:repeat(auto-fit, minmax(280px,1fr)); gap:18px;
          }}
          .p-card {{
            background: var(--card);
            border: 1px solid var(--line);
            border-radius: 18px;
            padding: 16px;
            display:grid; grid-template-rows:auto auto 1fr auto; gap:10px;
            transition: transform .15s ease, box-shadow .15s ease, border-color .15s ease;
          }}
          .p-card:hover {{
            transform: translateY(-3px);
            box-shadow: 0 12px 26px rgba(2,6,23,.08);
            border-color:#d8dce6;
          }}
          .p-logo {{
            height: 64px; width: 100%; display:flex; align-items:center; justify-content:center;
            border-radius: 12px; background:#fff; border:1px solid var(--line);
            overflow:hidden;
          }}
          .p-logo img {{ max-height:56px; width:auto; object-fit:contain; display:block; }}
          .p-name a {{ color:#0f3bdc; font-weight:900; text-decoration:none; }}
          .p-name a:hover {{ text-decoration:underline; }}
          .p-summary {{ color:var(--muted); line-height:1.45; font-size:.95rem; }}
          .p-badge {{ display:inline-block; padding:4px 10px; border-radius:999px;
                     border:1px solid #cbd5e1; background:#f8fafc; color:#334155; font-size:.8rem; }}
          /* funding wall (logo-only) */
          .grid-funding {{
            display:grid; grid-template-columns:repeat(auto-fit, minmax(140px,1fr)); gap:14px;
          }}
          .fund-tile {{
            background: var(--card);
            border: 1px solid var(--line);
            border-radius: 16px;
            padding: 14px;
            height: 92px;
            display:flex; align-items:center; justify-content:center;
            transition: transform .15s ease, box-shadow .15s ease, filter .15s ease, border-color .15s ease;
            filter: grayscale(22%);
          }}
          .fund-tile:hover {{
            transform: translateY(-3px);
            filter: grayscale(0%);
            box-shadow: 0 10px 22px rgba(2,6,23,.08);
            border-color:#d8dce6;
          }}
          .fund-tile img {{ max-height:64px; width:auto; object-fit:contain; display:block; }}
          @media (max-width:760px){{
            .p-logo {{ height: 60px; }}
            .p-logo img {{ max-height:52px; }}
          }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    # ---------- small render helpers ----------
    def _partner_cards(items: list[dict]) -> str:
        blocks = ['<div class="grid-partners">']
        for it in items:
            name = escape(it.get("name", ""))
            url = (it.get("website") or "").strip()
            summary = escape(it.get("summary", ""))
            country = escape(it.get("country", ""))
            logo_src = _img_src(PARTNER_DIR, it.get("logo"))
            logo_html = (
                f'<div class="p-logo"><img src="{logo_src}" alt="{name} logo"></div>'
                if logo_src else
                f'<div class="p-logo"><span style="font-weight:800;color:#9aa3af">{name[:2]}</span></div>'
            )
            title_html = f'<a href="{escape(url)}" target="_blank" rel="noopener">{name}</a>' if url else name
            summary_html = f'<div class="p-summary">{summary}</div>' if summary else ""
            country_html = f'<div class="p-badge">{country}</div>' if country else ""
            blocks.append(dedent(f"""
              <article class="p-card">
                {logo_html}
                <div class="p-name">{title_html}</div>
                {summary_html}
                {country_html}
              </article>
            """).strip())
        blocks.append("</div>")
        return "".join(blocks)

    def _funding_wall(items: list[dict]) -> str:
        if not items:
            return '<div style="color:#64748b">Add funder logos to <code>static/funding/</code> to show them here.</div>'
        tiles = ['<div class="grid-funding">']
        for it in items:
            img = it["src"]; alt = escape(it["file"])
            if it.get("href"):
                tiles.append(f'<a class="fund-tile" href="{escape(it["href"])}" target="_blank" rel="noopener"><img src="{img}" alt="{alt}"></a>')
            else:
                tiles.append(f'<div class="fund-tile"><img src="{img}" alt="{alt}"></div>')
        tiles.append("</div>")
        return "".join(tiles)

    st.markdown('<div class="partners-wrap">', unsafe_allow_html=True)

    st.markdown('<div class="page-title">Partners</div>', unsafe_allow_html=True)
    st.markdown(_partner_cards(PARTNERS), unsafe_allow_html=True)

    st.markdown('<section class="section" id="funding">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Funding</div>', unsafe_allow_html=True)
    st.markdown(_funding_wall(FUNDERS), unsafe_allow_html=True)
    st.markdown('</section>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    show_help_widget(
    name="Victor Fleck",
    email="Victor.fleck906438@gmail.com",
    role="Undergraduate Pharmacy Student • Web Developer :) ",
    phone="+55 61 991281207",
    label="Problems with the website?"
    )
def page_positions():

    def _find_video(basename: str = "open") -> Path | None:
        exts = [".mp4", ".webm", ".mov", ".m4v"]
        dirs = [
            Path(__file__).parent / "static" / "videos",
            Path(__file__).parent / "static",
            Path(__file__).parent,
        ]
        for d in dirs:
            for ext in exts:
                p = d / f"{basename}{ext}"
                if p.exists():
                    return p
        return None

    def _mime_for(p: Path | None) -> str:
        if not p:
            return "video/mp4"
        s = p.suffix.lower()
        if s in (".mp4", ".m4v"): return "video/mp4"
        if s == ".webm":           return "video/webm"
        if s == ".mov":            return "video/quicktime"
        return "video/mp4"

    def _video_data_uri(p: Path | None) -> str | None:
        try:
            return _data_uri_cached_core(str(p), mime=_mime_for(p)) if p else None
        except Exception:
            return None

    HERO_TITLE = "Open Positions"
    HERO_SUB = ""

    TXT_GRAD = (
        "All year around, the CancerLab is open to students interested in obtaining a "
        "Master and PhD degree through the institutional program of Molecular Pathology and Medical Sciences "
        "of the University of Brasilia. Please contact the respective professors via "
        '<a href="mailto:pittella@unb.br">pittella@unb.br</a> or '
        '<a href="mailto:andreabm@unb.br">andreabm@unb.br</a>.'
    )

    TXT_POSTDOC = (
        "Postdoctoral candidates with experience in genetics, biochemistry, molecular biology or bioinformatics "
        "should send a cover letter with expected starting date, Curriculum vitae and names of THREE references to "
        '<a href="mailto:pittella@unb.br">pittella@unb.br</a>.'
    )

    TXT_UNDERGRAD = (
        "The CancerLab is open to undergraduates with major in biology, biotechnology, pharmaceutical sciences, "
        "medicine and health sciences. Students should contact Prof. Fabio Pittella "
        '(<a href="mailto:pittella@unb.br">pittella@unb.br</a>) or Andrea Motoyama '
        '(<a href="mailto:andreabm@unb.br">andreabm@unb.br</a>).'
    )

    REQS = ["Cover letter", "CV", "Names of three references"]

    st.markdown(
        """
        <style>
          .wrap { max-width:1180px; margin:28px auto 44px; padding:0 16px; }
          .h1   { font-weight:900; font-size:clamp(24px,3.6vw,36px); color:#0f172a; margin:6px 0 6px; }
          .lead { color:#475569; margin:0 0 18px; font-size:1.02rem; }

          /* Anchor index */
          .index { display:flex; flex-wrap:wrap; gap:8px; margin: 8px 0 18px; }
          .chip  { display:inline-block; padding:6px 10px; border:1px solid #e6e8ee; border-radius:999px;
                   text-decoration:none; font-weight:700; font-size:.92rem; color:#0f172a; background:#fff; }
          .chip:hover { border-color:#cfd6e6; }

          /* Hero grid */
          .hero { display:grid; gap:18px; grid-template-columns: 1fr; }
          @media (min-width: 980px){ .hero { grid-template-columns: 1.05fr 0.95fr; } }

          .vcard, .tcard {
            background:#fff; border:1px solid #e6e8ee; border-radius:16px; overflow:hidden;
            box-shadow:0 10px 26px rgba(2,6,23,.06);
          }
          .vcard .media { background:#000; aspect-ratio:16/9; }
          .vcard video, .vcard iframe { width:100%; height:100%; display:block; }
          .tcard { padding:16px 18px; }
          .tcard h3 { margin:0 0 8px; font-size:18px; color:#0f172a; }
          .muted { color:#334155; font-size:.94rem; }
          .btns { display:flex; gap:10px; margin-top:12px; flex-wrap:wrap; }
          .btn  { display:inline-block; padding:8px 12px; border-radius:10px; text-decoration:none; font-weight:800;
                  background:var(--primary,#1351d8); color:#fff; }
          .btn,
          .btn:link,
          .btn:visited,
          .btn:hover,
          .btn:focus,
          .btn:active {
            color: #fff !important;       /* troque aqui se quiser outra cor */
            text-decoration: none !important;
          }
          /* Sections */
          .section { margin:26px 0 0; scroll-margin-top:90px; }
          .stitle   { display:flex; align-items:center; gap:10px;
                      font-weight:900; font-size:clamp(18px,2.6vw,24px); color:#0f172a; margin:0 0 8px; }
          .scard    { background:#fff; border:1px solid #e6e8ee; border-radius:14px; padding:14px 16px; }
          .scard p  { margin:8px 0; color:#1f2937; line-height:1.65; }
          .ul { margin:8px 0 0 18px; }
          .ul li { margin:4px 0; }

          /* Small accent on section cards */
          .scard { position:relative; }
          .scard::before { content:""; position:absolute; left:0; top:0; bottom:0; width:3px;
                           background:linear-gradient(#bfdbfe,#93c5fd); border-top-left-radius:14px; border-bottom-left-radius:14px; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    video_path = _find_video("open")
    video_data = _video_data_uri(video_path)  # prefer embedding to keep layout tight

    st.markdown('<div class="wrap">', unsafe_allow_html=True)
    st.markdown(f'<div class="h1">{HERO_TITLE}</div>', unsafe_allow_html=True)
    st.markdown(f'<p class="lead">{html.escape(HERO_SUB)}</p>', unsafe_allow_html=True)

    st.markdown(
        """
        <nav class="index">
          <a class="chip" href="#grad">Graduate Students</a>
          <a class="chip" href="#postdoc">Postdoctoral Fellows</a>
          <a class="chip" href="#undergrad">Undergraduate Students</a>
          <a class="chip" href="#requirements">Application Requirements</a>
        </nav>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<section class="hero">', unsafe_allow_html=True)

    if video_data:
        st.markdown(
            f"""
            <div class="vcard">
              <div class="media">
                <video controls preload="metadata" src="{video_data}"></video>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    elif video_path:
        left, right = st.columns((1.05, 0.95), gap="large")
        with left:
            st.video(str(video_path))
        with right:
            st.markdown(
                f"""
                <div class="tcard">
                  <h3>How to apply</h3>
                  <p class="muted">Send your materials to the appropriate contact below.</p>
                  <div class="btns">
                    <a class="btn" style="color:#EDE49D !important" href="mailto:pittella@unb.br?subject=Application%20-%20CancerLab">Email Prof. Pittella</a>
                    <a class="btn" style="color:#EDE49D !important" href="mailto:andreabm@unb.br?subject=Application%20-%20CancerLab">Email Prof. Motoyama</a>
                  </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
        st.markdown('</section>', unsafe_allow_html=True)
    else:
        st.markdown(
            """
            <div class="vcard">
              <div class="media" style="display:flex;align-items:center;justify-content:center;color:#e2e8f0;">
                <div style="text-align:center;padding:24px;">
                  <div style="font-weight:800;color:#f8fafc;margin-bottom:8px;">Presentation video not found</div>
                  <div style="color:#cbd5e1;font-size:.95rem;">
                    Put <code>open.mp4</code> (or .webm, .mov) in <code>static/videos/</code>, <code>static/</code> or the app folder.
                  </div>
                </div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown(
            """
            <div class="tcard">
              <h3>How to apply</h3>
              <p class="muted">Send your materials to the appropriate contact below.</p>
              <div class="btns">
                <a class="btn" href="mailto:pittella@unb.br?subject=Application%20-%20CancerLab">Email Prof. Pittella</a>
                <a class="btn" href="mailto:andreabm@unb.br?subject=Application%20-%20CancerLab">Email Prof. Motoyama</a>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown('</section>', unsafe_allow_html=True)

    if video_data:
        st.markdown(
            """
            <div class="tcard">
              <h3>How to apply</h3>
              <p class="muted">Send your materials to the appropriate contact below.</p>
              <div class="btns">
                <a class="btn" href="mailto:pittella@unb.br?subject=Application%20-%20CancerLab">Email Prof. Pittella</a>
                <a class="btn" href="mailto:andreabm@unb.br?subject=Application%20-%20CancerLab">Email Prof. Motoyama</a>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown('</section>', unsafe_allow_html=True)

    # Sections
    st.markdown(
        f"""
        <section class="section" id="grad">
          <div class="stitle">Graduate Students</div>
          <div class="scard"><p>{TXT_GRAD}</p></div>
        </section>

        <section class="section" id="postdoc">
          <div class="stitle">Postdoctoral Fellows</div>
          <div class="scard">
            <p>{TXT_POSTDOC}</p>
            <ul class="ul">
              <li>Expected start date in the cover letter</li>
              <li>Curriculum vitae</li>
              <li>Names of <strong>three</strong> references</li>
            </ul>
          </div>
        </section>
        <section class="section" id="undergrad">
          <div class="stitle">Undergraduate Students</div>
          <div class="scard"><p>{TXT_UNDERGRAD}</p></div>
        </section>
        <section class="section" id="requirements">
          <div class="stitle">Application Requirements</div>
          <div class="scard">
            <ul class="ul">
              {''.join(f'<li>{html.escape(req)}</li>' for req in REQS)}
            </ul>
          </div>
        </section>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("</div>", unsafe_allow_html=True)
    show_help_widget(
    name="Victor Fleck",
    email="Victor.fleck906438@gmail.com",
    role="Undergraduate Pharmacy Student • Web Developer :) ",
    phone="+55 61 991281207",
    label="Problems with the website?"
    )

def page_news():
    import re
    from datetime import datetime
    from urllib.parse import urlparse
    from html import escape as _e
    import streamlit as st

    NEWS_ITEMS = [
        {
            "kind": "press",
            "url": "https://www.fap.df.gov.br/w/cancer-de-mama-pesquisa-da-unb-investiga-novos-caminhos-para-conter-a-metastase",
            "date_iso": "2025-10-02",
            "title": "Câncer de mama: pesquisa da UnB investiga novos caminhos para conter a metástase",
            "dek": "FAP-DF feature about CancerLab/UnB research on strategies to contain breast cancer metastasis.",
            "thumb": None,  # e.g. "static/news/fapdf_oct2025.jpg"
        },
        {
            "kind": "video",
            "url": "https://www.youtube.com/watch?v=xiiLgiTCiZM",
            "date_iso": "2025-10-09",
            "title": "Pink October | Globo TV Interview",
            "dek": "Our team discusses ongoing breast cancer research in a special Globo TV interview",
        },
        {
            "kind": "press",
            "url": "https://noticias.unb.br/artigos-main/8116-outubro-rosa-o-elo-entre-pesquisa-sociedade-e-a-luta-contra-o-cancer-de-mama",
            "date_iso": "2025-10-17",
            "title": "Outubro Rosa: o elo entre pesquisa, sociedade e a luta contra o câncer de mama",
            "dek": "Nossos pesquisadores investigam novas estratégias para frear o câncer de mama e fortalecer a conscientização",
            "thumb": None,  # e.g. "static/news/fapdf_oct2025.jpg"
        },
    ]

    _MONTHS = ["January","February","March","April","May","June",
               "July","August","September","October","November","December"]

    def _fmt_date(iso_str):
        dt = datetime.fromisoformat(iso_str)
        return f"{_MONTHS[dt.month-1]} {dt.day}, {dt.year}"

    def _host(u: str):
        try:
            return urlparse(u).netloc.replace("www.", "")
        except Exception:
            return "link"

    def _trim_url(u: str):
        u = u.replace("https://", "").replace("http://", "")
        return u if len(u) <= 96 else (u[:93] + "…")

    def _youtube_id(u: str):
        """
        Extract a 11-char YouTube ID from common URL formats:
        - https://www.youtube.com/watch?v=ID
        - https://youtu.be/ID
        - https://www.youtube.com/embed/ID
        """
        m = re.search(r"(?:youtube\.com/(?:watch\?v=|embed/)|youtu\.be/)([A-Za-z0-9_-]{11})", u)
        if m:
            return m.group(1)
        m = re.search(r"[?&]v=([A-Za-z0-9_-]{11})", u)
        return m.group(1) if m else None

    st.markdown(
        """
        <style>
          :root{
            --text:#0f172a;      /* cor padrão do site para textos/títulos */
            --text-sub:#0f172a;  /* subtítulo também escuro */
            --muted:#475569;     /* detalhes/links finos */
            --chip-bg:#f3f4f6;   /* chip neutro */
            --chip-bd:#e5e7eb;
            --cyan:#06b6d4;      /* apenas para o botão */
            --cyan-soft: rgba(6,182,212,.08);
          }

          .wrap{max-width:1180px;margin:26px auto;padding:0 16px}
          .hero{text-align:center;padding:10px 0 6px;background:
            radial-gradient(1200px 260px at 50% -60px, rgba(19,81,216,.10), transparent 60%)}
          .title{font-weight:900;letter-spacing:-.01em;font-size:clamp(28px,4.4vw,46px);
            color:var(--text)}
          .sub{color:var(--text-sub);font-weight:750;font-size:clamp(16px,2.1vw,18px)}

          .grid{display:grid;gap:16px;margin-top:14px;grid-template-columns:1fr}
          @media (min-width:980px){.grid{grid-template-columns:1.15fr .85fr}}

          .card{background:#fff;border:1px solid #f0f2f8;border-radius:16px;
            box-shadow:0 10px 22px rgba(2,6,23,.06);overflow:hidden}
          .card--calm{box-shadow:0 6px 16px rgba(2,6,23,.05)}

          .media{position:relative;width:100%;aspect-ratio:16/9;background:#eef2ff}
          .media img,.media iframe{position:absolute;inset:0;width:100%;height:100%;border:0;object-fit:cover}

          .body{padding:14px 16px 16px}
          .h3{margin:6px 0 6px;font-weight:900;color:var(--text);
            font-size:clamp(18px,2.2vw,22px)}
          .dek{color:#1f2937;line-height:1.6}
          .meta{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px}
          .chip{display:inline-flex;align-items:center;gap:8px;background:var(--chip-bg);
            border:1px solid var(--chip-bd);color:var(--text);border-radius:999px;
            padding:6px 10px;font-weight:800;font-size:.85rem}
          .fine{color:var(--muted);font-size:.9rem;margin-top:6px;word-break:break-word}

          /* Botão: fundo branco + texto ciano (único com ciano) */
          .btn{
            display:inline-flex;align-items:center;gap:8px;padding:9px 12px;border-radius:12px;
            font-weight:800;border:1.5px solid var(--cyan);text-decoration:none;
            color:var(--cyan); background:#ffffff;
            box-shadow:0 6px 18px rgba(6,182,212,.12);
          }
          .btn:hover{ background:var(--cyan-soft); }

          /* Vídeo menor, lado a lado */
          .video-card{
            display:grid;
            grid-template-columns:minmax(240px,42%) 1fr; /* video | texto */
            gap:14px; align-items:start;
          }
          @media (max-width:880px){ .video-card{ grid-template-columns:1fr; } }
          .media--sm{ aspect-ratio:16/9; }
          .body--compact{ padding:12px 12px 14px; }
          .h4{ margin:4px 0 6px; font-weight:800;
               font-size:clamp(15px,1.8vw,18px); color:var(--text); }
          .dek--muted{ color:#334155; font-size:.95rem; line-height:1.5; }
          .meta .chip{ font-size:.8rem; padding:5px 9px; }
        </style>
        <div class="wrap">
          <section class="hero" aria-label="News overview">
            <div class="title">News</div>
            <div class="sub">Institutional coverage & media</div>
          </section>
          <div class="grid">
        """,
        unsafe_allow_html=True
    )
    # --------- LEFT: Press (feature) ----------
    press = next((i for i in NEWS_ITEMS if i["kind"] == "press"), None)
    st.markdown('<div>', unsafe_allow_html=True)
    if press:
        date_str = _fmt_date(press["date_iso"])
        host = _host(press["url"])
        thumb = press.get("thumb")
        title = press["title"] or _trim_url(press["url"])
        dek = press.get("dek") or ""

        media_html = f'<div class="media"><img src="{_e(thumb)}" alt="Press image"></div>' if thumb else ""

        st.markdown(
            f"""
            <article class="card" role="article" aria-label="{_e(title)}">
              {media_html}
              <div class="body">
                <div class="meta"><span class="chip">{_e(date_str)}</span><span class="chip">{_e(host)}</span></div>
                <h3 class="h3">{_e(title)}</h3>
                {f'<p class="dek">{_e(dek)}</p>' if dek else ''}
                <p class="fine">{_e(_trim_url(press["url"]))}</p>
                <div style="margin-top:10px">
                  <a class="btn" href="{_e(press['url'])}" target="_blank" rel="noopener">Read article</a>
                </div>
              </div>
            </article>
            """,
            unsafe_allow_html=True
        )
    st.markdown('</div>', unsafe_allow_html=True)
    video = next((i for i in NEWS_ITEMS if i["kind"] == "video"), None)
    st.markdown('<div>', unsafe_allow_html=True)
    if video:
        yt_id = _youtube_id(video["url"])
        date_str = _fmt_date(video["date_iso"])
        embed = f"https://www.youtube-nocookie.com/embed/{yt_id}?rel=0&modestbranding=1&playsinline=1" if yt_id else None
        st.markdown(
            f"""
            <article class="card card--calm video-card" role="article" aria-label="{_e(video['title'])}">
              <div class="media media--sm">
                {f'<iframe src="{_e(embed)}" loading="lazy" allow="accelerometer; autoplay; clipboard-write; encrypted-media; gyroscope; picture-in-picture; web-share" referrerpolicy="strict-origin-when-cross-origin" allowfullscreen title="YouTube video"></iframe>' if embed else ''}
              </div>
              <div class="body body--compact">
                <div class="meta"><span class="chip">{_e(date_str)}</span></div>
                <h3 class="h4">{_e(video['title'])}</h3>
                <p class="dek dek--muted">{_e(video['dek'])}</p>
              </div>
            </article>
            """,
            unsafe_allow_html=True
        )
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown("</div></div>", unsafe_allow_html=True)


show_help_widget(
    name="Victor Fleck",
    email="Victor.fleck906438@gmail.com",
    role="Undergraduate Pharmacy Student • Web Developer :) ",
    phone="+55 61 991281207",
    label="Problems with the website?"
    )




def go_login():
    st.session_state.page = "login"
    st.rerun()


  

APP_BASE_URL = os.getenv("APP_BASE_URL", "https://cancerlab.up.railway.app")
SMTP_HOST    = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT    = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER    = (os.getenv("SMTP_USER") or "").strip()
SMTP_PASS    = (os.getenv("SMTP_PASS") or "").replace(" ", "").strip()
FROM_EMAIL   = SMTP_USER



st.markdown("""
<style>
/* Remove Streamlit's built-in top chrome */
[data-testid="stHeader"],
[data-testid="stToolbar"],
[data-testid="stDecoration"] { display: none !important; }
/* Pull content up after removing the header */
main .block-container { padding-top: 0 !important; }
html, body { margin-top: 0 !important; }
</style>
""", unsafe_allow_html=True)




usuarios_table = Table(
    'usuarios', metadata,
    Column('id', Integer, primary_key=True),
    Column('CPF', String, unique=True),
    Column('nome', String),
    Column('email', String),
    Column('senha', String),   # hash em string
    Column('perfil', String),
)

password_reset_tokens = Table(
    'password_reset_tokens', metadata,
    Column('id', Integer, primary_key=True),
    Column('user_id', Integer, nullable=False),
    Column('token', String, unique=True, nullable=False),
    Column('expires_at', DateTime, nullable=False),
    Column('used', Integer, default=0, nullable=False),
)


def bootstrap_db_once_if_needed():
    alvo = st.session_state.get("page") or _get_param("page", "home")
    if alvo in {"login", "criar_conta", "reset_request", "reset"} and not st.session_state.get("_db_bootstrapped"):
        metadata.create_all(engine, checkfirst=True)
        st.session_state["_db_bootstrapped"] = True

bootstrap_db_once_if_needed()



def salvar_usuario(usuario):
    try:
        with engine.begin() as conn:  # begin() garante commit automático
            print("Tentando salvar:", usuario)
            conn.execute(usuarios_table.insert().values(**usuario))
            print("Salvou!")
    except Exception as e:
        print("❌ Erro ao salvar usuário:", e)

def carregar_usuarios():
    with engine.connect() as conn:
        result = conn.execute(select(usuarios_table))
        return [dict(row._mapping) for row in result]
# ta em teste essa porra

def get_user_by_email(email: str):
    with engine.connect() as conn:
        res = conn.execute(select(usuarios_table).where(usuarios_table.c.email == email)).fetchone()
        return dict(res._mapping) if res else None

def create_reset_token(user_id: int, ttl_minutes: int = 60) -> str:
    token = secrets.token_urlsafe(32)
    exp = datetime.now() + timedelta(minutes=ttl_minutes)
    with engine.begin() as conn:
        conn.execute(password_reset_tokens.insert().values(
            user_id=user_id, token=token, expires_at=exp, used=0
        ))
    return token

def validate_reset_token(token: str):
    with engine.connect() as conn:
        q = select(password_reset_tokens).where(password_reset_tokens.c.token == token)
        row = conn.execute(q).fetchone()
        if not row:
            return None
        rec = dict(row._mapping)
        if rec["used"] == 1:
            return None
        if rec["expires_at"] < datetime.now():
            return None
        return rec  

def mark_token_used(token: str):
    with engine.begin() as conn:
        conn.execute(
            password_reset_tokens.update()
            .where(password_reset_tokens.c.token == token)
            .values(used=1)
        )

def update_user_password(user_id: int, new_password: str):
    hashed = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
    with engine.begin() as conn:
        conn.execute(
            usuarios_table.update()
            .where(usuarios_table.c.id == user_id)
            .values(senha=hashed)
        )

def send_password_reset_email(to_email: str, reset_url: str) -> bool:
    msg = EmailMessage()
    msg["Subject"] = "Recuperação de senha - CancerLab"
    msg["From"] = FROM_EMAIL
    msg["To"] = to_email
    msg.set_content(
        f"Olá!\n\nClique para redefinir sua senha (válido por 60 minutos):\n{reset_url}\n\n"
        "Se não foi você, ignore este e-mail."
    )
    try:
        # 587 + STARTTLS
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as s:
            s.ehlo(); s.starttls(); s.ehlo()
            s.login(SMTP_USER, SMTP_PASS)
            s.send_message(msg)
        return True
    except smtplib.SMTPAuthenticationError:
        # fallback 465 + SSL
        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=20) as s:
                s.login(SMTP_USER, SMTP_PASS)
                s.send_message(msg)
            return True
        except Exception:
            return False
    except Exception:
        return False

# --- Verificar se matrícula já existe ---
def usuario_existe(CPF):
    with engine.connect() as conn:
        result = conn.execute(select(usuarios_table).where(usuarios_table.c.CPF == CPF))
        return result.fetchone() is not None

# --- Autenticar usuário ---
def autentica(CPF, senha):
    with engine.connect() as conn:
        result = conn.execute(select(usuarios_table).where(usuarios_table.c.CPF == CPF))
        user = result.fetchone()
        if user:
            user_dict = dict(user._mapping)
            if bcrypt.checkpw(senha.encode(), user_dict['senha'].encode()):
                return user_dict
    return None

# --- Função auxiliar de data ---
def add_months(data, n):
    mes = data.month - 1 + n
    ano = data.year + mes // 12
    mes = mes % 12 + 1
    dia = min(data.day, [31,
        29 if ano % 4 == 0 and not ano % 100 == 0 or ano % 400 == 0 else 28,
        31,30,31,30,31,31,30,31,30,31][mes-1])
    return date(ano, mes, dia)

INTERVALOS_BXL_MESES = {
    "BxL 3": 6,
    "BxL 4": 12,
    "BxL 5": 18,
    "BxL 6": 24,
    "BxL 7": 30
}




st.markdown("""
<style>
/* ===== Paleta ===== */
:root{
  --bg: #F3F9FB;           /* fundo da página */
  --text: #111827;         /* texto padrão */
  --muted: #1F2A37;        /* rótulos/legendas */
  --primary: #173975;      /* azul institucional */
  --primary-hover: #0A74B7;
  --input-bg: #EDF2F7;     /* fundo dos inputs/select */
  --input-text: #111827;   /* texto digitado */
  --placeholder: #6B7280;  /* placeholder */
  --border: #C5D1DE;       /* borda dos inputs */
  --alert-info-bg: #DBECFA;     /* fundo do st.info */
  --alert-info-text: #0B3E61;   /* texto do st.info */
  --alert-success-bg: #E7F4EA;  /* fundo do st.success */
  --alert-success-text:#064715; /* texto do st.success */
  /* Forçar cor dos headings (títulos) do Streamlit */
    h1, h2, h3, h4, h5, h6,
    [data-testid="stMarkdownContainer"] h1,
    [data-testid="stMarkdownContainer"] h2,
    .block-container h1 {
      color: var(--text) !important;   /* usa #111827 */
      text-shadow: none !important;    /* remove qualquer sombra clara */
}

/* ===== Fundo e texto ===== */
body, .stApp { 
  background: var(--bg) !important; 
  color: var(--text) !important; 
}

/* ===== Cabeçalho/título da página ===== */
.header-lab-text{
  color: var(--primary) !important;
  text-shadow: 0 2px 8px rgba(0,0,0,0.10);
}

/* ===== Botões ===== */
.stButton > button {
  background: var(--primary);
  color: white;
  border-radius: 1rem;
  padding: 0.75rem 2rem;
  border: none;
  font-size: 1.05rem;
  transition: .2s;
  margin-top: 0.5rem;
  margin-bottom: 1.5rem;
}
.stButton > button:hover { background: var(--primary-hover); }

/* ===== Inputs de texto/senha ===== */
.stTextInput > div > div > input,
.stPassword   > div > div > input{
  background: var(--input-bg) !important; color: var(--input-text) !important;
  border: 1px solid var(--border); border-radius: .6rem; padding: .55rem .75rem;
}
.stTextInput input::placeholder, .stPassword input::placeholder{
  color: var(--placeholder) !important; opacity: 1 !important;
}

/* ===== Labels/legendas ===== */
label, .stMarkdown, .stCaption, .stRadio, .stSelectbox label, .stTextInput label{
  color: var(--muted) !important;
}

/* ===== Selectbox (BaseWeb) ===== */
div[data-baseweb="select"] > div{
  background: var(--input-bg) !important; color: var(--input-text) !important;
  border: 1px solid var(--border); border-radius: .6rem; min-height: 42px;
}
div[data-baseweb="select"] *{ color: var(--input-text) !important; }
div[data-baseweb="select"] svg{ opacity: .85; }

/* Dropdown de opções do select */
div[role="listbox"]{
  background: #fff !important; border: 1px solid var(--border);
}
div[role="listbox"] *{ color: var(--text) !important; }

/* Estados desabilitados */
.stTextInput input:disabled{ color:#6B7280 !important; }
div[data-baseweb="select"][aria-disabled="true"] > div{
  opacity:.8; border-style: dashed;
}

/* ===== Alertas (success/info/warn/error) com texto legível ===== */
.stAlert{ border: none !important; }
.stAlert div[role="alert"]{ box-shadow:none; }
.stAlert[data-baseweb] { box-shadow:none; }

/* Streamlit 1.3x – alvo genérico do conteúdo do alerta */
div[role="alert"]{
  border-radius: .6rem; padding: .8rem 1rem;
}

/* Sucesso */
div[role="alert"].stAlert-success,
div[role="alert"]:has(.markdown-success){
  background: var(--alert-success-bg) !important; color: var(--alert-success-text) !important;
}

/* Info */
div[role="alert"].stAlert-info,
div[role="alert"]:has(.markdown-info){
  background: var(--alert-info-bg) !important; color: var(--alert-info-text) !important;
}

/* Fallback para o componente de notificação recente */
div[data-testid="stNotificationContent"]{
  background: var(--alert-info-bg) !important; color: var(--alert-info-text) !important;
}
.stButton > button,
.stDownloadButton > button,
.stFormSubmitButton > button {
  background: var(--primary) !important;
  color: white !important;
  border-radius: 1rem !important;
  padding: 0.75rem 2rem !important;
  border: none !important;
  font-size: 1.05rem !important;
  transition: .2s !important;
  margin-top: 0.5rem !important;
  margin-bottom: 1.5rem !important;
}
.stButton > button:hover,
.stDownloadButton > button:hover,
.stFormSubmitButton > button:hover {
  background: var(--primary-hover) !important;
}
/* Força estilos legíveis em TODOS os alerts nativos do Streamlit */
.stAlert, .stAlert > div[role="alert"] {
  box-shadow: none !important;
  border: none !important;
  border-radius: .6rem !important;
  padding: .8rem 1rem !important;
}

/* Info */
.stAlert:has([data-testid="stIconInfo"]),
.stAlert[data-testid="stNotification"] {
  background: var(--alert-info-bg) !important;   /* #DBECFA */
  color: var(--alert-info-text) !important;      /* #0B3E61 */
}

/* Success */
.stAlert:has([data-testid="stIconSuccess"]) {
  background: var(--alert-success-bg) !important;   /* #E7F4EA */
  color: var(--alert-success-text) !important;      /* #064715 */
}

/* Warning */
.stAlert:has([data-testid="stIconWarning"]) {
  background: #FFF7E6 !important;
  color: #7A4D00 !important;
}

/* Error */
.stAlert:has([data-testid="stIconError"]) {
  background: #FDE8E8 !important;
  color: #7F1D1D !important;
}

/* Conteúdo interno do alerta (markdown/texto) */
.stAlert [data-testid="stMarkdownContainer"],
.stAlert p, .stAlert span, .stAlert div {
  color: inherit !important;
}
</style>
""", unsafe_allow_html=True)





# === Top navbar + public routing (executa antes do app interno) ===
active_nav = _get_param("page", "home")
if isinstance(active_nav, (list, tuple)):   # <- defensivo, caso venha lista
    active_nav = active_nav[0]
show_lab_header(
    image_relpath="static/imgfront",
    address_lines=[
        "Laboratory of Molecular Pathology of Cancer",
        "Faculty of Health Sciences",
        "University of Brasília",
        "Brasília – DF – Brazil",
    ],
)
render_navbar(active_nav)





is_internal = st.session_state.get("usuario_logado") is not None or \
              st.session_state.get("page") in {"index", "biopsia", "clinicos"}

if not is_internal:
    if active_nav == "login":
        st.session_state.page = "login"
    elif active_nav in {"home", "research", "publications", "people", "partners", "positions", "equipment", "news"}:  # + two slugs
        {
         "home": page_home,
         "research": page_research,
         "publications": page_publications,
         "equipment": page_equipment,
         "people": page_people,
         "partners": page_partners,
         "positions": page_positions,
         "news": page_news, # + map       # + map
        }[active_nav]()
        st.stop()






if "page" not in st.session_state:
    st.session_state.page = "login"

if "usuario_logado" not in st.session_state:
    st.session_state.usuario_logado = None

try:
    params = st.query_params  
except Exception:
    params = st.experimental_get_query_params()  # fallback

if "page" in params and isinstance(params["page"], (list, tuple)):
    qp_page = params["page"][0]
else:
    qp_page = params.get("page")

if qp_page in {"reset_request", "reset"}:
    st.session_state.page = qp_page

if "reset_token" not in st.session_state:
    st.session_state.reset_token = None

qp_token = None
if "token" in params and isinstance(params["token"], (list, tuple)):
    qp_token = params["token"][0]
else:
    qp_token = params.get("token")

if qp_token:
    st.session_state.reset_token = qp_token

# ---- PÁGINA DE LOGIN ----
if st.session_state.page == "login":
    st.title("Login")

    with st.form("login_form"):
        CPF = st.text_input("CPF",placeholder="Digite o CPF sem pontos ou traços")
        senha = st.text_input("Senha",type="password",placeholder="Digite sua senha") 
        submitted = st.form_submit_button("Entrar")

    if submitted:
        try:
            user = autentica(CPF, senha)  # bcrypt check inside
            if user:
                st.session_state.usuario_logado = user
                st.session_state.page = "index"
                _set_param(page="index") 
                st.rerun()
            else:
                st.error("Usuário ou senha inválidos.")
        except Exception as e:
            st.error(f"Falha ao autenticar: {e}")

    st.markdown(
        """
        <div style="margin-top:-0.5rem;margin-bottom:0.5rem;">
          <a href="?page=reset_request" style="font-size:0.85rem; color:#00385C; text-decoration:none;">
            Esqueci minha senha
          </a>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown("---")
    if st.button("Criar conta"):
        st.session_state.page = "criar_conta"
        _set_param(page="criar_conta")
        st.rerun()




# CRIAÇÃO DE CONTA ----
elif st.session_state.page == "criar_conta":
    st.title("Criar nova conta")
    nome = st.text_input("Nome completo")
    CPF = st.text_input("CPF")
    email = st.text_input("Email")
    senha = st.text_input("Senha", type="password")
    perfil = st.radio("Tipo de usuário", ["aluno", "servidor"])

    if st.button("Registrar"):
        try:
            if usuario_existe(CPF):
                st.error("Matrícula já cadastrada!")
            elif not (nome and CPF and email and senha):
                st.error("Preencha todos os campos!")
            else:
                hashed = bcrypt.hashpw(senha.encode(), bcrypt.gensalt()).decode()
                usuario = {"CPF": CPF, "nome": nome, "email": email, "senha": hashed, "perfil": perfil}
                salvar_usuario(usuario)
                st.success("Conta criada com sucesso! Faça login.")
                st.session_state.page = "login"
                _set_param(page="login")   # >>> URL coerente para o login <<<
                st.rerun()
        except Exception as e:
            st.error(f"Erro ao registrar: {e}")

    if st.button("Voltar para login"):
        st.session_state.page = "login"
        _set_param(page="login")
        st.rerun()

elif st.session_state.page == "reset":
    st.title("Definir nova senha")
    token = st.session_state.reset_token

    if not token:
        st.error("Link inválido ou ausente. Solicite novamente a recuperação de senha.")
        if st.button("Voltar para solicitar link"):
            st.session_state.page = "reset_request"
            st.rerun()
    else:
        rec = validate_reset_token(token)
        if not rec:
            st.error("Este link é inválido, já foi utilizado ou expirou.")
            if st.button("Solicitar novo link"):
                st.session_state.page = "reset_request"
                st.rerun()
        else:
            nova = st.text_input("Nova senha", type="password")
            conf = st.text_input("Confirmar nova senha", type="password")
            if st.button("Salvar nova senha"):
                if not nova or not conf:
                    st.warning("Preencha a nova senha e a confirmação.")
                elif nova != conf:
                    st.error("As senhas não conferem.")
                elif len(nova) < 6:
                    st.warning("Use uma senha com pelo menos 6 caracteres.")
                else:
                    update_user_password(rec["user_id"], nova)
                    mark_token_used(token)
                    st.success("Senha alterada com sucesso! Faça login.")
                    # Limpa token em memória
                    st.session_state.reset_token = None
                    if st.button("Ir para o login"):
                        st.session_state.page = "login"
                        st.rerun()

# PÁGINA INICIAL
elif st.session_state.page == "index":
    user = st.session_state.usuario_logado
    st.title(f"Bem-vindo, {user['nome']} ({user['perfil']})")
    st.write(f"Email: {user['email']}")
    import streamlit as st
    import pandas as pd
    import numpy as np
    import plotly.express as px
    import plotly.graph_objects as go
    import unicodedata
    import re
    from typing import List, Tuple, Optional

    try:
        st.set_page_config(page_title="CancerLab – Analytics", layout="wide")
    except Exception:
        pass

    DARK_QUAL = ["#0f172a","#1d4ed8","#0ea5e9","#059669","#7c3aed","#dc2626","#d97706","#374151","#065f46","#7c2d12"]
    DARK_CONT = "Viridis"  # contínuo legível em fundo branco
    PLOTLY_TEMPLATE = "plotly_white"

    def _apply_fig_base(fig, h=360):
        fig.update_layout(
            template=PLOTLY_TEMPLATE,
            height=h,
            margin=dict(t=24, l=8, r=8, b=16),
            legend_title_text=None,
            colorway=DARK_QUAL,
        )
        return fig

    st.markdown("""
    <style>
    /* Tabs com contraste */
    .stTabs [role="tab"]{
      color:#111827 !important; background:rgba(2,6,23,.04) !important;
      border:1px solid rgba(2,6,23,.08) !important; border-bottom-color:transparent !important;
      font-weight:700; letter-spacing:.2px;
    }
    .stTabs [role="tab"][aria-selected="true"]{
      color:#111827 !important; background:#ffffff !important; border-bottom:2px solid #2563eb !important;
    }
    /* Cartões */
    .block-card{ background:#fff; border:1px solid rgba(2,6,23,.08); border-radius:14px; padding:14px;
                 box-shadow:0 10px 24px rgba(2,6,23,.06); }
    /* Botões: texto escuro no claro */
    .stButton > button, .stDownloadButton > button {
      background:#ffffff !important; color:#111827 !important; border:1px solid #e5e7eb !important;
    }
    /* Dataframe cabeçalho legível */
    [data-baseweb="table"] div { color:#111827; }
    </style>
    """, unsafe_allow_html=True)

    # ---------------- Helpers ----------------
    @st.cache_data(show_spinner=False)
    def _load_excel_cached(path: str) -> pd.DataFrame:
        try:
            df0 = pd.read_excel(path)
        except Exception as e:
            st.warning(f"Não foi possível ler {path}: {e}")
            return pd.DataFrame()
        df0.columns = [str(c).strip() for c in df0.columns]
        return df0

    @st.cache_data(show_spinner=False)
    def _read_file_bytes(path: str) -> bytes:
        try:
            with open(path, "rb") as f:
                return f.read()
        except Exception:
            return b""

    def _norm_txt(x: str) -> str:
        if x is None:
            return ""
        x = str(x).strip().lower()
        return unicodedata.normalize("NFKD", x).encode("ascii","ignore").decode("ascii")

    def _is_date(s: pd.Series) -> bool:
        return pd.api.types.is_datetime64_any_dtype(s)

    def _is_num(s: pd.Series) -> bool:
        return pd.api.types.is_numeric_dtype(s)

    def _safe_numeric(s: pd.Series) -> pd.Series:
        return pd.to_numeric(s, errors="coerce").astype(float).replace([np.inf, -np.inf], np.nan)

    def _pct_missing(df_: pd.DataFrame) -> float:
        return float(df_.isna().mean().mean() * 100)

    def _top_counts(s: pd.Series, top: int = 20) -> pd.DataFrame:
        ss = s.astype(str).replace({"nan":"(nulo)","None":"(nulo)"}).fillna("(nulo)")
        f = ss.value_counts(dropna=False).rename("contagem").head(top)
        out = f.to_frame().reset_index(names="valor")
        out["%"] = (out["contagem"] / max(1, len(s)) * 100).round(1)
        return out

    def _by_exact(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
        m = {str(c).strip().lower(): c for c in df.columns}
        for c in candidates:
            if c.lower() in m:
                return m[c.lower()]
        return None

    def _limit_bins(lo: float, hi: float, step: Optional[float], max_bins: int = 200) -> Tuple[np.ndarray, float]:
        if not np.isfinite(lo) or not np.isfinite(hi) or hi <= lo:
            hi = lo + 1.0
        width = float(hi - lo)
        if step is None or step <= 0:
            step = max(width / 10.0, 1.0)
        n_bins = int(np.ceil(width / step))
        if n_bins > max_bins:
            step = max(width / max_bins, np.finfo(float).eps)
            n_bins = int(np.ceil(width / step))
        if n_bins < 1:
            n_bins = 1
        edges = np.linspace(lo, lo + n_bins * step, n_bins + 1)
        edges[0] = np.floor(edges[0])
        edges[-1] = np.ceil(hi)
        edges = np.unique(edges)
        if edges.size < 2:
            edges = np.array([lo, lo + (step if step > 0 else 1.0)])
        return edges, float(step)

    def _make_intervals(s: pd.Series, step: Optional[float]):
        sn = _safe_numeric(s).dropna()
        if sn.empty:
            return None, None
        lo = float(np.nanpercentile(sn, 1))
        hi = float(np.nanpercentile(sn, 99))
        edges, step2 = _limit_bins(lo, hi, step, max_bins=200)
        labels = []
        width = edges[-1] - edges[0]
        fine = width < 10
        for i in range(len(edges) - 1):
            a, b = edges[i], edges[i + 1]
            labels.append(f"{a:.2f}–{b:.2f}" if fine else f"{int(np.floor(a))}–{int(np.ceil(b))}")
        binned = pd.cut(_safe_numeric(s), bins=edges, labels=labels, include_lowest=True, duplicates="drop")
        return binned, labels

    # ---------------- Navegação (router) ----------------
    def _goto(page: str):
        st.session_state.page = page
        st.rerun()





    # ---------------- BOTÕES PARA IR ÀS PÁGINAS JÁ EXISTENTES + INPUT DE VISUALIZAÇÃO ----------------
    st.divider()
    
    # CSS do bloco (não mexe no restante do seu tema)
    st.markdown("""
    <style>
      /* wrapper do bloco */
      .index-wrap { margin-top: 6px; }
    
      /* cards */
      .cl-card{
        background:#fff;
        border:1px solid rgba(2,6,23,.08);
        border-radius:16px;
        padding:16px 16px 14px 16px;
        box-shadow:0 10px 24px rgba(2,6,23,.06);
      }
      .cl-title{
        font-weight:800;
        letter-spacing:.2px;
        margin:0 0 4px 0;
        color:#0f172a;
      }
      .cl-muted{
        color: rgba(17,24,39,.70);
        font-size:.92rem;
        margin:0 0 12px 0;
      }
      .cl-divider{
        height:1px;
        background: rgba(2,6,23,.06);
        margin: 10px 0 14px 0;
      }
    
      /* Botões: estilo robusto e consistente */
      .stButton > button {
        border-radius: 12px !important;
        border: 1px solid rgba(2,6,23,.10) !important;
        padding: 0.70rem 0.9rem !important;
        font-weight: 800 !important;
      }
    
      /* Radio horizontal (pílulas) – melhora o visual do input */
      div[role="radiogroup"] > label {
        background: rgba(2,6,23,.04);
        border: 1px solid rgba(2,6,23,.08);
        padding: 8px 10px;
        border-radius: 999px;
        margin-right: 8px;
      }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="index-wrap">', unsafe_allow_html=True)
    st.subheader("Acessos e visualização")
    
    # ---------------- CARD 1: Acessos rápidos (SOMENTE 2 BOTÕES) ----------------
    with st.container():
        st.markdown('<p class="cl-title">Acessos rápidos</p>', unsafe_allow_html=True)
        st.markdown('<p class="cl-muted">Abra as telas de cadastro/edição (planilhas) do sistema.</p>', unsafe_allow_html=True)
    
        c1, c2 = st.columns(2, gap="large")
    
        with c1:
            if st.button("Biópsia Líquida (Entrada/Acompanhamento)", use_container_width=True, key="go_biopsia_top"):
                _goto("biopsia")
    
        with c2:
            if st.button("Dados Clínicos (Cadastro/Edição)", use_container_width=True, key="go_clinicos_top"):
                _goto("clinicos")
    
        st.markdown("</div>", unsafe_allow_html=True)
    
    st.markdown("")
    
    # ---------------- CARD 2: Inputs de visualização (SEM BOTÕES) ----------------
    with st.container():
        st.markdown('<p class="cl-title">Selecione o painel</p>', unsafe_allow_html=True)
        st.markdown('<p class="cl-muted">Escolha a fonte de dados e o tipo de visualização para exibir nesta página.</p>', unsafe_allow_html=True)
    
        # Linha 1: Fonte dos dados (mantém a variável "fonte_dados" para o resto do código)
        fonte_dados = st.selectbox(
            "Fonte dos dados",
            [
                "Dados Clínicos CCR",
                "Biópsia Líquida — Agenda (Entrada/Acompanhamento)",
            ],
            index=0,
            key="index_fonte_dados",
        )
    
        st.markdown('<div class="cl-divider"></div>', unsafe_allow_html=True)

        # Defaults (sem UI) — mantém compatibilidade com o resto do dashboard
        tipo_viz = "Resumo"
        top_n = 15
        show_missing = True
        modo = "Paper-ready"

    
        st.markdown("</div>", unsafe_allow_html=True)
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # A partir daqui, seu dashboard usa:
    # - fonte_dados (string)
    # - tipo_viz, top_n, show_missing, modo (inputs)
    #
    # Exemplo de roteamento (opcional):
    # if fonte_dados == "Dados Clínicos CCR":
    #     render_clinicos(tipo_viz=tipo_viz, top_n=top_n, show_missing=show_missing, modo=modo)
    # else:
    #     render_biopsia(tipo_viz=tipo_viz, top_n=top_n, show_missing=show_missing, modo=modo)
    

    # =====================================================================
    # 1) PAINEL CLÍNICO (CCR BxL) — 3 abas (Visão geral / Distribuições / Coorte)
    # =====================================================================
    def painel_clinico_ccr_bxl(df_in: Optional[pd.DataFrame] = None):
        from datetime import datetime, date  # <-- garante datetime/date aqui

        APP_TITLE = "Painel Clínico"

        # -------- Utils base (locais) --------
        def _slug(s: str) -> str:
            if s is None:
                return ""
            s = re.sub(r"\s+", " ", str(s)).strip()
            s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
            return s

        def _norm(s: str) -> str:
            return _slug(s).lower()

        def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
            df = df.copy()
            df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
            drop_me = [c for c in df.columns if str(c).lower().startswith("unnamed") and df[c].isna().all()]
            if drop_me:
                df = df.drop(columns=drop_me)
            return df

        def _chart_header(title: str, n_patients: int, subtitle: Optional[str] = None):
            n_txt = f"{n_patients:,}".replace(",", ".")
            if subtitle:
                st.markdown(f"**{title}**  \n{subtitle}  \n**N = {n_txt}**")
            else:
                st.markdown(f"**{title}**  \n**N = {n_txt}**")

        def _apply_plotly_theme(fig: go.Figure, height: int = 420) -> go.Figure:
            fig.update_layout(
                template="plotly_white",
                height=height,
                margin=dict(l=12, r=12, t=45, b=12),
                font=dict(family="Inter, Arial, sans-serif", size=13, color="#111827"),
                hovermode="closest",
                legend=dict(
                    orientation="h",
                    yanchor="bottom", y=1.02,
                    xanchor="left", x=0.0,
                    bgcolor="rgba(255,255,255,0.85)"
                ),
                paper_bgcolor="white",
                plot_bgcolor="white",
            )
            fig.update_xaxes(showgrid=True, gridcolor="rgba(17,24,39,0.08)", zeroline=False, ticks="outside")
            fig.update_yaxes(showgrid=True, gridcolor="rgba(17,24,39,0.08)", zeroline=False, ticks="outside")
            return fig

        def _to_sim_nao(x) -> str:
            if pd.isna(x):
                return "Não"
            s = str(x).strip().lower()
            if s in ("", "nan", "na", "n/a", "não", "nao", "n", "no", "0", "-", "--"):
                return "Não"
            if s in ("sim", "s", "yes", "y", "1", "true", "t"):
                return "Sim"
            return "Não"

        def _to_etilismo(x) -> str:
            if pd.isna(x):
                return "Não"
            s = str(x).strip().lower()
            if s in ("", "nan", "na", "n/a", "não", "nao", "n", "no", "0", "-", "--"):
                return "Não"
            if "social" in s:
                return "Social"
            if s in ("sim", "s", "yes", "y", "1", "true", "t"):
                return "Sim"
            return "Não"

        def _is_date_like_series(s: pd.Series) -> bool:
            if pd.api.types.is_datetime64_any_dtype(s):
                return True
            name = _norm(getattr(s, "name", ""))
            if "data" in name:
                return True
            if s.dropna().empty:
                return False
            sample = s.dropna().astype(str).head(60)
            dt = pd.to_datetime(sample, errors="coerce", dayfirst=True)
            return dt.notna().mean() >= 0.60

        def _format_date_only(x) -> str:
            if pd.isna(x):
                return "Não"

            if isinstance(x, (datetime, date)):
                d = x.date() if isinstance(x, datetime) else x
                if d.year < 1900 or d.year > 2200:
                    return "Não"
                return d.strftime("%d/%m/%Y")

            s = str(x).strip()
            if not s or s.lower() in ("nan", "na", "n/a"):
                return "Não"

            try:
                ts = pd.to_datetime(s, errors="coerce", dayfirst=True)
                if pd.notna(ts):
                    if ts.year < 1900 or ts.year > 2200:
                        return "Não"
                    return ts.strftime("%d/%m/%Y")
            except Exception:
                pass

            m = re.search(r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b", s)
            if m:
                try:
                    ts2 = pd.to_datetime(m.group(1), errors="coerce", dayfirst=True)
                    if pd.notna(ts2):
                        if ts2.year < 1900 or ts2.year > 2200:
                            return "Não"
                        return ts2.strftime("%d/%m/%Y")
                except Exception:
                    pass

            return s

        # Estadiamento
        _ROMAN_RE = re.compile(r"\b(IV|III|II|I|0)\s*([ABC])?\b", flags=re.IGNORECASE)

        def _split_multi_values(text: str) -> List[str]:
            text = str(text).replace("→", "->")
            parts = re.split(r"(?:->|>|;|\||/|,|\n|\r|\t)+", text)
            return [p.strip() for p in parts if p and p.strip()]

        def _normalize_stage_token(token: str) -> Optional[str]:
            if token is None or (isinstance(token, float) and np.isnan(token)):
                return None
            t = str(token).strip()
            if not t or t.lower() in ("na", "nan", "-"):
                return None

            m = _ROMAN_RE.search(t)
            if not m:
                mm = re.findall(r"(IV|III|II|I|0)([ABC])?", t.upper())
                if not mm:
                    return None
                r, l = mm[-1]
                return f"{r}{l or ''}".upper()

            roman = (m.group(1) or "").upper()
            letter = (m.group(2) or "").upper()
            return f"{roman}{letter}".upper()

        def _clean_stage_field(x) -> str:
            if pd.isna(x):
                return "Não"
            raw = str(x).strip()
            if not raw or raw.lower() in ("na", "nan", "-"):
                return "Não"

            parts = _split_multi_values(raw) or [raw]
            tokens = []
            for p in parts:
                tok = _normalize_stage_token(p)
                if tok:
                    tokens.append(tok)
                else:
                    mm = re.findall(r"(IV|III|II|I|0)([ABC])?", str(p).upper())
                    for r, l in mm:
                        tokens.append(f"{r}{l or ''}".upper())

            uniq = []
            for t in tokens:
                if t not in uniq:
                    uniq.append(t)

            if not uniq:
                return "Não"

            out = uniq[-1]
            if len(uniq) > 1:
                out = f"{out}*"
            return out

        def _stage_sort_key(stage: str) -> Tuple[int, int, int, str]:
            if stage is None:
                return (99, 99, 1, "")
            s = str(stage).strip().upper()
            if s in ("NÃO", "NAO", "NA", "NAN", "", "—", "-"):
                return (99, 99, 1, "NAO")
            starred = 1 if s.endswith("*") else 0
            s0 = s[:-1] if starred else s
            m = re.match(r"^(0|IV|III|II|I)([ABC])?$", s0)
            if not m:
                return (98, 98, starred, s0)
            roman, letter = m.group(1), m.group(2) or ""
            roman_map = {"0": 0, "I": 1, "II": 2, "III": 3, "IV": 4}
            letter_map = {"": 0, "A": 1, "B": 2, "C": 3}
            return (roman_map.get(roman, 98), letter_map.get(letter, 98), starred, s0)

        # Numérico robusto
        _DATE_DMY = re.compile(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b")
        _DATE_YMD = re.compile(r"\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b")
        _TIME = re.compile(r"\b\d{1,2}:\d{2}(:\d{2})?\b")

        def _extract_numeric_sequence(x) -> List[float]:
            if pd.isna(x):
                return []
            if isinstance(x, (int, float, np.integer, np.floating)):
                if np.isfinite(x):
                    return [float(x)]
                return []

            s = str(x).strip()
            if not s or s.lower() in ("nan", "na", "-"):
                return []

            s = _DATE_DMY.sub(" ", s)
            s = _DATE_YMD.sub(" ", s)
            s = _TIME.sub(" ", s)

            s = s.replace("→", "->").replace(">", " -> ").replace("<", " -> ").replace("->", " -> ")
            s = s.replace("%", " ").replace(",", ".")
            s = re.sub(r"[^0-9\.\-\+\s]", " ", s)
            s = re.sub(r"\s+", " ", s).strip()

            nums = re.findall(r"[-+]?\d+(?:\.\d+)?", s)
            out = []
            for n in nums:
                try:
                    v = float(n)
                    if np.isfinite(v):
                        out.append(v)
                except Exception:
                    pass
            return out

        def _robust_center(vals: List[float]) -> float:
            if not vals:
                return np.nan
            v = np.array(sorted(vals), dtype=float)
            if len(v) == 1:
                return float(v[0])
            if len(v) == 2:
                return float((v[0] + v[1]) / 2)
            if len(v) <= 4:
                return float(np.median(v))
            lo = np.quantile(v, 0.10)
            hi = np.quantile(v, 0.90)
            vw = np.clip(v, lo, hi)
            return float(np.mean(vw))

        def _to_numeric_robust_series(s: pd.Series) -> pd.Series:
            seqs = s.map(_extract_numeric_sequence)
            return seqs.map(_robust_center).astype(float)

        # -------- Execução do painel --------
        st.markdown(f"## {APP_TITLE}")
        st.caption("Selecione os filtros se necessário")

        if df_in is None or df_in.empty:
            st.warning("Dados Clínicos não carregados.")
            return

        df = _clean_columns(df_in.copy())

        if " " in df.columns and "CBL" not in df.columns:
            df = df.rename(columns={" ": "CBL"})
        if "CBL" not in df.columns:
            df["CBL"] = np.arange(len(df)).astype(str)

        def _clean_cbl(x):
            if pd.isna(x):
                return ""
            s = str(x).strip()
            m = re.search(r"\bCBL\s*\d+\b", s, flags=re.IGNORECASE)
            return m.group(0).replace(" ", "").upper() if m else s

        df["CBL"] = df["CBL"].map(_clean_cbl)

        for c in list(df.columns):
            if "pront" in _norm(c):
                df = df.drop(columns=[c])

        date_cols = [c for c in df.columns if _is_date_like_series(df[c])]
        for c in date_cols:
            try:
                if pd.api.types.is_numeric_dtype(df[c]):
                    dt = pd.to_datetime(df[c], unit="D", origin="1899-12-30", errors="coerce")
                else:
                    dt = pd.to_datetime(df[c], errors="coerce", dayfirst=True)

                dt = dt.dt.normalize()
                dt = dt.where((dt.dt.year >= 1900) & (dt.dt.year <= 2200), pd.NaT)
                df[c] = dt
            except Exception:
                df[c] = pd.NaT

        stage_path_col = "Estadio patologico" if "Estadio patologico" in df.columns else None
        stage_clin_col = "estágio clínico/radiologico" if "estágio clínico/radiologico" in df.columns else None

        df["_estadio_pat_clean"] = df[stage_path_col].map(_clean_stage_field) if stage_path_col else "Não"
        df["_estadio_clin_clean"] = df[stage_clin_col].map(_clean_stage_field) if stage_clin_col else "Não"

        df["Estadiamento"] = df["_estadio_pat_clean"].copy()
        mask_fallback = df["Estadiamento"].isin(["Não", "", "NA"]) | df["Estadiamento"].isna()
        df.loc[mask_fallback, "Estadiamento"] = df.loc[mask_fallback, "_estadio_clin_clean"]
        df["Estadiamento"] = df["Estadiamento"].fillna("Não")

        SIMNAO_COLS = [
            "HF 1° grau", "HF 2° grau", "HAS", "Diabetes ", "Tabagismo",
            "Sintoma", "sangramento", "alt de hab int", "perda de peso", "dor abd", "incontinencia",
            "mucinoso", "angiolinf", "perineural", "perfuração", "margem",
            "budding", "linfonodo lateral", "recrescimento WW",
            "neoadjuvancia", "Oxaliplatina", "Adjuvância",
        ]
        for c in SIMNAO_COLS:
            if c in df.columns:
                df[c] = df[c].map(_to_sim_nao)

        if "Etilismo" in df.columns:
            df["Etilismo"] = df["Etilismo"].map(_to_etilismo)

        NUM_ROBUST = [
            "Idade na entrada",
            "Peso (kg)",
            "Altura (m)",
            "IMC à Cx",
            "IMC ult 3 meses",
            "TISD (meses)",
            "CEA pré",
            "CEA pós",
            "cfDNA BxL1",
            "cfDNA BxLC1 ",
            "cfDNA BxL2",
            "Sum MAF% BXL1",
            "Sum MAF% BXL2",
        ]
        for c in NUM_ROBUST:
            if c in df.columns:
                if pd.api.types.is_numeric_dtype(df[c]):
                    df[f"{c} (robust)"] = pd.to_numeric(df[c], errors="coerce")
                else:
                    df[f"{c} (robust)"] = _to_numeric_robust_series(df[c])

        st.sidebar.markdown("### Coorte")
        include_dates = st.sidebar.toggle("Incluir colunas de data no coorte (sem horário)", value=False)

        st.markdown("#### Filtros clínicos")
        f1, f2, f3 = st.columns([1.6, 1.0, 1.0])

        est_opts = sorted(df["Estadiamento"].dropna().astype(str).unique().tolist(), key=_stage_sort_key)
        sel_est = f1.multiselect("Estadiamento", options=est_opts, default=[])

        dff = df.copy()
        if sel_est:
            dff = dff[dff["Estadiamento"].astype(str).isin(sel_est)]

        if "Sexo" in dff.columns:
            sex_opts = sorted(dff["Sexo"].dropna().astype(str).unique().tolist())
            sel_sex = f2.multiselect("Sexo", options=sex_opts, default=[])
            if sel_sex:
                dff = dff[dff["Sexo"].astype(str).isin(sel_sex)]

        if "ASA" in dff.columns:
            asa_opts = sorted(dff["ASA"].dropna().astype(str).unique().tolist())
            sel_asa = f3.multiselect("ASA", options=asa_opts, default=[])
            if sel_asa:
                dff = dff[dff["ASA"].astype(str).isin(sel_asa)]

        n_patients = int(dff["CBL"].nunique())

        with st.expander("CBLs incluídos (rastreabilidade)", expanded=False):
            cbls = sorted(dff["CBL"].dropna().astype(str).unique().tolist())
            st.write(f"CBLs únicos: {len(cbls)}")
            st.code(", ".join(cbls))

        tab_overview, tab_dist, tab_coorte = st.tabs(["Visão geral", "Distribuições", "Coorte"])

        with tab_overview:
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Pacientes (CBL únicos)", f"{n_patients:,}".replace(",", "."))

            age = dff["Idade na entrada (robust)"] if "Idade na entrada (robust)" in dff.columns else None
            m2.metric("Idade mediana", f"{np.nanmedian(age):.1f}" if age is not None and pd.Series(age).notna().any() else "—")

            imc = None
            if "IMC ult 3 meses (robust)" in dff.columns and dff["IMC ult 3 meses (robust)"].notna().any():
                imc = dff["IMC ult 3 meses (robust)"]
            elif "IMC à Cx (robust)" in dff.columns and dff["IMC à Cx (robust)"].notna().any():
                imc = dff["IMC à Cx (robust)"]
            m3.metric("IMC mediano", f"{np.nanmedian(imc):.1f}" if imc is not None and pd.Series(imc).notna().any() else "—")

            cea_pre = dff["CEA pré (robust)"] if "CEA pré (robust)" in dff.columns else None
            m4.metric("CEA pré mediana", f"{np.nanmedian(cea_pre):.2f}" if cea_pre is not None and pd.Series(cea_pre).notna().any() else "—")

            c1, c2 = st.columns(2)

            _chart_header("Distribuição por Estadiamento", n_patients)
            est_vc = dff["Estadiamento"].fillna("Não").astype(str).value_counts().reset_index()
            est_vc.columns = ["Estadiamento", "Pacientes"]
            est_vc["__ord"] = est_vc["Estadiamento"].map(_stage_sort_key)
            est_vc = est_vc.sort_values("__ord").drop(columns="__ord")

            fig = px.bar(est_vc, x="Estadiamento", y="Pacientes", text="Pacientes")
            fig.update_traces(textposition="outside", cliponaxis=False)
            fig.update_layout(xaxis_title="Estadiamento", yaxis_title="Pacientes")
            fig = _apply_plotly_theme(fig, height=460)
            c1.plotly_chart(fig, use_container_width=True)

            _chart_header("Distribuição por Estádio Patológico", n_patients, subtitle="")
            pat = dff["_estadio_pat_clean"].fillna("Não").astype(str)
            pat_vc = pat.value_counts().reset_index()
            pat_vc.columns = ["Estádio patológico", "Pacientes"]
            pat_vc["__ord"] = pat_vc["Estádio patológico"].map(_stage_sort_key)
            pat_vc = pat_vc.sort_values("__ord").drop(columns="__ord")

            fig2 = px.bar(pat_vc, x="Estádio patológico", y="Pacientes", text="Pacientes")
            fig2.update_traces(textposition="outside", cliponaxis=False)
            fig2 = _apply_plotly_theme(fig2, height=460)
            c2.plotly_chart(fig2, use_container_width=True)

        with tab_dist:
            st.markdown("#### Distribuições")

            candidates = []
            for c in dff.columns:
                if c in ("CBL",):
                    continue
                if c.startswith("_"):
                    continue
                if c in date_cols:
                    continue
                if dff[c].dropna().empty:
                    continue
                candidates.append(c)

            priority = ["Estadiamento", "_estadio_pat_clean", "_estadio_clin_clean", "Sexo", "ASA"]
            ordered = []
            for p in priority:
                if p in candidates and p not in ordered:
                    ordered.append(p)
            for c in candidates:
                if c not in ordered:
                    ordered.append(c)

            col = st.selectbox("Variável", options=ordered, index=0)
            _chart_header(f"Distribuição — {col}", n_patients)

            s = dff[col]
            is_num = pd.api.types.is_numeric_dtype(s)

            if is_num:
                tmp = pd.DataFrame({col: s}).dropna()
                if tmp.empty:
                    st.info("Sem dados suficientes.")
                else:
                    fig = px.histogram(tmp, x=col, nbins=24)
                    fig.update_layout(xaxis_title=col, yaxis_title="Pacientes")
                    fig = _apply_plotly_theme(fig, height=520)
                    st.plotly_chart(fig, use_container_width=True)
            else:
                tmp = s.where(s.notna(), "Não").astype(str).replace({"nan": "Não", "na": "Não", "": "Não"})
                vc = tmp.value_counts().head(40).reset_index()
                vc.columns = [col, "Pacientes"]
                fig = px.bar(vc, x=col, y="Pacientes", text="Pacientes")
                fig.update_traces(textposition="outside", cliponaxis=False)
                fig.update_layout(xaxis_title=col, yaxis_title="Pacientes")
                fig = _apply_plotly_theme(fig, height=520)
                st.plotly_chart(fig, use_container_width=True)

        with tab_coorte:
            st.markdown("#### Coorte clínico")

            BASE = [
                "CBL",
                "Sexo",
                "Idade na entrada",
                "Idade na entrada (robust)",
                "HF 1° grau",
                "HF 2° grau",
                "HAS",
                "Diabetes ",
                "Tabagismo",
                "Etilismo",
                "localização",
                "Sincronico / metacronico",
                "TNM",
                "Estadiamento",
                "_estadio_pat_clean",
                "_estadio_clin_clean",
                "grau de dif",
                "mucinoso",
                "angiolinf",
                "perineural",
                "perfuração",
                "margem",
                "budding",
                "CEA pré",
                "CEA pré (robust)",
                "CEA pós",
                "CEA pós (robust)",
                "cfDNA BxL1",
                "cfDNA BxL1 (robust)",
                "cfDNA BxLC1 ",
                "cfDNA BxLC1  (robust)",
                "cfDNA BxL2",
                "cfDNA BxL2 (robust)",
                "Sum MAF% BXL1",
                "Sum MAF% BXL1 (robust)",
                "Sum MAF% BXL2",
                "Sum MAF% BXL2 (robust)",
                "neoadjuvancia",
                "Esquema RT (0 short, 1 long)",
                "Esquema QT",
                "Oxaliplatina",
                "Adjuvância",
            ]
            cols = [c for c in BASE if c in dff.columns]

            if include_dates:
                cols += [c for c in date_cols if c in dff.columns]

            coorte = dff[cols].copy()

            for c in coorte.columns:
                if c in date_cols:
                    coorte[c] = coorte[c].map(_format_date_only)
                    continue

                if c in ("Estadiamento", "_estadio_pat_clean", "_estadio_clin_clean"):
                    coorte[c] = coorte[c].fillna("Não").astype(str)
                    continue

                if c in SIMNAO_COLS:
                    coorte[c] = coorte[c].map(_to_sim_nao)
                    continue

                if c == "Etilismo":
                    coorte[c] = coorte[c].map(_to_etilismo)
                    continue

                if not pd.api.types.is_numeric_dtype(coorte[c]):
                    coorte[c] = coorte[c].where(coorte[c].notna(), "Não").astype(str)
                    coorte[c] = coorte[c].replace({"nan": "Não", "na": "Não", "": "Não"})

            coorte = coorte.rename(columns={
                "_estadio_pat_clean": "Estádio patológico (limpo)",
                "_estadio_clin_clean": "Estádio clínico/radiológico (limpo)",
            })

            st.dataframe(coorte, use_container_width=True, hide_index=True, height=640)

            csv = coorte.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Exportar CSV (coorte clínico limpo)",
                data=csv,
                file_name="coorte_clinica_ccr_bxl.csv",
                mime="text/csv",
                use_container_width=True
            )

    def painel_agenda_bxl(file_path="Entrada e Acompanhamento de Pacientes Biópsia Líquida CCR.xlsx"):
        """
        PAINEL AGENDA BxL (com WhatsApp) — fluxo direto e persistente
    
        - Lê eventos do Excel (datas nas colunas BXL*) SEM cache
        - Status vem da aba REGISTRO_COLETAS (upsert por chave CBL+Coleta_short+Data_iso)
        - Botões:
          ✅ Registrar coleta  -> grava REGISTRADA em REGISTRO_COLETAS + marca célula (verde) + comentário
          ⏰ Marcar atrasada   -> grava ATRASADA    em REGISTRO_COLETAS + marca célula (amarelo) + comentário
        - WhatsApp:
          - Detecta colunas de telefone/whatsapp
          - Extrai múltiplos números, normaliza para 55DDDNXXXXXXXX e gera link wa.me com templates
          - (AJUSTE) Só permite WhatsApp para coletas de HOJE até +30 dias (não inclui atrasadas)
        """
    
        import re
        import unicodedata
        from datetime import datetime, date
    
        import numpy as np
        import pandas as pd
        import streamlit as st
    
        import openpyxl
        from openpyxl.utils.datetime import from_excel
        from openpyxl.styles import PatternFill, Font
        from openpyxl.comments import Comment
        from urllib.parse import quote
    
        # ============================================================
        # Config / estilos
        # ============================================================
        REG_SHEET = "REGISTRO_COLETAS"
        REG_HEADERS = ["CBL", "Coleta_short", "Data_iso", "Status", "Motivo", "Obs", "Updated_at", "Sheet", "Cell"]
    
        FILL_REG = PatternFill("solid", fgColor="C6EFCE")   # verde
        FILL_LATE = PatternFill("solid", fgColor="FFEB9C")  # amarelo
        FONT_BOLD = Font(bold=True)
        FONT_NORMAL = Font(bold=False)
    
        # ============================================================
        # Session state
        # ============================================================
        if "reg_version_bxl" not in st.session_state:
            st.session_state["reg_version_bxl"] = 0
        if "bxl_auto_guard" not in st.session_state:
            st.session_state["bxl_auto_guard"] = {"day": "", "keys": set()}
    
        # (novo) guard WhatsApp para resetar texto automaticamente sem bug de key
        if "wa_bxl_guard" not in st.session_state:
            st.session_state["wa_bxl_guard"] = {"sig": ""}
    
        # ============================================================
        # Helpers: string / normalização
        # ============================================================
        def _slug(s):
            if s is None:
                return ""
            s = re.sub(r"\s+", " ", str(s)).strip()
            s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
            return s
    
        def _norm(s):
            return _slug(s).lower()
    
        def _valid_year(d):
            return (d is not None) and (1900 <= d.year <= 2200)
    
        def _fmt_date(d):
            if not d or not _valid_year(d):
                return "—"
            return d.strftime("%d/%m/%y")
    
        # ============================================================
        # CBL normalize
        # ============================================================
        _RE_CBL_STRICT = re.compile(r"^CBL\d+$", flags=re.IGNORECASE)
        _RE_CBL_IN_TEXT = re.compile(r"\bCBL\s*(\d+)\b", flags=re.IGNORECASE)
    
        def _normalize_cbl(raw):
            s = str(raw or "").strip()
            m = _RE_CBL_IN_TEXT.search(s)
            if m:
                return "CBL{}".format(int(m.group(1)))
            return s
    
        def _is_in_study_id(raw):
            s = str(raw or "").strip().replace(" ", "")
            return bool(_RE_CBL_STRICT.match(s))
    
        # ============================================================
        # Coleta short label
        # ============================================================
        def _coleta_short_label(coleta):
            cn = _norm(coleta)
            m = re.search(r"\bbxl\s*(\d+)\b", cn)
            if m:
                return "BxL {}".format(int(m.group(1)))
            if "pos" in cn and "rt" in cn:
                return "BxL pós-RT"
            if "cirurg" in cn:
                return "Cirurgia"
            if "rm" in cn:
                return "RM"
            if "tc" in cn:
                return "TC"
            return str(coleta).strip()
    
        # ============================================================
        # Datas — robusta (serial Excel / datetime / strings)
        # ============================================================
        RE_YMD = re.compile(r"(?<!\d)(\d{4})[\/\-.](\d{1,2})[\/\-.](\d{1,2})(?!\d)")
        RE_DMY = re.compile(r"(?<!\d)(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})(?!\d)")
        RE_DM  = re.compile(r"(?<!\d)(\d{1,2})[\/\-.](\d{1,2})(?!\d)")
    
        def _two_digit_year_to_yyyy(yy):
            return 2000 + yy if yy <= 69 else 1900 + yy
    
        def _infer_year_for_ddmm(d, m, today_):
            y = today_.year
            try:
                cand = date(y, m, d)
            except Exception:
                return None
    
            if cand <= today_:
                if (today_ - cand).days <= 7:
                    return cand
                try:
                    return date(y + 1, m, d)
                except Exception:
                    return None
            return cand
    
        def _extract_first_date_from_string(s, today_):
            if not s:
                return None
            s0 = str(s).strip()
            if not s0 or s0.lower() in ("nan", "na", "n/a", "-"):
                return None
    
            s1 = (" " + s0.lower().replace("\\", "/") + " ")
            s1 = re.sub(r"[,\t]+", " ", s1)
            s1 = re.sub(r"\s+", " ", s1).strip()
    
            m = RE_YMD.search(s1)
            if m:
                y, mo, d = m.groups()
                try:
                    dd = date(int(y), int(mo), int(d))
                    return dd if _valid_year(dd) else None
                except Exception:
                    pass
    
            m = RE_DMY.search(s1)
            if m:
                d, mo, y = m.groups()
                try:
                    yv = _two_digit_year_to_yyyy(int(y)) if len(y) == 2 else int(y)
                    dd = date(yv, int(mo), int(d))
                    return dd if _valid_year(dd) else None
                except Exception:
                    pass
    
            m = RE_DM.search(s1)
            if m:
                d, mo = m.groups()
                return _infer_year_for_ddmm(int(d), int(mo), today_)
    
            return None
    
        def _extract_date(x, today_, epoch):
            if x is None or (isinstance(x, float) and np.isnan(x)):
                return None
    
            if isinstance(x, datetime):
                d = x.date()
                return d if _valid_year(d) else None
    
            if isinstance(x, date):
                return x if _valid_year(x) else None
    
            if isinstance(x, (int, float, np.integer, np.floating)):
                try:
                    v = float(x)
                    if 1 <= v <= 60000:
                        dtv = from_excel(v, epoch=epoch)
                        if isinstance(dtv, datetime):
                            d = dtv.date()
                            return d if _valid_year(d) else None
                except Exception:
                    pass
    
            return _extract_first_date_from_string(str(x), today_)
    
        # ============================================================
        # (NOVO) Tag atrasada: baseado APENAS no texto da célula
        # ============================================================
        def _has_late_tag(raw):
            s = ("" if raw is None else str(raw)).strip().lower()
            # pega "(atrasada)" ou "(atrasado)" em qualquer variação
            return ("(atrasad" in s)
    
        # ============================================================
        # Registro: ler / status map
        # ============================================================
        def _read_registro_df(wb):
            if REG_SHEET not in wb.sheetnames:
                return pd.DataFrame(columns=REG_HEADERS)
    
            ws = wb[REG_SHEET]
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                return pd.DataFrame(columns=REG_HEADERS)
    
            hdr = [str(x).strip() if x is not None else "" for x in rows[0]]
            data = rows[1:]
            df_reg = pd.DataFrame(data, columns=hdr if (data and len(hdr) == len(data[0])) else None)
    
            for c in REG_HEADERS:
                if c not in df_reg.columns:
                    df_reg[c] = None
            df_reg = df_reg[REG_HEADERS].copy()
    
            df_reg["CBL"] = df_reg["CBL"].astype(str).str.strip()
            df_reg["Coleta_short"] = df_reg["Coleta_short"].astype(str).str.strip()
            df_reg["Data_iso"] = df_reg["Data_iso"].astype(str).str.strip()
            df_reg["Status"] = df_reg["Status"].astype(str).str.strip().str.upper()
            df_reg["Updated_at"] = df_reg["Updated_at"].astype(str).str.strip()
    
            return df_reg
    
        def _status_map(df_reg):
            m = {}
            if df_reg is None or df_reg.empty:
                return m
    
            df2 = df_reg.copy()
            try:
                df2["_ts"] = pd.to_datetime(df2["Updated_at"], errors="coerce")
                df2 = df2.sort_values("_ts")
            except Exception:
                pass
    
            for _, r in df2.iterrows():
                cbl = str(r.get("CBL", "")).strip()
                col = str(r.get("Coleta_short", "")).strip()
                diso = str(r.get("Data_iso", "")).strip()
                stt = str(r.get("Status", "")).strip().upper() or "PENDENTE"
                if cbl and col and diso:
                    m["{}||{}||{}".format(cbl, col, diso)] = stt
            return m
    
        # ============================================================
        # Save: upsert + marcar célula + salvar direto
        # ============================================================
        def _ensure_reg_sheet(wb):
            if REG_SHEET not in wb.sheetnames:
                ws_reg = wb.create_sheet(REG_SHEET)
                ws_reg.append(REG_HEADERS)
                return ws_reg
            ws_reg = wb[REG_SHEET]
            if ws_reg.max_row < 1:
                ws_reg.append(REG_HEADERS)
            return ws_reg
    
        def _build_reg_col_map(ws_reg):
            header_vals = [ws_reg.cell(row=1, column=i).value for i in range(1, ws_reg.max_column + 1)]
            header_vals = [str(x).strip() if x is not None else "" for x in header_vals]
            col_map = {h: i + 1 for i, h in enumerate(header_vals) if h}
    
            for h in REG_HEADERS:
                if h not in col_map:
                    new_col = ws_reg.max_column + 1
                    ws_reg.cell(row=1, column=new_col).value = h
                    col_map[h] = new_col
            return col_map
    
        def _apply_cell_mark(ws, cell_ref, d, status, motivo, obs):
            """
            Marca a célula original no Excel.
            - REGISTRADA: escreve datetime real (dd/mm/yy)
            - ATRASADA/ATRASADO: escreve TEXTO "dd/mm/yy (atrasada)"
            """
            import re
            from datetime import datetime
        
            if not cell_ref:
                raise ValueError("cell_ref vazio (não dá pra marcar no Excel).")
        
            cell = ws[cell_ref]
        
            # normaliza status (tolerante a variações)
            stt_raw = "" if status is None else str(status)
            stt = re.sub(r"\s+", " ", stt_raw).strip().upper()
        
            # aceita ATRASADA / ATRASADO / qualquer coisa que contenha "ATRASAD"
            is_late = ("ATRASAD" in stt)
        
            if is_late:
                # escreve como TEXTO pra garantir que o sufixo apareça no Excel
                cell.value = f"{d.strftime('%d/%m/%y')} (atrasada)"
                cell.number_format = "@"
                cell.fill = FILL_LATE
                cell.font = FONT_BOLD
            else:
                # data real
                cell.value = datetime(d.year, d.month, d.day)
                cell.number_format = "dd/mm/yy"
        
                if stt == "REGISTRADA":
                    cell.fill = FILL_REG
                    cell.font = FONT_BOLD
                else:
                    cell.fill = PatternFill()
                    cell.font = FONT_NORMAL
        
            # comentário
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            txt = "{} | {}\n{}".format(stt if stt else ("ATRASADA" if is_late else "PENDENTE"), now, motivo).strip()
            if obs:
                txt += "\nObs: {}".format(str(obs).strip())
        
            try:
                cell.comment = Comment(txt[:700], "AgendaBxL")
            except Exception:
                pass

        def _save_status_to_excel(cbl, coleta_short, dd, status, motivo, obs, sheet_name, cell_ref):
            wb = openpyxl.load_workbook(file_path)
            ws_reg = _ensure_reg_sheet(wb)
            col_map = _build_reg_col_map(ws_reg)
    
            k_cbl = str(cbl).strip()
            k_col = str(coleta_short).strip()
            k_dat = dd.isoformat()
    
            # upsert por chave
            target_row = None
            for r in range(2, ws_reg.max_row + 1):
                v1 = str(ws_reg.cell(row=r, column=col_map["CBL"]).value or "").strip()
                v2 = str(ws_reg.cell(row=r, column=col_map["Coleta_short"]).value or "").strip()
                v3 = str(ws_reg.cell(row=r, column=col_map["Data_iso"]).value or "").strip()
                if v1 == k_cbl and v2 == k_col and v3 == k_dat:
                    target_row = r
                    break
            if target_row is None:
                target_row = ws_reg.max_row + 1
    
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            rec = {
                "CBL": k_cbl,
                "Coleta_short": k_col,
                "Data_iso": k_dat,
                "Status": str(status).upper(),
                "Motivo": motivo,
                "Obs": obs,
                "Updated_at": now,
                "Sheet": sheet_name,
                "Cell": cell_ref,
            }
    
            for h in REG_HEADERS:
                ws_reg.cell(row=target_row, column=col_map[h]).value = rec.get(h)
    
            # marca célula original
            if sheet_name in wb.sheetnames and cell_ref:
                try:
                    ws = wb[sheet_name]
                    _apply_cell_mark(ws, cell_ref, dd, rec["Status"], motivo, obs)
                except Exception:
                    pass
    
            wb.save(file_path)
    
        # ============================================================
        # Load events: Excel -> DataFrame (SEM cache)
        # ============================================================
        def _load_events(sheet_name):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            epoch = wb.epoch
    
            headers = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=1, column=c).value
                headers.append("" if v is None else str(v).strip())
    
            # coluna CBL
            cbl_col = None
            for idx, h in enumerate(headers, start=1):
                nh = _norm(h)
                if "cbl" in nh or ("id" in nh and "paciente" in nh):
                    cbl_col = idx
                    break
            if cbl_col is None:
                cbl_col = 1
    
            # colunas de telefone
            phone_cols = []
            for idx, h in enumerate(headers, start=1):
                nh = _norm(h)
                if any(k in nh for k in ["telefone", "celular", "whats", "whatsapp", "contato", "fone"]):
                    phone_cols.append(idx)
    
            # colunas BxL
            bxl_cols = []
            for idx, h in enumerate(headers, start=1):
                nh = _norm(h)
                if ("bxl" in nh) and ("obs" not in nh):
                    bxl_cols.append(idx)
    
            today_ = date.today()
            events = []
            phones_map = {}
    
            for r in range(2, ws.max_row + 1):
                cbl_val = ws.cell(row=r, column=cbl_col).value
                if cbl_val is None or str(cbl_val).strip() == "":
                    continue
    
                cbl_raw = str(cbl_val).strip()
                cbl = _normalize_cbl(cbl_raw)
    
                # telefones por coluna
                phone_raw_by_col = {}
                for pc in phone_cols:
                    hname = headers[pc - 1] or "Telefone_{}".format(pc)
                    phone_raw_by_col[hname] = ws.cell(row=r, column=pc).value
    
                if cbl not in phones_map:
                    phones_map[cbl] = phone_raw_by_col
                else:
                    for k, v in phone_raw_by_col.items():
                        cur = phones_map[cbl].get(k)
                        if (cur is None or str(cur).strip() == "") and (v is not None and str(v).strip() != ""):
                            phones_map[cbl][k] = v
    
                for bc in bxl_cols:
                    head = headers[bc - 1] or "BxL_{}".format(bc)
                    cell = ws.cell(row=r, column=bc)
                    raw = cell.value
                    raw_str = "" if raw is None else str(raw).strip()
    
                    dd = _extract_date(raw, today_=today_, epoch=epoch)
                    if not dd:
                        continue
    
                    events.append({
                        "CBL": cbl,
                        "CBL_RAW": cbl_raw,
                        "IN_STUDY": _is_in_study_id(cbl_raw),
                        "Coleta_short": _coleta_short_label(str(head)),
                        "Data": dd,
                        "Data_str": _fmt_date(dd),
                        "Dias": (dd - today_).days,
                        "Sheet": sheet_name,
                        "Cell": cell.coordinate,
                        "Raw": raw_str,
                        "LateTag": _has_late_tag(raw_str),   # (NOVO) atraso por TAG na célula
                    })
    
            df = pd.DataFrame(events)
            if not df.empty:
                df = df.drop_duplicates(subset=["CBL", "Coleta_short", "Data", "Sheet", "Cell"]).reset_index(drop=True)
    
            wb2 = openpyxl.load_workbook(file_path, data_only=True)
            df_reg = _read_registro_df(wb2)
            smap = _status_map(df_reg)
    
            if df.empty:
                return df, phones_map, headers, phone_cols, df_reg, smap
    
            df["KeyStatus"] = df.apply(lambda r: "{}||{}||{}".format(r["CBL"], r["Coleta_short"], r["Data"].isoformat()), axis=1)
            df["Status"] = df["KeyStatus"].apply(lambda k: smap.get(k, "PENDENTE")).astype(str).str.upper()
    
            return df, phones_map, headers, phone_cols, df_reg, smap
    
        # ============================================================
        # (NOVO) Painel de Pacientes: cadastrados, excluídos, motivos
        # ============================================================
        def painel_pacientes(sheet_name):
            st.divider()
            st.markdown("### Painel de Pacientes (cadastro / exclusões)")
    
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                if sheet_name not in wb.sheetnames:
                    st.warning("Aba não encontrada para painel de pacientes.")
                    return
                ws = wb[sheet_name]
            except Exception as e:
                st.warning(f"Falha ao ler planilha para painel de pacientes: {e}")
                return
    
            # cabeçalhos
            headers = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=1, column=c).value
                headers.append("" if v is None else str(v).strip())
    
            def _find_col(pred):
                for idx, h in enumerate(headers, start=1):
                    if pred(_norm(h)):
                        return idx
                return None
    
            col_id = _find_col(lambda x: ("cbl" in x) or ("id" in x and "paciente" in x))
            col_status = _find_col(lambda x: ("status" in x and "estudo" in x) or ("status" in x and "study" in x))
            col_data_exc = _find_col(lambda x: ("data" in x and "exclus" in x))
            col_motivo = _find_col(lambda x: ("motivo" in x and "exclus" in x))
    
            if not col_id:
                st.info("Não encontrei a coluna de ID/CBL nesta aba para montar o painel de pacientes.")
                return
    
            rows = []
            today_ = date.today()
            epoch = wb.epoch
    
            for r in range(2, ws.max_row + 1):
                rid = ws.cell(row=r, column=col_id).value
                if rid is None or str(rid).strip() == "":
                    continue
    
                cbl = _normalize_cbl(rid)
                status = "" if not col_status else (ws.cell(row=r, column=col_status).value or "")
                motivo = "" if not col_motivo else (ws.cell(row=r, column=col_motivo).value or "")
                d_exc_raw = "" if not col_data_exc else ws.cell(row=r, column=col_data_exc).value
    
                d_exc = _extract_date(d_exc_raw, today_=today_, epoch=epoch)
    
                status_s = str(status).strip()
                motivo_s = str(motivo).strip()
    
                excl_flag = False
                if status_s and ("exclu" in status_s.lower()):
                    excl_flag = True
                if motivo_s and motivo_s.lower() not in ("nan", "na", "n/a", "-"):
                    excl_flag = True
                if d_exc is not None:
                    excl_flag = True
    
                rows.append({
                    "CBL": cbl,
                    "Status_estudo": status_s,
                    "Data_exclusao": d_exc.strftime("%d/%m/%y") if d_exc else "",
                    "Motivo_exclusao": motivo_s,
                    "Excluido": excl_flag
                })
    
            dfp = pd.DataFrame(rows)
            if dfp.empty:
                st.info("Sem pacientes encontrados nesta aba.")
                return
    
            # métricas por CBL (evita duplicados)
            dfp_u = dfp.drop_duplicates(subset=["CBL"]).copy()
            total = int(dfp_u["CBL"].nunique())
            excl = int(dfp_u[dfp_u["Excluido"] == True]["CBL"].nunique())
            ativos = total - excl
    
            c1, c2, c3 = st.columns(3)
            c1.metric("Pacientes cadastrados", total)
            c2.metric("Excluídos", excl)
            c3.metric("Ativos", ativos)
    
            # motivos
            if "Motivo_exclusao" in dfp_u.columns:
                motivos = dfp_u[dfp_u["Excluido"] == True]["Motivo_exclusao"].replace("", np.nan).dropna()
                if not motivos.empty:
                    vc = motivos.value_counts().reset_index()
                    vc.columns = ["Motivo", "Qtd"]
                    st.markdown("#### Motivos de exclusão")
                    st.dataframe(vc, use_container_width=True, hide_index=True)
    
            # lista de excluídos
            st.markdown("#### Pacientes excluídos (detalhes)")
            df_exc = dfp_u[dfp_u["Excluido"] == True].copy()
            if df_exc.empty:
                st.info("Nenhum excluído detectado nesta aba.")
            else:
                st.dataframe(
                    df_exc[["CBL", "Data_exclusao", "Motivo_exclusao", "Status_estudo"]],
                    use_container_width=True,
                    hide_index=True,
                    height=260
                )
    
        # ============================================================
        # WhatsApp — painel robusto (AJUSTADO: apenas 0..30)
        # ============================================================
        def painel_whatsapp_bxl(df_all_events, phones_map, headers, phone_cols):
            import re
            import numpy as np
            import pandas as pd
            import streamlit as st
            from urllib.parse import quote
    
            def _extract_phones(raw):
                if raw is None or (isinstance(raw, float) and np.isnan(raw)):
                    return []
                s = str(raw).strip()
                if not s:
                    return []
                s = s.replace("\n", " ").replace("\t", " ")
                s = re.sub(r"[;,/|]+", " ", s)
                s = re.sub(r"\s+", " ", s).strip()
    
                candidates = re.findall(
                    r"(?:\+?\s*55\s*)?\(?\s*(\d{2})\s*\)?\s*(9?\d{4})[-\s]?(\d{4})",
                    s
                )
    
                out = []
                for ddd, p1, p2 in candidates:
                    n = "{}{}{}".format(ddd, p1, p2)
                    if 10 <= len(n) <= 11 and n not in out:
                        out.append(n)
                return out
    
            def _normalize_to_e164_br(digits):
                p = re.sub(r"\D", "", str(digits or ""))
                if not p:
                    return ""
                if p.startswith("55") and 12 <= len(p) <= 13:
                    return p
                if 10 <= len(p) <= 11:
                    return "55" + p
                return ""
    
            def _format_phone_br(digits):
                p = re.sub(r"\D", "", str(digits or ""))
                if p.startswith("55"):
                    p = p[2:]
                if len(p) == 10:
                    return "({}) {}-{}".format(p[:2], p[2:6], p[6:])
                if len(p) == 11:
                    return "({}) {}-{}".format(p[:2], p[2:7], p[7:])
                return str(digits)
    
            def _wa_link(phone_e164_no_plus, msg):
                if not phone_e164_no_plus:
                    return ""
                return "https://wa.me/{}?text={}".format(phone_e164_no_plus, quote(msg or ""))
    
            st.divider()
            st.markdown("### WhatsApp — Enviar mensagens")
            st.caption("Apenas coletas de hoje até +30 dias (não inclui atrasadas).")
    
            if df_all_events is None or df_all_events.empty:
                st.info("Sem eventos para WhatsApp.")
                return
    
            df = df_all_events.copy()
            if "Dias" not in df.columns:
                df["Dias"] = np.nan
            if "Status" not in df.columns:
                df["Status"] = "PENDENTE"
            if "Status_display" not in df.columns:
                df["Status_display"] = df["Status"].astype(str)
    
            df["Status"] = df["Status"].astype(str).str.upper()
            dias_num = pd.to_numeric(df["Dias"], errors="coerce")
    
            # (AJUSTE) Só 0..30 e NÃO REGISTRADA e NÃO ATRASADA (tag/célula)
            status_mix = (df["Status_display"].astype(str) + " " + df["Status"].astype(str)).str.upper()
            is_reg = status_mix.str.contains("REGISTRADA", na=False)
            is_late = status_mix.str.contains("ATRASADA", na=False) | df.get("LateTag", False).astype(bool)
    
            is_futura_0_30 = dias_num.between(0, 30, inclusive="both")
            df_src = df[(~is_reg) & (~is_late) & (is_futura_0_30)].copy()
    
            if df_src.empty:
                st.info("Nenhuma coleta elegível (0..30 dias) para WhatsApp.")
                return
    
            # Colunas de telefone detectadas
            if phone_cols and isinstance(headers, list) and len(headers) > 0:
                phone_colnames = [(headers[pc - 1] if (pc - 1) < len(headers) else f"Telefone_{pc}") for pc in phone_cols]
                phone_colnames = [
                    str(x).strip() if x is not None and str(x).strip() else f"Telefone_{i+1}"
                    for i, x in enumerate(phone_colnames)
                ]
            else:
                phone_colnames = ["Telefone"]
    
            wa_scope = "wa_bxl_0_30"
    
            c1, c2 = st.columns([1.2, 1.0])
            with c1:
                cbl_options = sorted(df_src["CBL"].dropna().astype(str).unique().tolist())
                if not cbl_options:
                    st.warning("Sem CBL válido após o filtro.")
                    return
                sel_cbl = st.selectbox("CBL (paciente)", options=cbl_options, index=0, key=f"{wa_scope}_cbl")
            with c2:
                sel_phone_col = st.selectbox("Coluna de contato", options=phone_colnames, index=0, key=f"{wa_scope}_phonecol")
    
            ev = df_src[df_src["CBL"].astype(str) == str(sel_cbl)].copy()
            if "Data" in ev.columns:
                ev = ev.sort_values(["Data"], ascending=True)
            ev = ev.reset_index(drop=True)
    
            if ev.empty:
                st.info("Não há eventos elegíveis para este CBL.")
                return
    
            def _safe_int(x, default=0):
                try:
                    if pd.isna(x):
                        return default
                    return int(float(x))
                except Exception:
                    return default
    
            ev["Label"] = ev.apply(
                lambda r: "{} — {} (D{})".format(
                    str(r.get("Coleta_short", "")).strip(),
                    str(r.get("Data_str", "")).strip(),
                    _safe_int(r.get("Dias", 0), 0)
                ),
                axis=1
            )
    
            sel_label = st.selectbox("Evento (coleta)", options=ev["Label"].tolist(), index=0, key=f"{wa_scope}_event")
            row = ev[ev["Label"] == sel_label].iloc[0]
    
            # Telefone
            raw_contact = ""
            if isinstance(phones_map, dict) and sel_cbl in phones_map:
                try:
                    raw_contact = phones_map[sel_cbl].get(sel_phone_col, "") or ""
                except Exception:
                    raw_contact = ""
    
            phones = _extract_phones(raw_contact)
            phones_e164 = []
            for p in phones:
                e = _normalize_to_e164_br(p)
                if e:
                    phones_e164.append((e, p))
    
            if phones_e164:
                display_opts = ["{}  (→ {})".format(_format_phone_br(p), e) for (e, p) in phones_e164]
                sel_idx = st.selectbox(
                    "Contato (telefone encontrado)",
                    options=list(range(len(display_opts))),
                    format_func=lambda i: display_opts[i],
                    index=0,
                    key=f"{wa_scope}_phoneidx"
                )
                phone_e164 = phones_e164[int(sel_idx)][0]
                phone_disp = _format_phone_br(phones_e164[int(sel_idx)][1])
            else:
                st.warning("Não encontrei telefone válido na coluna escolhida. Cole um número manualmente.")
                manual = st.text_input("Telefone manual (DDD + número; opcionalmente com 55)", value="", key=f"{wa_scope}_manual")
                phone_e164 = _normalize_to_e164_br(manual)
                phone_disp = _format_phone_br(manual) if manual else ""
    
            # Mensagem (corrigido: reseta quando muda seleção/modelo)
            st.markdown("#### Mensagem")
    
            template = st.selectbox(
                "Modelo",
                ["Lembrete padrão", "Confirmação de presença", "Solicitar reagendamento", "Mensagem neutra"],
                index=0,
                key=f"{wa_scope}_template"
            )
    
            coleta_txt = str(row.get("Coleta_short", "")).strip()
            data_txt = str(row.get("Data_str", "")).strip()
    
            if template == "Lembrete padrão":
                default_msg = (
                    "Olá! Lembrete: sua coleta {} está agendada para {}. "
                    "Se precisar reagendar, por favor responda por aqui."
                ).format(coleta_txt, data_txt)
            elif template == "Confirmação de presença":
                default_msg = (
                    "Olá! Você pode confirmar sua presença para a coleta {} em {}? "
                    "Se precisar ajustar a data, me avise por aqui."
                ).format(coleta_txt, data_txt)
            elif template == "Solicitar reagendamento":
                default_msg = (
                    "Olá! Sobre a coleta {} prevista para {}: você precisa reagendar? "
                    "Se sim, me informe sua disponibilidade."
                ).format(coleta_txt, data_txt)
            else:
                default_msg = "Olá! Estou entrando em contato sobre sua coleta. Pode me responder por aqui, por favor?"
    
            # (CRÍTICO) assinatura para reset do texto
            sig = f"{sel_cbl}||{sel_label}||{template}"
            msg_key = f"{wa_scope}_msg"
    
            if st.session_state["wa_bxl_guard"].get("sig", "") != sig:
                st.session_state["wa_bxl_guard"]["sig"] = sig
                st.session_state[msg_key] = default_msg
    
            msg = st.text_area("Texto", height=120, key=msg_key)
            link = _wa_link(phone_e164, msg)
    
            a1, a2 = st.columns([1.0, 1.0])
            with a1:
                if not phone_e164:
                    st.error("Número inválido. Informe um telefone válido (DDD + número, opcionalmente com 55).")
                else:
                    st.link_button("Abrir WhatsApp com mensagem", link, use_container_width=True)
            with a2:
                st.caption("Contato: {}".format(phone_disp if phone_disp else "—"))
                st.caption("CBL: {} | Coleta: {} | Data: {}".format(sel_cbl, coleta_txt, data_txt))
    
            with st.expander("Ver campo bruto da planilha", expanded=False):
                st.write("Coluna selecionada: {}".format(sel_phone_col))
                st.code(str(raw_contact))
    
        # ============================================================
        # UI principal
        # ============================================================
        st.markdown("## Agenda de Coletas (BxL)")
        st.caption("Leitura e gravação direta no Excel.")
    
        try:
            wb0 = openpyxl.load_workbook(file_path, data_only=True)
            sheetnames = wb0.sheetnames
            if not sheetnames:
                st.error("A planilha não contém abas.")
                return
        except Exception as e:
            st.error("Não consegui abrir o Excel: {}".format(e))
            return
    
        st.sidebar.markdown("### Filtros")
        sheet = st.sidebar.selectbox("Aba da planilha", sheetnames, index=0, key="bxl_sheet_sel")
        cbl_search = st.sidebar.text_input("Buscar CBL (contém)", value="", key="bxl_cbl_search").strip()
        hide_withdrawn = st.sidebar.checkbox("Ocultar retirados do estudo (ID != CBL123)", value=True, key="bxl_hide_withdrawn")
    
        df_all, phones_map, headers, phone_cols, df_reg, smap = _load_events(sheet)
        if df_all.empty:
            st.warning("Nenhuma data BxL encontrada nesta aba.")
            return
    
        # (NOVO) painel pacientes
        painel_pacientes(sheet)
    
        # aplica filtros
        df = df_all.copy()
        if hide_withdrawn and "IN_STUDY" in df.columns:
            df = df[df["IN_STUDY"] == True].copy()
        if cbl_search:
            df = df[df["CBL"].astype(str).str.contains(cbl_search, case=False, na=False)].copy()
        if df.empty:
            st.info("Após os filtros, não restaram eventos.")
            return
    
        # ============================================================
        # Status display (ALTERADO: atrasada por TAG na célula)
        # ============================================================
        def _status_display(r):
            stt = str(r.get("Status", "")).upper().strip()
            late_tag = bool(r.get("LateTag", False))
    
            if stt == "REGISTRADA":
                return "REGISTRADA"
            if late_tag:
                return "ATRASADA"
            if stt == "ATRASADA":
                return "ATRASADA"
            return stt if stt else "PENDENTE"
    
        df["Status_display"] = df.apply(_status_display, axis=1)
    
        def _prefix_status(stt, coleta):
            stt = (stt or "").upper()
            if stt == "ATRASADA":
                return "⏰ {}".format(coleta)
            if stt == "REGISTRADA":
                return "✅ {}".format(coleta)
            return "📌 {}".format(coleta)
    
        df["Coleta_show"] = df.apply(lambda r: _prefix_status(r["Status_display"], r["Coleta_short"]), axis=1)
    
        # recortes (ALTERADO)
        atras = df[df["Status_display"] == "ATRASADA"].copy().sort_values(["Data", "CBL", "Coleta_short"]).reset_index(drop=True)
        fut = df[(df["Dias"] >= 0) & (df["Dias"] <= 30) & (df["Status_display"] != "ATRASADA")].copy() \
                .sort_values(["Data", "CBL", "Coleta_short"]).reset_index(drop=True)
    
        fut_0_7 = fut[(fut["Dias"] >= 0) & (fut["Dias"] <= 7)]
        fut_8_14 = fut[(fut["Dias"] >= 8) & (fut["Dias"] <= 14)]
        fut_15_30 = fut[(fut["Dias"] >= 15) & (fut["Dias"] <= 30)]
    
        # métricas
        st.markdown("### Visão geral")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Atrasadas", int(len(atras)))
        m2.metric("0–7 dias", int(len(fut_0_7)))
        m3.metric("8–14 dias", int(len(fut_8_14)))
        m4.metric("15–30 dias", int(len(fut_15_30)))
    
        # abas
        tabA, tab1, tab2, tab3 = st.tabs(["Atrasadas", "0–7 dias", "8–14 dias", "15–30 dias"])
        SHOW_COLS = ["CBL", "Coleta_show", "Data_str", "Dias", "Status_display"]
    
        with tabA:
            if atras.empty:
                st.info("Nenhuma coleta marcada como atrasada na planilha.")
            else:
                st.dataframe(atras[SHOW_COLS].rename(columns={"Coleta_show": "Coleta", "Status_display": "Status"}),
                             use_container_width=True, hide_index=True, height=420)
    
        with tab1:
            if fut_0_7.empty:
                st.info("Nenhuma coleta entre 0 e 7 dias.")
            else:
                st.dataframe(fut_0_7[SHOW_COLS].rename(columns={"Coleta_show": "Coleta", "Status_display": "Status"}),
                             use_container_width=True, hide_index=True, height=420)
    
        with tab2:
            if fut_8_14.empty:
                st.info("Nenhuma coleta entre 8 e 14 dias.")
            else:
                st.dataframe(fut_8_14[SHOW_COLS].rename(columns={"Coleta_show": "Coleta", "Status_display": "Status"}),
                             use_container_width=True, hide_index=True, height=420)
    
        with tab3:
            if fut_15_30.empty:
                st.info("Nenhuma coleta entre 15 e 30 dias.")
            else:
                st.dataframe(fut_15_30[SHOW_COLS].rename(columns={"Coleta_show": "Coleta", "Status_display": "Status"}),
                             use_container_width=True, hide_index=True, height=420)
    
        # ============================================================
        # Auto D-1 (mantido) — agora marca a célula como "dd/mm/yy (atrasada)"
        # ============================================================
        today_ = date.today()
        guard_day = today_.isoformat()
        if st.session_state["bxl_auto_guard"]["day"] != guard_day:
            st.session_state["bxl_auto_guard"] = {"day": guard_day, "keys": set()}
    
        auto = df_all[
            (df_all["Dias"] == -1) &
            (df_all["Status"] != "REGISTRADA") &
            (~df_all.get("LateTag", False).astype(bool))
        ].copy()
    
        if not auto.empty:
            auto = auto[~auto["KeyStatus"].isin(st.session_state["bxl_auto_guard"]["keys"])].copy()
    
        if not auto.empty:
            attempted = auto["KeyStatus"].tolist()
            ok = 0
            for _, r in auto.iterrows():
                try:
                    _save_status_to_excel(
                        cbl=str(r["CBL"]).strip(),
                        coleta_short=str(r["Coleta_short"]).strip(),
                        dd=r["Data"],
                        status="ATRASADA",
                        motivo="Auto: D-1 e não registrada como coletada",
                        obs="",
                        sheet_name=str(r.get("Sheet", sheet)),
                        cell_ref=str(r.get("Cell", "")),
                    )
                    ok += 1
                except Exception:
                    break
    
            st.session_state["bxl_auto_guard"]["keys"].update(set(attempted))
            if ok > 0:
                st.session_state["reg_version_bxl"] += 1
                st.rerun()
    
        # ============================================================
        # Registro manual (ALTERADO: só atrasadas tag + futuras 0..30)
        # ============================================================
        st.divider()
        st.markdown("### Registro de coleta")
        st.caption("Futuras (0–30) e atrasadas.")
    
        action_df = pd.concat([atras, fut], ignore_index=True)
        action_df = action_df.drop_duplicates(subset=["CBL", "Coleta_short", "Data", "Sheet", "Cell"]).copy()
        action_df = action_df.sort_values(["Dias", "Data", "CBL", "Coleta_short"]).reset_index(drop=True)
        if action_df.empty:
            st.info("Nada para registrar no recorte atual.")
            return
    
        action_df["KeyUI"] = action_df.apply(
            lambda r: "{}||{}||{}||{}||{}".format(
                r["CBL"], r["Coleta_short"], r["Data"].isoformat(),
                r.get("Sheet", ""), r.get("Cell", "")
            ),
            axis=1
        )
        action_df["LabelUI"] = action_df.apply(
            lambda r: "{} — {} — {} — {} (D{})".format(
                r["CBL"], r["Coleta_short"], r["Data_str"], r["Status_display"], int(r["Dias"])
            ),
            axis=1
        )
        label_map = dict(zip(action_df["KeyUI"], action_df["LabelUI"]))
    
        sel_key = st.selectbox(
            "Evento para atualizar",
            options=action_df["KeyUI"].tolist(),
            index=0,
            format_func=lambda k: label_map.get(k, str(k)),
            key="agenda_bxl_sel_event"
        )
        sel_row = action_df[action_df["KeyUI"] == sel_key].iloc[0]
    
        obs_reg = st.text_input("Observação (opcional)", value="", key="agenda_bxl_obs").strip()
    
        def _do_set_status(new_status, motivo):
            try:
                _save_status_to_excel(
                    cbl=str(sel_row["CBL"]).strip(),
                    coleta_short=str(sel_row["Coleta_short"]).strip(),
                    dd=sel_row["Data"],
                    status=str(new_status).upper(),
                    motivo=motivo,
                    obs=obs_reg,
                    sheet_name=str(sel_row.get("Sheet", sheet)),
                    cell_ref=str(sel_row.get("Cell", "")),
                )
                st.session_state["reg_version_bxl"] += 1
                return True
            except Exception as e:
                st.error("Erro ao salvar: {}".format(e))
                return False
    
        b1, b2, b3 = st.columns([1.0, 1.0, 1.6])
        with b1:
            if st.button("✅ Registrar coleta", use_container_width=True, key="agenda_bxl_btn_reg"):
                if _do_set_status("REGISTRADA", "Manual: registrada no painel"):
                    st.rerun()
    
        with b2:
            if st.button("⏰ Marcar como atrasada", use_container_width=True, key="agenda_bxl_btn_late"):
                if _do_set_status("ATRASADA", "Manual: marcada como atrasada no painel"):
                    st.rerun()
    
        with b3:
            st.caption(
                "CBL: {} | Coleta: {} | Data: {} | Dias: {} | Status: {}".format(
                    sel_row["CBL"], sel_row["Coleta_short"], sel_row["Data_str"], int(sel_row["Dias"]), sel_row["Status_display"]
                )
            )
    
        # ============================================================
        # WhatsApp (AGORA: só futuras 0–30)
        # ============================================================
        painel_whatsapp_bxl(df, phones_map, headers, phone_cols)
            

            



    
    



    
       

    # =========================
    # ROUTER ÚNICO DO PAINEL (render)
    # =========================
    PATH_BXL = "Entrada e Acompanhamento de Pacientes Biópsia Líquida CCR.xlsx"
    PATH_CCR = "Dados Clínicos Pacientes CCR Biopsia Liquida.xlsx"

    SRC_CCR = "Dados Clínicos CCR"
    SRC_BXL = "Biópsia Líquida — Agenda (Entrada/Acompanhamento)"

    if fonte_dados == SRC_CCR:
        df_src = _load_excel_cached(PATH_CCR)
        if df_src.empty:
            st.error("A planilha de Dados Clínicos está vazia ou não pôde ser carregada.")
            st.stop()
        painel_clinico_ccr_bxl(df_src)

    elif fonte_dados == SRC_BXL:
        painel_agenda_bxl(PATH_BXL)

    else:
        st.warning("Fonte de dados inválida. Verifique as opções do selectbox.")


            
elif st.session_state.page == "biopsia":
    # =================== Entrada e Acompanhamento de Pacientes - Biópsia Líquida CCR ===================
    import os
    import re
    import time
    from datetime import datetime, timedelta
    import numpy as np
    import pandas as pd
    import streamlit as st

    st.title("Entrada e Acompanhamento de Pacientes - Biópsia Líquida CCR")

    FILE_PATH = "Entrada e Acompanhamento de Pacientes Biópsia Líquida CCR.xlsx"

    # -------- Utils de datas (robustos) --------
    def _excel_serial_to_ts(x):
        try:
            if pd.isna(x):
                return None
            if isinstance(x, (int, float)) and -100000 < x < 400000:
                # Base 1899-12-30 (padrão Excel)
                return (pd.to_datetime("1899-12-30") + pd.to_timedelta(int(x), unit="D")).to_pydatetime()
        except Exception:
            pass
        return None

    # aceita dd/mm/aa, dd-mm-aaaa, yyyy-mm-dd, yyyy/mm/dd
    _DATE_RE = re.compile(
        r"(\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b|\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b)"
    )

    def _parse_date_any(v):
        """Extrai a primeira data de uma célula (string com data + texto, ou número serial do Excel)."""
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return None
        # datetime/np.datetime
        try:
            if isinstance(v, (pd.Timestamp, datetime)):
                return pd.to_datetime(v).normalize().to_pydatetime()
        except Exception:
            pass
        # serial excel
        ts = _excel_serial_to_ts(v)
        if ts:
            return ts.replace(hour=0, minute=0, second=0, microsecond=0)
        # texto -> regex dd/mm/aaaa, dd/mm/aa, yyyy-mm-dd, yyyy/mm/dd
        s = str(v).strip()
        m = _DATE_RE.search(s)
        if not m:
            return None
        token = m.group(1)
        # normaliza separadores
        token = token.replace(".", "/").replace("-", "/")
        parts = token.split("/")
        if len(parts) != 3:
            return None

        # detecta formato ano-mês-dia vs dia-mês-ano
        if len(parts[0]) == 4:  # yyyy/mm/dd
            yy, mm, dd = parts
        else:  # dd/mm/yy(yy)
            dd, mm, yy = parts

        # yy curto -> século
        if len(yy) == 2:
            y = int(yy)
            yy = str(2000 + y if y <= 30 else 1900 + y)
        try:
            d = datetime(int(yy), int(mm), int(dd))
        except ValueError:
            return None
        return d

    # formato curto dd/mm/aa (2 dígitos)
    def _fmt_br(d):
        return d.strftime("%d/%m/%y") if isinstance(d, (datetime, pd.Timestamp)) else d

    # -------- helper centralizado de salvamento --------
    def _salvar_df_biopsia(show_toast: bool = True):
        """Persiste o df_biopsia em disco de forma segura."""
        try:
            st.session_state.df_biopsia.to_excel(FILE_PATH, index=False)
            if show_toast:
                st.success("Alterações salvas no arquivo Excel!")
        except Exception as e:
            st.error(f"Falha ao salvar no Excel: {e}")

    # =================== HELPERS: autosave + dirty tracking ===================
    def _df_hash(df_: pd.DataFrame) -> int:
        try:
            h = pd.util.hash_pandas_object(df_, index=True).sum()
            return int(h)
        except Exception:
            return hash(df_.to_csv(index=False))

    def _is_empty_cell(v) -> bool:
        if v is None:
            return True
        if isinstance(v, float) and np.isnan(v):
            return True
        s = str(v).strip()
        return (s == "" or s.lower() in {"nan", "none"})

    # >>> MANTIDO: reconhece "AUTO" em qualquer lugar (pra limpar legado antigo)
    def _is_auto_cell(v) -> bool:
        if v is None:
            return False
        s = str(v).upper()
        return re.search(r"\bAUTO\b", s) is not None

    # >>> NOVO: extrai só a parte da data (dd/mm/aa) se existir texto legado
    def _extract_date_str(val):
        if val is None:
            return None
        m = _DATE_RE.search(str(val))
        if not m:
            return None
        token = m.group(1).replace(".", "/").replace("-", "/")
        parts = token.split("/")
        if len(parts) != 3:
            return None
        # normaliza para dd/mm/aa
        if len(parts[0]) == 4:  # yyyy/mm/dd -> dd/mm/yy
            yy, mm, dd = parts
        else:  # dd/mm/yy(yy)
            dd, mm, yy = parts
        if len(yy) == 4:
            yy = yy[-2:]
        return f"{dd.zfill(2)}/{mm.zfill(2)}/{yy.zfill(2)}"

    # >>> MODIFICADO: automação agora retorna APENAS a data (dd/mm/aa), sem texto
    def _auto_text(d: datetime, suffix: str = "agendar coleta") -> str:
        return f"{_fmt_br(d)}".strip()

    def _mark_dirty():
        st.session_state["_biopsia_dirty"] = True

    def _autosave_if_dirty():
        """Salva automaticamente se houve alteração (com debounce)."""
        if not st.session_state.get("_biopsia_dirty", False):
            return

        now = time.time()
        last_save = st.session_state.get("_biopsia_last_save_ts", 0.0)
        if (now - last_save) < 0.8:
            return

        cur_hash = _df_hash(st.session_state.df_biopsia)
        last_hash = st.session_state.get("_biopsia_last_saved_hash", None)
        if last_hash is not None and cur_hash == last_hash:
            st.session_state["_biopsia_dirty"] = False
            return

        _salvar_df_biopsia(show_toast=False)
        st.session_state["_biopsia_last_save_ts"] = now
        st.session_state["_biopsia_last_saved_hash"] = cur_hash
        st.session_state["_biopsia_dirty"] = False

    # -------- Carregar DF na sessão --------
    if "df_biopsia" not in st.session_state:
        st.session_state.df_biopsia = pd.read_excel(FILE_PATH)

    if "_biopsia_last_saved_hash" not in st.session_state:
        st.session_state["_biopsia_last_saved_hash"] = _df_hash(st.session_state.df_biopsia)
    if "_biopsia_last_save_ts" not in st.session_state:
        st.session_state["_biopsia_last_save_ts"] = 0.0
    if "_biopsia_dirty" not in st.session_state:
        st.session_state["_biopsia_dirty"] = False

    df = st.session_state.df_biopsia

    # -------- Navegação --------
    if st.button("Voltar para o menu"):
        _autosave_if_dirty()
        st.session_state.page = "index"
        st.rerun()

    # -------- Tabela (esconde colunas totalmente vazias) --------
    df_visu = df.copy()
    colunas_vazias = [
        col for col in df_visu.columns
        if df_visu[col].isna().all() or (df_visu[col].astype(str) == "").all()
    ]
    df_visu = df_visu.drop(columns=colunas_vazias)

    st.subheader("Tabela de pacientes")

    def _highlight_excluidos(row):
        val = row.get("STATUS ESTUDO", "")
        if isinstance(val, str) and val.startswith("EXCLUÍDO DO ESTUDO"):
            return ['background-color: #fee2e2; color: #7f1d1d'] * len(row)
        return [''] * len(row)

    # >>> MANTIDO: se tiver legado com "AUTO", mostra só a data em azul.
    # (Se tudo estiver só data, não vai ter azul, só o verde do futuro.)
    def _cell_display(val):
        if _is_auto_cell(val):
            ds = _extract_date_str(val)
            return ds if ds else val
        return val

    def _cell_style(val):
        d = _parse_date_any(val)
        is_auto = _is_auto_cell(val)

        if is_auto:
            return "color: #1d4ed8; font-weight: 700;"
        if d and d.date() > datetime.now().date():
            return "background-color: #dcfce7; color: #166534; font-weight: 600;"
        return ""

    # aplica display (substitui texto mostrado) sem alterar df real
    df_show = df_visu.copy()
    for c in df_show.columns:
        df_show[c] = df_show[c].apply(_cell_display)

    if "STATUS ESTUDO" in df_show.columns:
        styler = df_show.style.apply(_highlight_excluidos, axis=1)
    else:
        styler = df_show.style

    styler = styler.applymap(_cell_style)

    st.dataframe(styler, use_container_width=True)

    # -------- Seleção do paciente --------
    st.markdown("---")
    st.subheader("Paciente selecionado")

    nomes_pacientes = df.iloc[:, 0].astype(str).tolist()
    paciente_selecionado = st.selectbox("Paciente", nomes_pacientes)
    idx = df[df.iloc[:, 0].astype(str) == str(paciente_selecionado)].index[0]
    dados_paciente = df.iloc[idx, 1:].to_dict()
    st.write(f"**Nome:** {paciente_selecionado}")

    # =================== Automação BxL baseada em CIRURGIA (datas fixas) ===================


    def add_months(dt, months):
        y = dt.year + (dt.month - 1 + months) // 12
        m = (dt.month - 1 + months) % 12 + 1
        d = min(dt.day, [31,
            29 if (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)) else 28,
            31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m-1])
        return dt.replace(year=y, month=m, day=d, hour=0, minute=0, second=0, microsecond=0)

    def _bxl_num(c):
        m = re.search(r"bxl\s*(\d+)", str(c).lower())
        return int(m.group(1)) if m else 0

    def _find_cirurgia_col(cols):
        for c in cols:
            nc = str(c).lower()
            if "cirurgia" in nc:
                return c
        return None

    bxl_cols = [c for c in df.columns if str(c).strip().upper().startswith("BXL")]
    bxl_cols = sorted(bxl_cols, key=_bxl_num)
    cirurgia_col = _find_cirurgia_col(df.columns)

    OFFSETS_MESES = {2: 1, 3: 6, 4: 12, 5: 18, 6: 24, 7: 30, 8: 36}

    sequencia = []
    if cirurgia_col:
        sequencia.append(("CIRURGIA", cirurgia_col, 0))
    for c in bxl_cols:
        n = _bxl_num(c)
        if 2 <= n <= 8:
            sequencia.append((f"BxL {n}", c, OFFSETS_MESES.get(n, None)))

    def _grade_alvo(row):
        base = _parse_date_any(row.get(cirurgia_col)) if cirurgia_col else None
        grade = {}
        if base:
            grade["CIRURGIA"] = base
            for rotulo, col, off in sequencia:
                if rotulo == "CIRURGIA":
                    continue
                if off is not None:
                    grade[rotulo] = add_months(base, off)
        return grade

    def _ultima_etapa_com_data(row):
        ultima_idx = -1
        ultima_rotulo = None
        for i, (rotulo, col, off) in enumerate(sequencia):
            dt = _parse_date_any(row.get(col))
            if dt:
                ultima_idx = i
                ultima_rotulo = rotulo
        return ultima_idx, ultima_rotulo

    def _proximas_calculadas(row):
        grade = _grade_alvo(row)
        ult_i, _ = _ultima_etapa_com_data(row)
        proximas = []
        if ult_i >= 0:
            for j in range(ult_i + 1, len(sequencia)):
                rotulo, col, off = sequencia[j]
                alvo = grade.get(rotulo)
                if alvo:
                    proximas.append((rotulo, col, alvo))
                else:
                    rot_ult, col_ult, off_ult = sequencia[ult_i]
                    d_ult = _parse_date_any(row.get(col_ult))
                    if d_ult and off is not None and off_ult is not None:
                        delta_m = off - off_ult
                        proximas.append((rotulo, col, add_months(d_ult, delta_m)))
        return proximas

    # =================== auto-preenchimento (APENAS DATA) ===================
    def _auto_preencher_datas_futuras(idx_patient: int) -> bool:
        changed = False
        row = st.session_state.df_biopsia.loc[idx_patient, :]

        grade = _grade_alvo(row)
        if grade:
            for rotulo, col, off in sequencia:
                if rotulo == "CIRURGIA":
                    continue
                alvo = grade.get(rotulo)
                if not alvo:
                    continue
                atual = st.session_state.df_biopsia.at[idx_patient, col]
                # Preenche se vazio OU se tinha legado com AUTO
                if _is_empty_cell(atual) or _is_auto_cell(atual):
                    novo_txt = _auto_text(alvo)  # <<< APENAS DATA
                    if str(atual).strip() != novo_txt:
                        st.session_state.df_biopsia.at[idx_patient, col] = novo_txt
                        changed = True
            return changed

        # fallback: ainda existe, mas grava só a data
        proximas = _proximas_calculadas(row)
        for rotulo, col, d in proximas:
            atual = st.session_state.df_biopsia.at[idx_patient, col]
            if _is_empty_cell(atual) or _is_auto_cell(atual):
                novo_txt = _auto_text(d)  # <<< APENAS DATA
                if str(atual).strip() != novo_txt:
                    st.session_state.df_biopsia.at[idx_patient, col] = novo_txt
                    changed = True

        return changed

    # ====== REMOVIDO: UI legado "Calcular próximas BxL's" + cache + botão agendar_next ======

    # -------- Bloco BxL (registro de ação) --------
    st.markdown("---")
    st.subheader("Registrar Ação BxL")
    
    import os, uuid, shutil
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter
    from datetime import datetime, date
    
    # ==============================
    # Config / estilos
    # ==============================
    FILE_PATH = FILE_PATH if "FILE_PATH" in globals() else "Entrada e Acompanhamento de Pacientes Biópsia Líquida CCR.xlsx"
    REG_SHEET = "REGISTRO_COLETAS"
    REG_HEADERS = ["CBL", "Coleta_short", "Data_iso", "Status", "Motivo", "Obs", "Updated_at", "Sheet", "Cell"]
    
    FILL_REG = PatternFill("solid", fgColor="C6EFCE")   # verde claro
    FONT_BOLD = Font(bold=True)
    
    def _ensure_writable_workbook_path(path: str) -> str:
        try:
            base_dir = os.path.dirname(path) or "."
            test = os.path.join(base_dir, f".__write_test__{uuid.uuid4().hex}.tmp")
            with open(test, "w", encoding="utf-8") as f:
                f.write("ok")
            os.remove(test)
            return path
        except Exception:
            base, ext = os.path.splitext(path)
            alt = f"{base}__EDITAVEL{ext}"
            try:
                shutil.copy2(path, alt)
                return alt
            except Exception:
                alt2 = f"REGISTRO__EDITAVEL__{uuid.uuid4().hex}.xlsx"
                shutil.copy2(path, alt2)
                return alt2
    
    def _atomic_save_workbook(wb, target_path: str):
        base_dir = os.path.dirname(target_path) or "."
        tmp_path = os.path.join(base_dir, f".__tmp__{uuid.uuid4().hex}.xlsx")
        wb.save(tmp_path)
        os.replace(tmp_path, target_path)
    
    def _coleta_short_label(coleta: str) -> str:
        import re
        c = str(coleta or "").strip().lower()
        m = re.search(r"\bbxl\s*(\d+)\b", c)
        if m:
            return f"BxL {int(m.group(1))}"
        if "pos" in c and "rt" in c:
            return "BxL pós-RT"
        if "cirurg" in c:
            return "Cirurgia"
        if "rm" in c:
            return "RM"
        if "tc" in c:
            return "TC"
        return str(coleta or "").strip()
    
    def _find_header_map(ws):
        """
        Mapa: nome_do_header (strip) -> índice da coluna (1-based)
        """
        hmap = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v is None:
                continue
            key = str(v).strip()
            if key and key not in hmap:
                hmap[key] = c
        return hmap
    
    def _find_row_by_cbl(ws, cbl_value: str, cbl_col_idx: int | None) -> int | None:
        """
        Procura a linha do Excel pelo CBL na coluna identificada.
        Retorna número da linha (2..max_row) ou None.
        """
        if not cbl_value:
            return None
        if not cbl_col_idx:
            return None
        target = str(cbl_value).strip()
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=cbl_col_idx).value
            if v is None:
                continue
            if str(v).strip() == target:
                return r
        return None
    
    def _upsert_registro_and_mark_cell(file_path: str, rec: dict, sheet_name: str, cell_ref: str, cell_text: str):
        """
        1) upsert em REGISTRO_COLETAS (chave: CBL+Coleta_short+Data_iso)
        2) marca a célula (fill verde + bold + comentário)
        3) mantém o valor do texto que você escreveu
        """
        file_path = _ensure_writable_workbook_path(file_path)
        wb = openpyxl.load_workbook(file_path)
    
        # --- registro
        if REG_SHEET not in wb.sheetnames:
            ws_reg = wb.create_sheet(REG_SHEET)
            ws_reg.append(REG_HEADERS)
        else:
            ws_reg = wb[REG_SHEET]
            if ws_reg.max_row < 1:
                ws_reg.append(REG_HEADERS)
    
        header_vals = [ws_reg.cell(row=1, column=i).value for i in range(1, ws_reg.max_column + 1)]
        header_vals = [str(x).strip() if x is not None else "" for x in header_vals]
        col_map = {h: i + 1 for i, h in enumerate(header_vals) if h}
    
        for h in REG_HEADERS:
            if h not in col_map:
                new_col = ws_reg.max_column + 1
                ws_reg.cell(row=1, column=new_col).value = h
                col_map[h] = new_col
    
        k_cbl = str(rec.get("CBL", "")).strip()
        k_col = str(rec.get("Coleta_short", "")).strip()
        k_dat = str(rec.get("Data_iso", "")).strip()
    
        target_row = None
        for r in range(2, ws_reg.max_row + 1):
            v1 = str(ws_reg.cell(row=r, column=col_map["CBL"]).value or "").strip()
            v2 = str(ws_reg.cell(row=r, column=col_map["Coleta_short"]).value or "").strip()
            v3 = str(ws_reg.cell(row=r, column=col_map["Data_iso"]).value or "").strip()
            if v1 == k_cbl and v2 == k_col and v3 == k_dat:
                target_row = r
                break
        if target_row is None:
            target_row = ws_reg.max_row + 1
    
        for h in REG_HEADERS:
            ws_reg.cell(row=target_row, column=col_map[h]).value = rec.get(h)
    
        # --- marca célula
        if sheet_name in wb.sheetnames and cell_ref:
            ws = wb[sheet_name]
            cell = ws[cell_ref]
            cell.value = cell_text  # garante que fica igual seu texto
            cell.fill = FILL_REG
            cell.font = FONT_BOLD
    
            try:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                txt = f"REGISTRADA | {now}\n{rec.get('Motivo','')}".strip()
                if rec.get("Obs"):
                    txt += f"\nObs: {str(rec['Obs']).strip()}"
                cell.comment = Comment(txt[:700], "AgendaBxL")
            except Exception:
                pass
    
        try:
            _atomic_save_workbook(wb, file_path)
            return file_path
        except PermissionError:
            base, ext = os.path.splitext(file_path)
            alt = f"{base}__EDITAVEL{ext}"
            try:
                shutil.copy2(file_path, alt)
            except Exception:
                alt = f"REGISTRO__EDITAVEL__{uuid.uuid4().hex}.xlsx"
            _atomic_save_workbook(wb, alt)
            return alt
    
    # ==============================
    # UI original (igual seu)
    # ==============================
    bxl_cols = [c for c in df.columns if str(c).strip().upper().startswith("BXL")]
    col_bxl = st.selectbox(
        "Selecione a coluna BxL para registrar a ação:",
        bxl_cols,
        key=f"bxlcol_{paciente_selecionado}_{idx}"
    )
    obs_bxl = st.text_input(
        "Observação (opcional)",
        key=f"obs_{paciente_selecionado}_{col_bxl}_{idx}"
    )
    
    cols_bxl = st.columns([2, 1])
    
    with cols_bxl[0]:
        if st.button("Registrar data de hoje na BxL selecionada", key=f"bxlbtn_{paciente_selecionado}_{col_bxl}_{idx}"):
    
            # 1) Texto na célula / df (como você já fazia)
            hoje_txt = datetime.now().strftime("%d/%m/%y")
            texto = f"{hoje_txt}"
            if obs_bxl.strip():
                texto += f" {obs_bxl.strip()}"
    
            st.session_state.df_biopsia.at[idx, col_bxl] = texto
    
            # 2) Atualiza registro + marca célula no Excel
            #    - tenta achar sheet atual a partir do seu state (se existir); senão usa a primeira aba.
            try:
                sheet_name = st.session_state.get("sheet_biopsia", None)
            except Exception:
                sheet_name = None
    
            wb_path = _ensure_writable_workbook_path(FILE_PATH)
            wb = openpyxl.load_workbook(wb_path)
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
    
            hmap = _find_header_map(ws)
    
            # tenta achar coluna CBL/ID do paciente
            cbl_col_idx = None
            for name, col_idx in hmap.items():
                n = name.strip().lower()
                if "cbl" in n or ("id" in n and "paciente" in n):
                    cbl_col_idx = col_idx
                    break
            if cbl_col_idx is None:
                cbl_col_idx = 1  # fallback
    
            # linha do excel:
            # 1) procura pelo CBL do paciente (mais confiável)
            excel_row = _find_row_by_cbl(ws, str(paciente_selecionado).strip(), cbl_col_idx)
            # 2) fallback para idx+2 (padrão: cabeçalho na linha 1)
            if excel_row is None:
                excel_row = int(idx) + 2
    
            # coluna do excel para a BxL selecionada:
            excel_col = hmap.get(str(col_bxl).strip(), None)
            if excel_col is None:
                # fallback: tenta achar por normalização
                target = str(col_bxl).strip().lower()
                for k, v in hmap.items():
                    if str(k).strip().lower() == target:
                        excel_col = v
                        break
    
            if excel_col is None:
                st.error("Não consegui mapear a coluna BxL no Excel (cabeçalho não encontrado).")
            else:
                cell_ref = get_column_letter(excel_col) + str(excel_row)
    
                rec = {
                    "CBL": str(paciente_selecionado).strip(),
                    "Coleta_short": _coleta_short_label(col_bxl),
                    "Data_iso": date.today().isoformat(),  # registro do dia (hoje)
                    "Status": "REGISTRADA",
                    "Motivo": "Manual: registrada na página de registro",
                    "Obs": obs_bxl.strip(),
                    "Updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Sheet": sheet_name,
                    "Cell": cell_ref,
                }
    
                new_path = _upsert_registro_and_mark_cell(wb_path, rec, sheet_name, cell_ref, texto)
                if new_path != wb_path:
                    FILE_PATH = new_path
                    st.warning(f"Excel bloqueado: continuei na cópia editável: {FILE_PATH}")
    
            # 3) mantém seu fluxo original
            if _auto_preencher_datas_futuras(idx):
                _mark_dirty()
            _mark_dirty()
            _autosave_if_dirty()
    
            st.success(f"Registrado '{texto}' para {paciente_selecionado} em '{col_bxl}'!")
            st.rerun()
    
    with cols_bxl[1]:
        st.caption("Observação opcional junto com a data.")
    
    # -------- Edição dos dados do paciente (AUTOSAVE SEM BOTÃO) --------
    st.markdown("---")
    st.subheader("Editar dados do paciente")
    st.caption("Qualquer alteração aqui é gravada automaticamente.")

    def _on_edit_field_change(col_name: str, widget_key: str, idx_patient: int):
        novo_val = st.session_state.get(widget_key, "")
        st.session_state.df_biopsia.at[idx_patient, col_name] = novo_val

        if ("bxl" in str(col_name).lower()) or ("cirurgia" in str(col_name).lower()):
            if _auto_preencher_datas_futuras(idx_patient):
                _mark_dirty()

        _mark_dirty()

    for coluna, valor in dados_paciente.items():
        wkey = f"edit_{coluna}_{paciente_selecionado}_{idx}_biopsia"
        st.text_input(
            coluna,
            value=str("" if pd.isna(valor) else valor),
            key=wkey,
            on_change=_on_edit_field_change,
            args=(coluna, wkey, idx)
        )

    # =================== Salvar na planilha (persistência manual – mantido, mas agora é opcional) ===================
    st.markdown("---")
    with st.expander("Opções avançadas de salvamento / backup", expanded=False):
        st.caption(f"Arquivo: `{FILE_PATH}` (será sobrescrito ao salvar).")
        if st.button("Salvar novamente no Excel (backup manual)", key="salvar_excel_manual"):
            _salvar_df_biopsia(show_toast=True)

    # -------- Retirar paciente do estudo (fluxo moderno em card) --------
    st.markdown("---")
    st.subheader("Retirar paciente do estudo")
    
    st.markdown("""
    <style>
    .excluir-card {
        padding: 1.25rem 1.5rem;
        margin-top: 0.5rem;
        border-radius: 0.9rem;
        background: #f9fafb;
        border: 1px solid #e5e7eb;
        box-shadow: 0 10px 25px rgba(15,23,42,0.10);
    }
    .excluir-header {
        font-weight: 600;
        margin-bottom: 0.25rem;
        color: #0f172a;
        display:flex;
        align-items:center;
        gap:0.5rem;
    }
    .badge-alerta {
        display:inline-block;
        padding: 0.15rem 0.6rem;
        font-size: 0.75rem;
        border-radius: 999px;
        background:#f97316;
        color:white;
    }
    .texto-discreto {
        font-size:0.85rem;
        color:#4b5563;
    }
    </style>
    """, unsafe_allow_html=True)
    
    flow_key = f"fluxo_exclusao_{idx}"
    if st.button("Retirar paciente do estudo", key=f"btn_iniciar_exclusao_{idx}"):
        st.session_state[flow_key] = True
    
    if st.session_state.get(flow_key, False):
    
        df_bio = st.session_state.df_biopsia
    
        nome_col = df_bio.columns[0]
        ident_str = str(df_bio.at[idx, nome_col]).strip()
    
        # >>> NOVO: data editável
        data_exclusao = st.date_input(
            "Data da exclusão",
            value=datetime.now().date(),
            key=f"data_exclusao_{idx}"
        )
        hoje_txt = data_exclusao.strftime("%d/%m/%y")
    
        MOTIVOS_EXCLUSAO = {
            "Faleceu": "✝",
            "Desistência": "🚪",
            "Sem amostra cirúrgica": "🧪🚫",
            "Não tem BxL 1": "🩸🚫",
        }
    
        st.markdown("<div class='excluir-card'>", unsafe_allow_html=True)
    
        st.markdown(
            "<div class='excluir-header'>Resumo do paciente "
            "<span class='badge-alerta'>ação definitiva</span></div>",
            unsafe_allow_html=True
        )
        st.markdown(
            f"<div class='texto-discreto'><b>Identificador (CBL/ID):</b> {ident_str}<br>"
            f"<b>Nome exibido:</b> {paciente_selecionado}</div>",
            unsafe_allow_html=True
        )
    
        st.divider()
    
        motivo_exclusao = st.selectbox(
            "Motivo da retirada do estudo",
            list(MOTIVOS_EXCLUSAO.keys()),
            key=f"motivo_exclusao_{idx}"
        )
    
        st.markdown(
            f"<div class='texto-discreto'>Você está prestes a retirar "
            f"<b>{ident_str}</b> do estudo por <b>{motivo_exclusao}</b> "
            f"na data de <b>{hoje_txt}</b>.</div>",
            unsafe_allow_html=True
        )
    
        confirm_flag = st.checkbox(
            "Confirmar retirada do estudo",
            key=f"chk_confirma_exclusao_{idx}"
        )
        st.markdown(
            "<span class='texto-discreto'>Ao confirmar, este paciente será marcado como "
            "EXCLUÍDO DO ESTUDO e não deve ser reativado sem justificativa formal.</span>",
            unsafe_allow_html=True
        )
    
        col_ok, col_cancel = st.columns(2)
    
        with col_ok:
            if st.button("Confirmar retirada do estudo", key=f"btn_confirmar_exclusao_{idx}"):
                if not confirm_flag:
                    st.error("Marque a opção 'Confirmar retirada do estudo' antes de continuar.")
                else:
                    icon = MOTIVOS_EXCLUSAO[motivo_exclusao]
                    status_txt = f"EXCLUÍDO DO ESTUDO – {motivo_exclusao} – {hoje_txt}"
    
                    if "STATUS ESTUDO" not in df_bio.columns:
                        df_bio["STATUS ESTUDO"] = ""
                    if "DATA EXCLUSÃO ESTUDO" not in df_bio.columns:
                        df_bio["DATA EXCLUSÃO ESTUDO"] = ""
                    if "MOTIVO EXCLUSÃO" not in df_bio.columns:
                        df_bio["MOTIVO EXCLUSÃO"] = ""
    
                    COLUNAS_PROTEGIDAS = {
                        "TCLE FOTO",
                        "TCLE",
                        "data TCLE",
                        "STATUS ESTUDO",
                        "DATA EXCLUSÃO ESTUDO",
                        "MOTIVO EXCLUSÃO",
                    }
    
                    for col in df_bio.columns[1:]:
                        if col in COLUNAS_PROTEGIDAS:
                            continue
                        val = df_bio.at[idx, col]
                        if pd.isna(val) or str(val).strip() == "":
                            df_bio.at[idx, col] = status_txt
    
                    df_bio.at[idx, "STATUS ESTUDO"] = status_txt
                    df_bio.at[idx, "DATA EXCLUSÃO ESTUDO"] = hoje_txt
                    df_bio.at[idx, "MOTIVO EXCLUSÃO"] = motivo_exclusao
    
                    nome_atual = str(df_bio.at[idx, nome_col])
                    if not nome_atual.startswith("⛔"):
                        df_bio.at[idx, nome_col] = f"⛔ {nome_atual} {icon}"
    
                    st.session_state[flow_key] = False
    
                    _mark_dirty()
                    _autosave_if_dirty()
    
                    st.success(f"Paciente {ident_str} retirado do estudo por {motivo_exclusao}.")
                    st.rerun()
    
        with col_cancel:
            if st.button("Cancelar operação", key=f"btn_cancelar_exclusao_{idx}"):
                st.session_state[flow_key] = False
                st.info("Retirada do estudo cancelada.")
    
        st.markdown("</div>", unsafe_allow_html=True)


    # -------- TCLE --------
    st.markdown("---")
    st.markdown("TCLE do paciente")
    PASTA_TCLE = "tcle_uploads"
    os.makedirs(PASTA_TCLE, exist_ok=True)

    if "TCLE FOTO" not in df.columns:
        df["TCLE FOTO"] = ""
    if "data TCLE" not in df.columns:
        df["data TCLE"] = ""

    tcle_path = df.at[idx, "TCLE FOTO"]

    def _tcle_valido(path: object) -> bool:
        if not isinstance(path, str):
            return False
        path = path.strip()
        if not path:
            return False
        if not path.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp")):
            return False
        return os.path.isfile(path)

    if _tcle_valido(tcle_path):
        if st.session_state.get(f"tentar_novo_tcle_biopsia_{idx}", False):
            st.error("Já existe um TCLE enviado para este paciente. Para substituir, remova o anterior primeiro.")
            if st.button("Cancelar envio de novo TCLE", key=f"cancelar_envio_biopsia_{idx}"):
                st.session_state[f"tentar_novo_tcle_biopsia_{idx}"] = False
        else:
            st.image(tcle_path, width=320, caption="TCLE já enviado anteriormente")
            st.info("Já existe um TCLE enviado para este paciente.")
            if st.button("Tentar enviar novo TCLE", key=f"novo_tcle_biopsia_{idx}"):
                st.session_state[f"tentar_novo_tcle_biopsia_{idx}"] = True

    else:
        if st.button("Enviar foto do TCLE", key=f"btn_tcle_biopsia_{idx}"):
            st.session_state[f"tcle_upload_biopsia_{idx}"] = True

        if st.session_state.get(f"tcle_upload_biopsia_{idx}", False):
            tcle_foto = st.file_uploader(
                "Selecione a foto do TCLE",
                type=["png", "jpg", "jpeg"],
                key=f"tcle_file_biopsia_{idx}"
            )
            if tcle_foto is not None:
                if st.button("Registrar novo TCLE", key=f"reg_tcle_biopsia_{idx}"):
                    caminho = os.path.join(
                        PASTA_TCLE,
                        f"tcle_{paciente_selecionado}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                    )
                    with open(caminho, "wb") as f:
                        f.write(tcle_foto.getbuffer())

                    st.session_state.df_biopsia.at[idx, "TCLE FOTO"] = caminho
                    st.session_state.df_biopsia.at[idx, "data TCLE"] = datetime.now().strftime("%d/%m/%y %H:%M")

                    _mark_dirty()
                    _autosave_if_dirty()

                    st.success("TCLE registrado com sucesso!")
                    st.session_state[f"tcle_upload_biopsia_{idx}"] = False
                    st.rerun()

    # -------- Remoção de paciente (mantido) --------
    def css_remover_minibutton():
        st.markdown("""
        <style>
        .btn-danger + div button{
            padding:.25rem .6rem; font-size:.8rem; border-radius:6px;
            background:#FFF1F2; color:#7F1D1D; border:1px solid #FECACA;
        }
        .btn-danger + div button:hover{ background:#FFE4E6; }
        </style>
        """, unsafe_allow_html=True)

    def remover_paciente(df_key: str, idx_rm: int):
        df_loc = st.session_state.get(df_key)
        if df_loc is None:
            st.error(f"DataFrame '{df_key}' não está carregado.")
            return
        st.session_state[df_key] = df_loc.drop(idx_rm).reset_index(drop=True)
        _mark_dirty()
        _autosave_if_dirty()
        st.success("Paciente removido!")
        st.rerun()

    def botao_remover_paciente(df_key: str, idx_rm: int, nome_exibicao: str, key_prefix: str = ""):
        css_remover_minibutton()
        flag_key = f"{key_prefix}confirmar_remove_{df_key}_{idx_rm}"
        st.markdown("<span class='btn-danger'></span>", unsafe_allow_html=True)
        if st.button("Remover paciente", key=f"{key_prefix}rm_{df_key}_{idx_rm}"):
            st.session_state[flag_key] = True
        if st.session_state.get(flag_key):
            st.warning(f"Tem certeza que deseja remover **{nome_exibicao}**? Essa ação é irreversível.")
            c1, c2 = st.columns(2)
            with c1:
                if st.button("Confirmar", key=f"{key_prefix}rm_yes_{df_key}_{idx_rm}"):
                    remover_paciente(df_key, idx_rm)
            with c2:
                if st.button("Cancelar", key=f"{key_prefix}rm_no_{df_key}_{idx_rm}"):
                    st.session_state[flag_key] = False

    # -------- Adicionar novo paciente (mantido + auto-preenchimento) --------
    st.markdown("---")
    st.subheader("Adicionar novo paciente")

    with st.form("novo_paciente_form_biopsia"):
        novo_nome = st.text_input("Nome do novo paciente", key="novo_nome_biopsia")
        novos_campos = {}
        for coluna in df.columns[1:]:
            if coluna not in ["TCLE", "TCLE FOTO", "TCLE"]:
                novos_campos[coluna] = st.text_input(coluna, key=f"novo_{coluna}_biopsia")

        novo_tcle_foto = st.file_uploader("Enviar foto do TCLE (obrigatório)", type=['png', 'jpg', 'jpeg'], key="novo_tcle_foto")
        submitted = st.form_submit_button("Adicionar")

        novo_tcle_enviado = False
        novo_caminho = ""
        if novo_tcle_foto is not None:
            PASTA_TCLE = "tcle_uploads"
            os.makedirs(PASTA_TCLE, exist_ok=True)
            novo_caminho = os.path.join(PASTA_TCLE, f"novo_tcle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg")
            with open(novo_caminho, "wb") as f:
                f.write(tcle_foto.getbuffer())
            novo_tcle_enviado = True

        if submitted:
            if not novo_nome:
                st.error("Preencha o nome do paciente!")
            elif not novo_tcle_enviado:
                st.error("É obrigatório enviar o TCLE para cadastrar!")
            else:
                d2 = _parse_date_any(novos_campos.get("BxL 2"))
                if d2:
                    def _bxl_num(c):
                        m = re.search(r"bxl\s*(\d+)", str(c).lower())
                        return int(m.group(1)) if m else 0

                    ordenadas = sorted([c for c in df.columns if str(c).strip().upper().startswith("BXL")], key=_bxl_num)

                    if "BxL 2" in ordenadas:
                        i2 = ordenadas.index("BxL 2")
                        if i2 + 1 < len(ordenadas):
                            prox_col = ordenadas[i2 + 1]
                            n_cur  = 2
                            n_prox = _bxl_num(prox_col)

                            off_cur  = OFFSETS_MESES.get(n_cur)
                            off_prox = OFFSETS_MESES.get(n_prox)

                            if off_cur is not None and off_prox is not None and not novos_campos.get(prox_col):
                                delta_meses = off_prox - off_cur
                                data_prox = add_months(d2, delta_meses)
                                # >>> MODIFICADO: APENAS DATA
                                novos_campos[prox_col] = _auto_text(data_prox)

                novos_campos["TCLE FOTO"] = novo_caminho
                novos_campos["TCLE"] = datetime.now().strftime("%d/%m/%y %H:%M")
                nova_linha = [novo_nome] + [novos_campos.get(col, "") for col in df.columns[1:]]
                st.session_state.df_biopsia.loc[len(st.session_state.df_biopsia)] = nova_linha

                _mark_dirty()
                idx_novo = len(st.session_state.df_biopsia) - 1
                if _auto_preencher_datas_futuras(idx_novo):
                    _mark_dirty()
                _autosave_if_dirty()

                st.success("Novo paciente adicionado!")
                st.rerun()

    # -------- autosave final --------
    _autosave_if_dirty()








elif st.session_state.page == "clinicos":
    st.title("Dados Clínicos - Pacientes CCR Biópsia Líquida")
    if st.button("Voltar para o menu"):
        st.session_state.page = "index"
        st.rerun()

    # Carrega a nova planilha, se ainda não estiver carregada
    if "df_clinicos" not in st.session_state:
        st.session_state.df_clinicos = pd.read_excel("Dados Clínicos Pacientes CCR Biopsia Liquida.xlsx")
    df = st.session_state.df_clinicos

    st.subheader("Tabela de dados clínicos")
    st.dataframe(df)
    # --- CAMPOS PRINCIPAIS: ajuste se os nomes das colunas forem diferentes na sua planilha! ---
    CAMPOS_PRINCIPAIS = ['CBL', 'TCLE', 'prontuário', 'Nome', 'Data de Nasc']

# Seleção do paciente
    nomes_pacientes = df['CBL'].astype(str).tolist()
    paciente_selecionado = st.selectbox('Selecione o paciente', nomes_pacientes)
    idx = df[df['CBL'].astype(str) == paciente_selecionado].index[0]
    dados_paciente = df.loc[idx].to_dict()



    
    st.subheader("Dados do paciente (principais)")
    campos_novos = {}
    
    # Inputs principais
    for campo in CAMPOS_PRINCIPAIS:
        if campo == "TCLE":
            continue
        valor = dados_paciente.get(campo, "")
        campos_novos[campo] = st.text_input(campo, value=str(valor), key=f"{campo}_{idx}")

    # --- REGISTRO TCLE (OBRIGATÓRIO FOTO) ---
    if 'TCLE FOTO' not in df.columns:
        df['TCLE FOTO'] = ""


    PASTA_TCLE = "tcle_uploads"
    os.makedirs(PASTA_TCLE, exist_ok=True)
    
    tcle_path = dados_paciente.get('TCLE FOTO', '')
    
    # Lógica para controlar tentativa de novo envio
    if isinstance(tcle_path, str) and tcle_path:
        # Se o usuário clicou para enviar novo TCLE mesmo já existindo
        if st.session_state.get(f"tentar_novo_tcle_{idx}", False):
            st.error("Já existe um TCLE enviado para este paciente. Se quiser substituir, remova o anterior primeiro.")
            # Botão para voltar para visualização normal
            if st.button("Cancelar envio de novo TCLE", key=f"cancelar_envio_{idx}"):
                st.session_state[f"tentar_novo_tcle_{idx}"] = False
        else:
            st.image(tcle_path, width=320, caption="TCLE já enviado anteriormente")
            st.info("Já existe um TCLE enviado para este paciente.")
            if st.button("Tentar enviar novo TCLE", key=f"novo_tcle_{idx}"):
                st.session_state[f"tentar_novo_tcle_{idx}"] = True
    else:
        if st.button("Enviar foto do TCLE", key=f"btn_tcle_{idx}"):
            st.session_state[f"tcle_upload_{idx}"] = True
    
        if st.session_state.get(f"tcle_upload_{idx}", False):
            tcle_foto = st.file_uploader("Selecione a foto do TCLE", type=["png", "jpg", "jpeg"], key=f"tcle_file_{idx}")
            if tcle_foto is not None:
                if st.button("Registrar novo TCLE", key=f"reg_tcle_{idx}"):
                    caminho = os.path.join(
                        PASTA_TCLE,
                        f"tcle_{paciente_selecionado}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                    )
                    with open(caminho, "wb") as f:
                        f.write(tcle_foto.getbuffer())
                    st.session_state.df_clinicos.at[idx, 'TCLE FOTO'] = caminho
                    st.session_state.df_clinicos.at[idx, 'TCLE'] = datetime.now().strftime("%d/%m/%Y %H:%M")
                    st.success("TCLE registrado com sucesso!")
                    st.session_state[f"tcle_upload_{idx}"] = False
                    st.rerun()



    

    # --- OUTROS DADOS ---
    st.markdown("#### Outros dados")
    CAMPOS_OCULTAR = set(CAMPOS_PRINCIPAIS + ["TCLE FOTO", "TCLE DATA ENVIO"])
    outros_campos = [col for col in df.columns if col not in CAMPOS_OCULTAR]
    
    campos_escolhidos = st.multiselect(
        "Quais outros dados deseja editar/registrar?", 
        outros_campos,
        key=f"outros_campos_{idx}")
    
    for campo in campos_escolhidos:
        valor = dados_paciente.get(campo, "")
        campos_novos[campo] = st.text_input(campo, value=str(valor), key=f"{campo}_{idx}_extra")
    
    # --- BOTÃO DE SALVAR ---
    if st.button("Salvar alterações do paciente", key=f"salvar_paciente_{idx}"):
            for campo, valor in campos_novos.items():
                st.session_state.df_clinicos.at[idx, campo] = valor
            st.success("Dados do paciente atualizados!")
    
    # --- CADASTRAR NOVO PACIENTE ---
    with st.expander("Cadastrar novo paciente"):
        st.markdown("#### Preencher campos principais")
        novos_campos = {}
        for campo in CAMPOS_PRINCIPAIS:
            if campo == "TCLE":
                continue
            novos_campos[campo] = st.text_input(f"[Novo] {campo}", key=f"novo_{campo}")
    
        novo_tcle_foto = st.file_uploader("Enviar foto do TCLE (obrigatório)", type=['png', 'jpg', 'jpeg'], key="novo_tcle_foto")
        novo_tcle_enviado = False
        novo_caminho = ""
        if novo_tcle_foto is not None:
            os.makedirs(PASTA_TCLE, exist_ok=True)
            novo_caminho = os.path.join(PASTA_TCLE, f"novo_tcle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg")
            with open(novo_caminho, "wb") as f:
                f.write(novo_tcle_foto.getbuffer())
            novo_tcle_enviado = True
    
        st.markdown("#### Outros dados do novo paciente")
        outros_novos = st.multiselect(
            "Quais outros dados deseja registrar?", 
            outros_campos,
            key="novo_outros_campos")
        for campo in outros_novos:
            novos_campos[campo] = st.text_input(f"[Novo] {campo}", key=f"novo_{campo}_extra")
    
        if st.button("Adicionar novo paciente"):
            if not novo_tcle_enviado:
                st.error("É obrigatório enviar o TCLE para cadastrar!")
            else:
                nova_linha = {**novos_campos, "TCLE FOTO": novo_caminho, "TCLE DATA ENVIO": datetime.now().strftime("%d/%m/%Y %H:%M")}
                for col in df.columns:
                    if col not in nova_linha:
                        nova_linha[col] = ""
                st.session_state.df_clinicos.loc[len(df)] = nova_linha
                st.success("Novo paciente cadastrado!")

    
































