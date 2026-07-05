# apps/internal/views.py
import os
from datetime import datetime, date

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from apps.bookings.email_utils import send_booking_cancelled, send_booking_decision
from apps.bookings.models import BookingAuditLog, EquipmentBooking, SCHEDULABLE_EQUIPMENT

from . import biopsia_logic as L

try:
    import numpy as np
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


# ─────────────────────────────────────────────
# Caminhos das planilhas
# ─────────────────────────────────────────────
BIOPSIA_FILE = "Entrada e Acompanhamento de Pacientes Biópsia Líquida CCR.xlsx"
CLINICOS_FILE = "Dados Clínicos Pacientes CCR Biopsia Liquida.xlsx"
REG_SHEET = "REGISTRO_COLETAS"
REG_HEADERS = ["CBL", "Coleta_short", "Data_iso", "Status", "Motivo", "Obs", "Updated_at", "Sheet", "Cell"]


def _biopsia_path():
    return settings.BASE_DIR / BIOPSIA_FILE


def _clinicos_path():
    return settings.BASE_DIR / CLINICOS_FILE


def _load_df(path):
    df = pd.read_excel(path)
    df.columns = [L.norm_header(c) for c in df.columns]
    # A planilha é de texto livre (datas gravadas como texto, observações etc.).
    # Convertendo para `object` evitamos erros de dtype ao gravar strings em
    # colunas que o pandas inferiu como numéricas (ex.: Prontuário/Telefone).
    return df.astype(object)


def _save_df(df, path):
    df.to_excel(path, index=False)


# ═════════════════════════════════════════════════════════════════════
# PAINEL PRINCIPAL (INDEX)
# ═════════════════════════════════════════════════════════════════════
@login_required
def index(request):
    pending_count = EquipmentBooking.objects.filter(status="pending").count()
    return render(request, "internal/index.html", {
        "user": request.user,
        "pending_count": pending_count,
    })


# ═════════════════════════════════════════════════════════════════════
# DASHBOARD ANALÍTICO (Painel Clínico + Agenda BxL)  — inalterado
# ═════════════════════════════════════════════════════════════════════
@login_required
def dashboard(request):
    from . import dashboard as dash

    ctx = {"user": request.user, "libs_ok": dash.HAS_LIBS}
    if not dash.HAS_LIBS:
        ctx["error"] = "Bibliotecas de análise (pandas/plotly) não instaladas."
        return render(request, "internal/dashboard.html", ctx)

    SRC_CCR = "Dados Clínicos CCR"
    SRC_BXL = "Biópsia Líquida — Agenda (Entrada/Acompanhamento)"
    PATH_CCR = _clinicos_path()
    PATH_BXL = _biopsia_path()

    fonte = request.GET.get("fonte", SRC_CCR)
    ctx["fonte"] = fonte
    ctx["SRC_CCR"] = SRC_CCR
    ctx["SRC_BXL"] = SRC_BXL

    if fonte == SRC_BXL:
        if not PATH_BXL.exists():
            ctx["error"] = f"Arquivo não encontrado: {PATH_BXL.name}"
        else:
            filters = {
                "sheet": request.GET.get("sheet"),
                "segmento": request.GET.get("segmento", "Todos"),
                "paciente": request.GET.get("paciente", "Todos"),
                "hide_events": request.GET.get("hide_events", "1") == "1",
                "cbl_search": request.GET.get("cbl_search", ""),
            }
            ctx["agenda"] = dash.build_painel_agenda(str(PATH_BXL), filters)
    else:
        if not PATH_CCR.exists():
            ctx["error"] = f"Arquivo não encontrado: {PATH_CCR.name}"
        else:
            import pandas as _pd
            try:
                df0 = _pd.read_excel(PATH_CCR)
                df0.columns = [str(c).strip() for c in df0.columns]
            except Exception as exc:
                ctx["error"] = f"Não foi possível ler a planilha: {exc}"
                return render(request, "internal/dashboard.html", ctx)
            filters = {
                "estadiamento": request.GET.getlist("estadiamento"),
                "sexo": request.GET.getlist("sexo"),
                "asa": request.GET.getlist("asa"),
                "dist_var": request.GET.get("dist_var"),
                "include_dates": request.GET.get("include_dates") == "1",
            }
            ctx["clinico"] = dash.build_painel_clinico(df0, filters)

    return render(request, "internal/dashboard.html", ctx)


# ═════════════════════════════════════════════════════════════════════
# GERENCIAR AGENDAMENTOS  — inalterado
# ═════════════════════════════════════════════════════════════════════
@login_required
def manage_bookings(request):
    all_bk = EquipmentBooking.objects.all().order_by("-created_at")
    pending = all_bk.filter(status="pending")

    eq_filter = request.GET.get("equipment", "")
    st_filter = request.GET.get("status", "")
    date_filter = request.GET.get("date", "")

    filtered = all_bk
    if eq_filter:
        filtered = filtered.filter(equipment_name=eq_filter)
    if st_filter:
        if st_filter == "approved":
            filtered = filtered.filter(status__in=["approved", "confirmed"])
        else:
            filtered = filtered.filter(status=st_filter)
    if date_filter:
        filtered = filtered.filter(booking_date=date_filter)

    paginator = Paginator(filtered, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    audit_logs = BookingAuditLog.objects.select_related("booking").order_by("-created_at")[:200]

    return render(request, "internal/manage_bookings.html", {
        "pending": pending,
        "page_obj": page_obj,
        "audit_logs": audit_logs,
        "equipment_list": SCHEDULABLE_EQUIPMENT,
        "eq_filter": eq_filter,
        "st_filter": st_filter,
        "date_filter": date_filter,
        "total": all_bk.count(),
        "n_pending": pending.count(),
        "n_approved": all_bk.filter(status__in=["approved", "confirmed"]).count(),
        "n_rejected": all_bk.filter(status="rejected").count(),
        "n_cancelled": all_bk.filter(status="cancelled").count(),
    })


@login_required
@require_http_methods(["POST"])
def approve_booking(request, booking_id: int):
    try:
        booking = EquipmentBooking.objects.get(pk=booking_id)
    except EquipmentBooking.DoesNotExist:
        messages.error(request, "Reserva não encontrada.")
        return redirect("internal:manage_bookings")

    review_note = request.POST.get("review_note", "").strip()
    booking.status = "approved"
    booking.reviewed_by = request.user.nome
    booking.reviewed_at = timezone.now()
    booking.review_note = review_note
    booking.save()

    BookingAuditLog.objects.create(
        booking=booking, action="approved",
        performed_by=f"internal:{request.user.nome}",
        note=review_note or "Approved.",
    )
    result = send_booking_decision(booking, "approved", review_note)
    ok_msg = "E-mail enviado." if result.get("user_sent") else f"E-mail não enviado ({result.get('user_error','?')})."
    messages.success(request, f"Reserva #{booking_id} aprovada. {ok_msg}")
    return redirect("internal:manage_bookings")


@login_required
@require_http_methods(["POST"])
def reject_booking(request, booking_id: int):
    try:
        booking = EquipmentBooking.objects.get(pk=booking_id)
    except EquipmentBooking.DoesNotExist:
        messages.error(request, "Reserva não encontrada.")
        return redirect("internal:manage_bookings")

    review_note = request.POST.get("review_note", "").strip()
    booking.status = "rejected"
    booking.reviewed_by = request.user.nome
    booking.reviewed_at = timezone.now()
    booking.review_note = review_note
    booking.save()

    BookingAuditLog.objects.create(
        booking=booking, action="rejected",
        performed_by=f"internal:{request.user.nome}",
        note=review_note or "Rejected.",
    )
    result = send_booking_decision(booking, "rejected", review_note)
    ok_msg = "E-mail enviado." if result.get("user_sent") else f"E-mail não enviado ({result.get('user_error','?')})."
    messages.success(request, f"Reserva #{booking_id} rejeitada. {ok_msg}")
    return redirect("internal:manage_bookings")


@login_required
@require_http_methods(["POST"])
def cancel_booking_internal(request, booking_id: int):
    try:
        booking = EquipmentBooking.objects.get(pk=booking_id)
    except EquipmentBooking.DoesNotExist:
        messages.error(request, "Reserva não encontrada.")
        return redirect("internal:manage_bookings")

    booking.status = "cancelled"
    booking.cancelled_by = f"internal:{request.user.nome}"
    booking.cancelled_at = timezone.now()
    booking.save()

    BookingAuditLog.objects.create(
        booking=booking, action="cancelled",
        performed_by=f"internal:{request.user.nome}",
        note="Cancelled by lab staff.",
    )
    send_booking_cancelled(booking, cancelled_by_label=f"equipe do laboratório ({request.user.nome})")
    messages.success(request, f"Reserva #{booking_id} cancelada.")
    return redirect("internal:manage_bookings")


# ═════════════════════════════════════════════════════════════════════
# BIÓPSIA LÍQUIDA — Entrada e Acompanhamento de Pacientes
# Reproduz a página `biopsia` do app Streamlit (acompanhar pacientes,
# agendar coletas, automação BxL, retirada do estudo, TCLE).
# ═════════════════════════════════════════════════════════════════════
def _tcle_valido(path) -> bool:
    if not isinstance(path, str):
        return False
    path = path.strip()
    if not path:
        return False
    if not path.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp")):
        return False
    return os.path.isfile(path)


def _tcle_url(path):
    """Converte um caminho de arquivo salvo em URL /media/ quando possível."""
    if not _tcle_valido(path):
        return None
    try:
        p = os.path.abspath(path)
        media_root = os.path.abspath(str(settings.MEDIA_ROOT))
        if p.startswith(media_root):
            rel = os.path.relpath(p, media_root).replace(os.sep, "/")
            return settings.MEDIA_URL + rel
    except Exception:
        pass
    return None


def _save_tcle_file(uploaded, cbl_hint=""):
    """Salva a foto do TCLE em MEDIA/tcle_uploads e devolve o caminho absoluto."""
    upload_dir = settings.TCLE_UPLOAD_DIR
    os.makedirs(upload_dir, exist_ok=True)
    slug = L.normalize_cbl_text(cbl_hint) or "novo"
    filename = f"tcle_{slug}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
    filepath = os.path.join(str(upload_dir), filename)
    with open(filepath, "wb") as f:
        for chunk in uploaded.chunks():
            f.write(chunk)
    return filepath


def _registrar_coleta_openpyxl(path, rec, cell_text, id_value, col_bxl):
    """
    Marca a célula da coleta na planilha (verde/negrito + comentário) e faz
    upsert na aba REGISTRO_COLETAS. Espelha _upsert_registro_and_mark_cell
    do Streamlit. Falhas aqui não interrompem o salvamento principal.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        from openpyxl.comments import Comment
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    try:
        wb = openpyxl.load_workbook(path)
        ws = wb[wb.sheetnames[0]]

        # mapa de cabeçalhos
        hmap = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v is not None:
                key = L.norm_header(v)
                hmap.setdefault(key, c)

        # coluna do identificador
        cbl_col_idx = 1
        for name, ci in hmap.items():
            n = str(name).strip().lower()
            if "cbl" in n or ("id" in n and "paciente" in n):
                cbl_col_idx = ci
                break

        # linha pelo identificador
        alvo = L.normalize_cbl_text(id_value)
        excel_row = None
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=cbl_col_idx).value
            if v is not None and L.normalize_cbl_text(v) == alvo:
                excel_row = r
                break
        if excel_row is None:
            return None

        excel_col = hmap.get(L.norm_header(col_bxl))
        if excel_col is None:
            return None

        cell = ws.cell(row=excel_row, column=excel_col)
        cell.value = cell_text
        cell.number_format = "@"
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.font = Font(bold=True)
        try:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            txt = f"REGISTRADA | {now}\n{rec.get('Motivo','')}".strip()
            if rec.get("Obs"):
                txt += f"\nObs: {str(rec['Obs']).strip()}"
            cell.comment = Comment(txt[:700], "AgendaBxL")
        except Exception:
            pass

        rec = dict(rec)
        rec["Cell"] = get_column_letter(excel_col) + str(excel_row)

        # upsert REGISTRO_COLETAS
        if REG_SHEET not in wb.sheetnames:
            ws_reg = wb.create_sheet(REG_SHEET)
            ws_reg.append(REG_HEADERS)
        else:
            ws_reg = wb[REG_SHEET]
            if ws_reg.max_row < 1:
                ws_reg.append(REG_HEADERS)
        header_vals = [str(ws_reg.cell(row=1, column=i).value or "").strip()
                       for i in range(1, ws_reg.max_column + 1)]
        col_map = {h: i + 1 for i, h in enumerate(header_vals) if h}
        for h in REG_HEADERS:
            if h not in col_map:
                nc = ws_reg.max_column + 1
                ws_reg.cell(row=1, column=nc).value = h
                col_map[h] = nc
        target_row = None
        for r in range(2, ws_reg.max_row + 1):
            v1 = str(ws_reg.cell(row=r, column=col_map["CBL"]).value or "").strip()
            v2 = str(ws_reg.cell(row=r, column=col_map["Coleta_short"]).value or "").strip()
            v3 = str(ws_reg.cell(row=r, column=col_map["Data_iso"]).value or "").strip()
            if v1 == str(rec.get("CBL", "")).strip() and v2 == str(rec.get("Coleta_short", "")).strip() \
               and v3 == str(rec.get("Data_iso", "")).strip():
                target_row = r
                break
        if target_row is None:
            target_row = ws_reg.max_row + 1
        for h in REG_HEADERS:
            ws_reg.cell(row=target_row, column=col_map[h]).value = rec.get(h)

        wb.save(path)
        return rec["Cell"]
    except Exception:
        return None


@login_required
def biopsia(request):
    path = _biopsia_path()
    ctx = {"user": request.user, "file_exists": path.exists()}

    if not HAS_PANDAS:
        ctx["error"] = "Pandas não está instalado."
        return render(request, "internal/biopsia.html", ctx)
    if not path.exists():
        ctx["error"] = f"Arquivo não encontrado: {path.name}"
        return render(request, "internal/biopsia.html", ctx)

    try:
        df = _load_df(path)
    except Exception as exc:
        ctx["error"] = f"Erro ao ler o arquivo: {exc}"
        return render(request, "internal/biopsia.html", ctx)

    id_col = df.columns[0]

    # ─────────── POST (ações) ───────────
    if request.method == "POST":
        action = request.POST.get("action")
        idx = request.POST.get("idx")
        try:
            idx = int(idx) if idx is not None else None
        except (TypeError, ValueError):
            idx = None

        # ---- salvar edição dos campos do paciente ----
        if action == "save_row" and idx is not None and idx in df.index:
            for col in df.columns:
                key = f"col_{col}"
                if key in request.POST:
                    df.at[idx, col] = request.POST[key]
            L.autofill_datas_futuras(df, idx)   # BxL automático após editar
            try:
                _save_df(df, path)
                messages.success(request, "Paciente atualizado com sucesso!")
            except Exception as exc:
                messages.error(request, f"Erro ao salvar: {exc}")
            return redirect(f"{request.path}?sel={idx}#edit-section")

        # ---- registrar coleta BxL (agendar/registrar data de hoje) ----
        elif action == "register_bxl" and idx is not None and idx in df.index:
            col_bxl = request.POST.get("col_bxl", "")
            obs = request.POST.get("obs_bxl", "").strip()
            data_str = request.POST.get("data_coleta", "").strip()
            if not col_bxl or col_bxl not in df.columns:
                messages.error(request, "Selecione uma coluna BxL válida.")
                return redirect(f"{request.path}?sel={idx}#registrar")
            d = L.parse_date_any(data_str) if data_str else datetime.now()
            if d is None:
                d = datetime.now()
            data_txt = d.strftime("%d/%m/%y")
            texto = f"{data_txt} (coletado)"
            if obs:
                texto += f" {obs}"
            df.at[idx, col_bxl] = texto
            L.autofill_datas_futuras(df, idx)
            try:
                _save_df(df, path)
            except Exception as exc:
                messages.error(request, f"Erro ao salvar: {exc}")
                return redirect(f"{request.path}?sel={idx}#registrar")

            rec = {
                "CBL": L.normalize_cbl_text(str(df.at[idx, id_col])),
                "Coleta_short": L.coleta_short_label(col_bxl),
                "Data_iso": d.date().isoformat(),
                "Status": "REGISTRADA",
                "Motivo": "Manual: registrada na página de registro",
                "Obs": obs,
                "Updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Sheet": "",
                "Cell": "",
            }
            _registrar_coleta_openpyxl(str(path), rec, texto, str(df.at[idx, id_col]), col_bxl)
            messages.success(request, f"Coleta registrada: '{texto}' em '{L.norm_header(col_bxl)}'.")
            return redirect(f"{request.path}?sel={idx}#registrar")

        # ---- preencher datas futuras automaticamente ----
        elif action == "autofill" and idx is not None and idx in df.index:
            changed = L.autofill_datas_futuras(df, idx)
            if changed:
                try:
                    _save_df(df, path)
                    messages.success(request, "Datas futuras de coleta preenchidas automaticamente.")
                except Exception as exc:
                    messages.error(request, f"Erro ao salvar: {exc}")
            else:
                messages.info(request, "Nenhuma data para preencher (defina a data de CIRURGIA primeiro).")
            return redirect(f"{request.path}?sel={idx}#registrar")

        # ---- recalcular o cronograma de TODOS os pacientes ----
        elif action == "recompute_all":
            n = L.recompute_all_schedules(df)
            if n:
                try:
                    _save_df(df, path)
                    messages.success(request, f"Cronograma recalculado para {n} paciente(s).")
                except Exception as exc:
                    messages.error(request, f"Erro ao salvar: {exc}")
            else:
                messages.info(request, "Todos os cronogramas já estavam atualizados.")
            return redirect("internal:biopsia")

        # ---- retirar paciente do estudo ----
        elif action == "exclude_patient" and idx is not None and idx in df.index:
            motivo = request.POST.get("motivo", "Outro")
            data_str = request.POST.get("data_exclusao", "").strip()
            d = L.parse_date_any(data_str) if data_str else datetime.now()
            if d is None:
                d = datetime.now()
            hoje_txt = d.strftime("%d/%m/%y")
            ident_str = str(df.at[idx, id_col]).strip()
            novo_id = L.padronizar_id_exclusao(ident_str, motivo)
            status_txt = f"EXCLUÍDO DO ESTUDO – {motivo} – {hoje_txt}"

            for col in ("STATUS ESTUDO", "DATA EXCLUSÃO ESTUDO", "MOTIVO EXCLUSÃO"):
                if col not in df.columns:
                    df[col] = ""

            protegidas = set(L.COLUNAS_PROTEGIDAS_EXCLUSAO) | {id_col}
            for col in df.columns[1:]:
                if col in protegidas:
                    continue
                val = df.at[idx, col]
                if pd.isna(val) or str(val).strip() == "":
                    df.at[idx, col] = status_txt

            df.at[idx, "STATUS ESTUDO"] = status_txt
            df.at[idx, "DATA EXCLUSÃO ESTUDO"] = hoje_txt
            df.at[idx, "MOTIVO EXCLUSÃO"] = motivo
            df.at[idx, id_col] = novo_id
            try:
                _save_df(df, path)
                messages.success(request, f"Paciente {ident_str} retirado do estudo por {motivo}. "
                                          f"Novo identificador: {novo_id}")
            except Exception as exc:
                messages.error(request, f"Erro ao salvar: {exc}")
            return redirect(f"{request.path}?sel={idx}#edit-section")

        # ---- remover paciente (excluir linha) ----
        elif action == "remove_patient" and idx is not None and idx in df.index:
            df = df.drop(idx).reset_index(drop=True)
            try:
                _save_df(df, path)
                messages.success(request, "Paciente removido!")
            except Exception as exc:
                messages.error(request, f"Erro ao salvar: {exc}")
            return redirect("internal:biopsia")

        # ---- upload de TCLE ----
        elif action == "upload_tcle" and idx is not None and idx in df.index:
            tcle_file = request.FILES.get("tcle_foto")
            if not tcle_file:
                messages.error(request, "Selecione uma foto do TCLE.")
                return redirect(f"{request.path}?sel={idx}#tcle")
            if "TCLE FOTO" not in df.columns:
                df["TCLE FOTO"] = ""
            if "data TCLE" not in df.columns:
                df["data TCLE"] = ""
            filepath = _save_tcle_file(tcle_file, str(df.at[idx, id_col]))
            df.at[idx, "TCLE FOTO"] = filepath
            df.at[idx, "data TCLE"] = datetime.now().strftime("%d/%m/%y %H:%M")
            try:
                _save_df(df, path)
                messages.success(request, "TCLE registrado com sucesso!")
            except Exception as exc:
                messages.error(request, f"Erro ao salvar: {exc}")
            return redirect(f"{request.path}?sel={idx}#tcle")

        # ---- adicionar novo paciente ----
        elif action == "add_patient":
            novo_nome = request.POST.get("new_id", "").strip()
            tcle_file = request.FILES.get("tcle_foto")
            if not novo_nome:
                messages.error(request, "Preencha o identificador do paciente.")
                return redirect(f"{request.path}#novo")
            if not tcle_file:
                messages.error(request, "É obrigatório enviar o TCLE para cadastrar.")
                return redirect(f"{request.path}#novo")

            nova = {c: request.POST.get(f"new_{c}", "") for c in df.columns[1:]
                    if c not in ("TCLE", "TCLE FOTO")}
            nova[id_col] = novo_nome
            filepath = _save_tcle_file(tcle_file, novo_nome)
            if "TCLE FOTO" not in df.columns:
                df["TCLE FOTO"] = ""
            if "data TCLE" not in df.columns:
                df["data TCLE"] = ""
            nova["TCLE FOTO"] = filepath
            nova["data TCLE"] = datetime.now().strftime("%d/%m/%y %H:%M")

            linha = {c: nova.get(c, "") for c in df.columns}
            df.loc[len(df)] = linha
            new_idx = len(df) - 1
            L.autofill_datas_futuras(df, new_idx)
            try:
                _save_df(df, path)
                messages.success(request, "Novo paciente adicionado!")
            except Exception as exc:
                messages.error(request, f"Erro ao salvar: {exc}")
            return redirect(f"{request.path}?sel={new_idx}#edit-section")

    # ─────────── GET (render) ───────────
    # Recalcula automaticamente o cronograma de TODOS os pacientes (cirurgia →
    # BxL 7), preenchendo coletas futuras e ajustando as atrasadas (meio-termo).
    # Só grava no Excel se algo mudou, evitando reescrita desnecessária.
    try:
        if L.recompute_all_schedules(df):
            _save_df(df, path)
    except Exception:
        pass

    # colunas totalmente vazias são ocultadas na tabela (como no Streamlit)
    colunas_vazias = [
        col for col in df.columns
        if df[col].isna().all() or (df[col].astype(str).str.strip() == "").all()
    ]
    display_cols = [c for c in df.columns if c not in colunas_vazias]

    display_df = df.head(120)
    rows = []
    for i, row in display_df.iterrows():
        cells = []
        for col in display_cols:
            raw = row[col]
            cells.append({
                "value": "" if pd.isna(raw) else L.cell_display(raw),
                "cls": L.cell_class(raw),
            })
        rows.append({
            "idx": i,
            "id_value": "" if pd.isna(row[id_col]) else row[id_col],
            "is_event": L.is_event_id(row[id_col]),
            "cells": cells,
        })

    # paciente selecionado
    sel_idx = request.GET.get("sel")
    selected_idx = None
    selected_patient = None
    selected_fields = []       # [(col, value)]
    tcle_url = None
    tcle_exists = False
    seq_info = []
    if sel_idx is not None:
        try:
            selected_idx = int(sel_idx)
            if selected_idx in df.index:
                srow = df.loc[selected_idx]
                selected_patient = srow[id_col]
                for col in df.columns:
                    v = srow[col]
                    selected_fields.append((col, "" if pd.isna(v) else str(v)))
                tp = srow.get("TCLE FOTO", "") if "TCLE FOTO" in df.columns else ""
                tcle_exists = _tcle_valido(tp if isinstance(tp, str) else "")
                tcle_url = _tcle_url(tp if isinstance(tp, str) else "")
                # cronograma BxL do paciente
                seq, cir, _ = L.build_sequencia(list(df.columns))
                for rotulo, col, off in seq:
                    val = srow[col]
                    seq_info.append({
                        "rotulo": rotulo,
                        "col": col,
                        "value": "" if pd.isna(val) else L.cell_display(val),
                        "cls": L.cell_class(val),
                    })
            else:
                selected_idx = None
        except (TypeError, ValueError):
            selected_idx = None

    bxl_cols = sorted([c for c in df.columns if L.is_bxl_col(c)], key=L.bxl_num)

    ctx.update({
        "columns": display_cols,
        "rows": rows,
        "total_rows": len(df),
        "id_col": id_col,
        "selected_idx": selected_idx,
        "selected_patient": selected_patient,
        "selected_fields": selected_fields,
        "seq_info": seq_info,
        "bxl_cols": bxl_cols,
        "tcle_url": tcle_url,
        "tcle_exists": tcle_exists,
        "motivos_exclusao": L.MOTIVOS_EXCLUSAO,
        "today_iso": date.today().isoformat(),
        "patient_list": list(enumerate(df.iloc[:, 0].astype(str).tolist())),
    })
    return render(request, "internal/biopsia.html", ctx)


# ═════════════════════════════════════════════════════════════════════
# DADOS CLÍNICOS
# ═════════════════════════════════════════════════════════════════════
@login_required
def clinicos(request):
    path = _clinicos_path()
    ctx = {"user": request.user, "file_exists": path.exists()}

    if not HAS_PANDAS:
        ctx["error"] = "Pandas não está instalado."
        return render(request, "internal/clinicos.html", ctx)
    if not path.exists():
        ctx["error"] = f"Arquivo não encontrado: {path.name}"
        return render(request, "internal/clinicos.html", ctx)

    try:
        df = _load_df(path)
    except Exception as exc:
        ctx["error"] = f"Erro ao ler o arquivo: {exc}"
        return render(request, "internal/clinicos.html", ctx)

    # coluna identificadora: 'CBL' se existir, senão a primeira
    id_col = "CBL" if "CBL" in df.columns else df.columns[0]

    CANONICOS = ["CBL", "TCLE", "prontuário", "Nome", "Data de Nasc"]
    campos_principais = [c for c in CANONICOS if c in df.columns and c != "TCLE"]
    if not campos_principais:
        campos_principais = [c for c in df.columns[:5] if c != "TCLE"]

    ocultar = set(campos_principais) | {"TCLE", "TCLE FOTO", "TCLE DATA ENVIO", "data TCLE"}
    outros_campos = [c for c in df.columns if c not in ocultar]

    if request.method == "POST":
        action = request.POST.get("action")
        idx = request.POST.get("idx")
        try:
            idx = int(idx) if idx is not None else None
        except (TypeError, ValueError):
            idx = None

        if action == "save_patient" and idx is not None and idx in df.index:
            for campo in campos_principais:
                key = f"campo_{campo}"
                if key in request.POST:
                    df.at[idx, campo] = request.POST[key]
            for campo in request.POST.getlist("outros"):
                key = f"extra_{campo}"
                if campo in df.columns and key in request.POST:
                    df.at[idx, campo] = request.POST[key]

            tcle_file = request.FILES.get("tcle_foto")
            if tcle_file:
                if "TCLE FOTO" not in df.columns:
                    df["TCLE FOTO"] = ""
                filepath = _save_tcle_file(tcle_file, str(df.at[idx, id_col]))
                df.at[idx, "TCLE FOTO"] = filepath
                if "TCLE" in df.columns:
                    df.at[idx, "TCLE"] = datetime.now().strftime("%d/%m/%Y %H:%M")
            try:
                _save_df(df, path)
                messages.success(request, "Dados do paciente atualizados!")
            except Exception as exc:
                messages.error(request, f"Erro ao salvar: {exc}")
            return redirect(f"{request.path}?sel={idx}#edit-section")

        elif action == "add_patient":
            nova = {}
            for campo in campos_principais:
                nova[campo] = request.POST.get(f"new_{campo}", "")
            for campo in request.POST.getlist("novo_outros"):
                if campo in df.columns:
                    nova[campo] = request.POST.get(f"new_extra_{campo}", "")

            tcle_file = request.FILES.get("tcle_foto")
            if not tcle_file:
                messages.error(request, "É obrigatório enviar o TCLE para cadastrar.")
                return redirect(f"{request.path}#novo")
            if "TCLE FOTO" not in df.columns:
                df["TCLE FOTO"] = ""
            if "TCLE DATA ENVIO" not in df.columns:
                df["TCLE DATA ENVIO"] = ""
            filepath = _save_tcle_file(tcle_file, nova.get(id_col, ""))
            nova["TCLE FOTO"] = filepath
            nova["TCLE DATA ENVIO"] = datetime.now().strftime("%d/%m/%Y %H:%M")

            linha = {c: nova.get(c, "") for c in df.columns}
            df.loc[len(df)] = linha
            try:
                _save_df(df, path)
                messages.success(request, "Novo paciente cadastrado!")
            except Exception as exc:
                messages.error(request, f"Erro ao salvar: {exc}")
            return redirect("internal:clinicos")

    # ─────────── GET ───────────
    sel_idx = request.GET.get("sel")
    selected_idx = None
    selected_main = []      # [(campo, value)]
    selected_extras = []    # [(campo, value)]
    selected_patient = None
    tcle_url = None
    tcle_exists = False
    if sel_idx is not None:
        try:
            selected_idx = int(sel_idx)
            if selected_idx in df.index:
                srow = df.loc[selected_idx]
                selected_patient = srow[id_col]
                for campo in campos_principais:
                    v = srow.get(campo, "")
                    selected_main.append((campo, "" if pd.isna(v) else str(v)))
                for campo in outros_campos:
                    v = srow.get(campo, "")
                    selected_extras.append((campo, "" if pd.isna(v) else str(v)))
                tp = srow.get("TCLE FOTO", "") if "TCLE FOTO" in df.columns else ""
                tcle_exists = _tcle_valido(tp if isinstance(tp, str) else "")
                tcle_url = _tcle_url(tp if isinstance(tp, str) else "")
            else:
                selected_idx = None
        except (TypeError, ValueError):
            selected_idx = None

    ctx.update({
        "columns": list(df.columns),
        "campos_principais": campos_principais,
        "outros_campos": outros_campos,
        "id_col": id_col,
        "total_rows": len(df),
        "selected_idx": selected_idx,
        "selected_patient": selected_patient,
        "selected_main": selected_main,
        "selected_extras": selected_extras,
        "tcle_url": tcle_url,
        "tcle_exists": tcle_exists,
        "patient_list": list(enumerate(df[id_col].astype(str).tolist())),
        "df_head": df.head(100).to_html(
            classes="table table-sm table-hover", border=0, index=False
        ) if not df.empty else "",
    })
    return render(request, "internal/clinicos.html", ctx)
