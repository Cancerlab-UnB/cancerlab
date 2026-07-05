# apps/internal/dashboard.py
"""
Reproduz os painéis analíticos que no app Streamlit ficavam na página `index`:

  1) Painel Clínico (CCR BxL)  — 3 abas: Visão geral / Distribuições / Coorte
  2) Painel Agenda BxL          — eventos, segmentos, pendências e WhatsApp

A lógica de limpeza/normalização foi portada 1:1 do arquivo Streamlit
(`Cancerlabteste__5_.py`). Os gráficos Plotly são renderizados como HTML
embutido (fig.to_html) e injetados no template.
"""
import re
import unicodedata
import warnings
import zipfile
from datetime import datetime, date
from typing import List, Optional, Tuple
from urllib.parse import quote

warnings.filterwarnings("ignore", message="Could not infer format")

try:
    import numpy as np
    import pandas as pd
    import plotly.express as px
    import plotly.graph_objects as go
    HAS_LIBS = True
except Exception:  # pragma: no cover
    HAS_LIBS = False


# =====================================================================
# THEME
# =====================================================================
def _apply_plotly_theme(fig, height: int = 420):
    fig.update_layout(
        template="plotly_white",
        height=height,
        margin=dict(l=12, r=12, t=45, b=12),
        font=dict(family="Inter, Arial, sans-serif", size=13, color="#111827"),
        hovermode="closest",
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    xanchor="left", x=0.0, bgcolor="rgba(255,255,255,0.85)"),
        paper_bgcolor="white", plot_bgcolor="white",
    )
    fig.update_xaxes(showgrid=True, gridcolor="rgba(17,24,39,0.08)", zeroline=False, ticks="outside")
    fig.update_yaxes(showgrid=True, gridcolor="rgba(17,24,39,0.08)", zeroline=False, ticks="outside")
    return fig


def _fig_html(fig) -> str:
    return fig.to_html(full_html=False, include_plotlyjs=False, config={"displayModeBar": False})


# =====================================================================
# HELPERS (portados do Streamlit)
# =====================================================================
def _slug(s: str) -> str:
    if s is None:
        return ""
    s = re.sub(r"\s+", " ", str(s)).strip()
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")


def _norm(s: str) -> str:
    return _slug(s).lower()


def _clean_columns(df):
    df = df.copy()
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    drop_me = [c for c in df.columns if str(c).lower().startswith("unnamed") and df[c].isna().all()]
    if drop_me:
        df = df.drop(columns=drop_me)
    return df


def _to_sim_nao(x) -> str:
    if pd.isna(x):
        return "Não"
    s = str(x).strip().lower()
    if s in ("", "nan", "na", "n/a", "não", "nao", "n", "no", "0", "-", "--"):
        return "Não"
    if s in ("sim", "s", "yes", "y", "1", "true", "t"):
        return "Sim"
    return "Não"


def _to_etilismo(x) -> str:
    if pd.isna(x):
        return "Não"
    s = str(x).strip().lower()
    if s in ("", "nan", "na", "n/a", "não", "nao", "n", "no", "0", "-", "--"):
        return "Não"
    if "social" in s:
        return "Social"
    if s in ("sim", "s", "yes", "y", "1", "true", "t"):
        return "Sim"
    return "Não"


def _is_date_like_series(s) -> bool:
    if pd.api.types.is_datetime64_any_dtype(s):
        return True
    name = _norm(getattr(s, "name", ""))
    if "data" in name:
        return True
    if s.dropna().empty:
        return False
    sample = s.dropna().astype(str).head(60)
    dt = pd.to_datetime(sample, errors="coerce", dayfirst=True)
    return dt.notna().mean() >= 0.60


def _format_date_only(x) -> str:
    if pd.isna(x):
        return "Não"
    if isinstance(x, (datetime, date)):
        d = x.date() if isinstance(x, datetime) else x
        if d.year < 1900 or d.year > 2200:
            return "Não"
        return d.strftime("%d/%m/%Y")
    s = str(x).strip()
    if not s or s.lower() in ("nan", "na", "n/a"):
        return "Não"
    try:
        ts = pd.to_datetime(s, errors="coerce", dayfirst=True)
        if pd.notna(ts):
            if ts.year < 1900 or ts.year > 2200:
                return "Não"
            return ts.strftime("%d/%m/%Y")
    except Exception:
        pass
    m = re.search(r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b", s)
    if m:
        try:
            ts2 = pd.to_datetime(m.group(1), errors="coerce", dayfirst=True)
            if pd.notna(ts2):
                if ts2.year < 1900 or ts2.year > 2200:
                    return "Não"
                return ts2.strftime("%d/%m/%Y")
        except Exception:
            pass
    return s


_ROMAN_RE = re.compile(r"\b(IV|III|II|I|0)\s*([ABC])?\b", flags=re.IGNORECASE)


def _split_multi_values(text: str) -> List[str]:
    text = str(text).replace("→", "->")
    parts = re.split(r"(?:->|>|;|\||/|,|\n|\r|\t)+", text)
    return [p.strip() for p in parts if p and p.strip()]


def _normalize_stage_token(token: str) -> Optional[str]:
    if token is None or (isinstance(token, float) and np.isnan(token)):
        return None
    t = str(token).strip()
    if not t or t.lower() in ("na", "nan", "-"):
        return None
    m = _ROMAN_RE.search(t)
    if not m:
        mm = re.findall(r"(IV|III|II|I|0)([ABC])?", t.upper())
        if not mm:
            return None
        r, l = mm[-1]
        return f"{r}{l or ''}".upper()
    roman = (m.group(1) or "").upper()
    letter = (m.group(2) or "").upper()
    return f"{roman}{letter}".upper()


def _clean_stage_field(x) -> str:
    if pd.isna(x):
        return "Não"
    raw = str(x).strip()
    if not raw or raw.lower() in ("na", "nan", "-"):
        return "Não"
    parts = _split_multi_values(raw) or [raw]
    tokens = []
    for p in parts:
        tok = _normalize_stage_token(p)
        if tok:
            tokens.append(tok)
        else:
            mm = re.findall(r"(IV|III|II|I|0)([ABC])?", str(p).upper())
            for r, l in mm:
                tokens.append(f"{r}{l or ''}".upper())
    uniq = []
    for t in tokens:
        if t not in uniq:
            uniq.append(t)
    if not uniq:
        return "Não"
    out = uniq[-1]
    if len(uniq) > 1:
        out = f"{out}*"
    return out


def _stage_sort_key(stage: str) -> Tuple[int, int, int, str]:
    if stage is None:
        return (99, 99, 1, "")
    s = str(stage).strip().upper()
    if s in ("NÃO", "NAO", "NA", "NAN", "", "—", "-"):
        return (99, 99, 1, "NAO")
    starred = 1 if s.endswith("*") else 0
    s0 = s[:-1] if starred else s
    m = re.match(r"^(0|IV|III|II|I)([ABC])?$", s0)
    if not m:
        return (98, 98, starred, s0)
    roman, letter = m.group(1), m.group(2) or ""
    roman_map = {"0": 0, "I": 1, "II": 2, "III": 3, "IV": 4}
    letter_map = {"": 0, "A": 1, "B": 2, "C": 3}
    return (roman_map.get(roman, 98), letter_map.get(letter, 98), starred, s0)


_DATE_DMY = re.compile(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b")
_DATE_YMD = re.compile(r"\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b")
_TIME = re.compile(r"\b\d{1,2}:\d{2}(:\d{2})?\b")


def _extract_numeric_sequence(x) -> List[float]:
    if pd.isna(x):
        return []
    if isinstance(x, (int, float, np.integer, np.floating)):
        if np.isfinite(x):
            return [float(x)]
        return []
    s = str(x).strip()
    if not s or s.lower() in ("nan", "na", "-"):
        return []
    s = _DATE_DMY.sub(" ", s)
    s = _DATE_YMD.sub(" ", s)
    s = _TIME.sub(" ", s)
    s = s.replace("→", "->").replace(">", " -> ").replace("<", " -> ").replace("->", " -> ")
    s = s.replace("%", " ").replace(",", ".")
    s = re.sub(r"[^0-9\.\-\+\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    nums = re.findall(r"[-+]?\d+(?:\.\d+)?", s)
    out = []
    for n in nums:
        try:
            v = float(n)
            if np.isfinite(v):
                out.append(v)
        except Exception:
            pass
    return out


def _robust_center(vals: List[float]) -> float:
    if not vals:
        return np.nan
    v = np.array(sorted(vals), dtype=float)
    if len(v) == 1:
        return float(v[0])
    if len(v) == 2:
        return float((v[0] + v[1]) / 2)
    if len(v) <= 4:
        return float(np.median(v))
    lo = np.quantile(v, 0.10)
    hi = np.quantile(v, 0.90)
    vw = np.clip(v, lo, hi)
    return float(np.mean(vw))


def _to_numeric_robust_series(s):
    seqs = s.map(_extract_numeric_sequence)
    return seqs.map(_robust_center).astype(float)


SIMNAO_COLS = [
    "HF 1° grau", "HF 2° grau", "HAS", "Diabetes ", "Tabagismo",
    "Sintoma", "sangramento", "alt de hab int", "perda de peso", "dor abd", "incontinencia",
    "mucinoso", "angiolinf", "perineural", "perfuração", "margem",
    "budding", "linfonodo lateral", "recrescimento WW",
    "neoadjuvancia", "Oxaliplatina", "Adjuvância",
]

NUM_ROBUST = [
    "Idade na entrada", "Peso (kg)", "Altura (m)", "IMC à Cx", "IMC ult 3 meses",
    "TISD (meses)", "CEA pré", "CEA pós", "cfDNA BxL1", "cfDNA BxLC1 ", "cfDNA BxL2",
    "Sum MAF% BXL1", "Sum MAF% BXL2",
]


# =====================================================================
# PAINEL CLÍNICO (CCR BxL)
# =====================================================================
def build_painel_clinico(df_in, filters: dict) -> dict:
    """Retorna dict de contexto para o template do painel clínico."""
    ctx = {"ok": False}
    if df_in is None or df_in.empty:
        ctx["warning"] = "Dados Clínicos não carregados."
        return ctx

    df = _clean_columns(df_in.copy())

    if " " in df.columns and "CBL" not in df.columns:
        df = df.rename(columns={" ": "CBL"})
    if "CBL" not in df.columns:
        df["CBL"] = np.arange(len(df)).astype(str)

    def _clean_cbl(x):
        if pd.isna(x):
            return ""
        s = str(x).strip()
        m = re.search(r"\bCBL\s*\d+\b", s, flags=re.IGNORECASE)
        return m.group(0).replace(" ", "").upper() if m else s

    df["CBL"] = df["CBL"].map(_clean_cbl)

    for c in list(df.columns):
        if "pront" in _norm(c):
            df = df.drop(columns=[c])

    date_cols = [c for c in df.columns if _is_date_like_series(df[c])]
    for c in date_cols:
        try:
            if pd.api.types.is_numeric_dtype(df[c]):
                dt = pd.to_datetime(df[c], unit="D", origin="1899-12-30", errors="coerce")
            else:
                dt = pd.to_datetime(df[c], errors="coerce", dayfirst=True)
            dt = dt.dt.normalize()
            dt = dt.where((dt.dt.year >= 1900) & (dt.dt.year <= 2200), pd.NaT)
            df[c] = dt
        except Exception:
            df[c] = pd.NaT

    stage_path_col = "Estadio patologico" if "Estadio patologico" in df.columns else None
    stage_clin_col = "estágio clínico/radiologico" if "estágio clínico/radiologico" in df.columns else None

    df["_estadio_pat_clean"] = df[stage_path_col].map(_clean_stage_field) if stage_path_col else "Não"
    df["_estadio_clin_clean"] = df[stage_clin_col].map(_clean_stage_field) if stage_clin_col else "Não"

    df["Estadiamento"] = df["_estadio_pat_clean"].copy()
    mask_fallback = df["Estadiamento"].isin(["Não", "", "NA"]) | df["Estadiamento"].isna()
    df.loc[mask_fallback, "Estadiamento"] = df.loc[mask_fallback, "_estadio_clin_clean"]
    df["Estadiamento"] = df["Estadiamento"].fillna("Não")

    for c in SIMNAO_COLS:
        if c in df.columns:
            df[c] = df[c].map(_to_sim_nao)
    if "Etilismo" in df.columns:
        df["Etilismo"] = df["Etilismo"].map(_to_etilismo)

    for c in NUM_ROBUST:
        if c in df.columns:
            if pd.api.types.is_numeric_dtype(df[c]):
                df[f"{c} (robust)"] = pd.to_numeric(df[c], errors="coerce")
            else:
                df[f"{c} (robust)"] = _to_numeric_robust_series(df[c])

    # -------- Filtros --------
    est_opts = sorted(df["Estadiamento"].dropna().astype(str).unique().tolist(), key=_stage_sort_key)
    sel_est = filters.get("estadiamento") or []
    dff = df.copy()
    if sel_est:
        dff = dff[dff["Estadiamento"].astype(str).isin(sel_est)]

    sex_opts = []
    if "Sexo" in dff.columns:
        sex_opts = sorted(df["Sexo"].dropna().astype(str).unique().tolist())
        sel_sex = filters.get("sexo") or []
        if sel_sex:
            dff = dff[dff["Sexo"].astype(str).isin(sel_sex)]

    asa_opts = []
    if "ASA" in dff.columns:
        asa_opts = sorted(df["ASA"].dropna().astype(str).unique().tolist())
        sel_asa = filters.get("asa") or []
        if sel_asa:
            dff = dff[dff["ASA"].astype(str).isin(sel_asa)]

    n_patients = int(dff["CBL"].nunique())
    cbls = sorted(dff["CBL"].dropna().astype(str).unique().tolist())

    # -------- Métricas --------
    def _med(series, fmt):
        if series is not None and pd.Series(series).notna().any():
            return fmt.format(np.nanmedian(series))
        return "—"

    age = dff["Idade na entrada (robust)"] if "Idade na entrada (robust)" in dff.columns else None
    imc = None
    if "IMC ult 3 meses (robust)" in dff.columns and dff["IMC ult 3 meses (robust)"].notna().any():
        imc = dff["IMC ult 3 meses (robust)"]
    elif "IMC à Cx (robust)" in dff.columns and dff["IMC à Cx (robust)"].notna().any():
        imc = dff["IMC à Cx (robust)"]
    cea_pre = dff["CEA pré (robust)"] if "CEA pré (robust)" in dff.columns else None

    metrics = {
        "n_patients": f"{n_patients:,}".replace(",", "."),
        "idade_mediana": _med(age, "{:.1f}"),
        "imc_mediano": _med(imc, "{:.1f}"),
        "cea_pre_mediana": _med(cea_pre, "{:.2f}"),
    }

    # -------- Visão geral: gráficos --------
    est_vc = dff["Estadiamento"].fillna("Não").astype(str).value_counts().reset_index()
    est_vc.columns = ["Estadiamento", "Pacientes"]
    est_vc["__ord"] = est_vc["Estadiamento"].map(_stage_sort_key)
    est_vc = est_vc.sort_values("__ord").drop(columns="__ord")
    fig = px.bar(est_vc, x="Estadiamento", y="Pacientes", text="Pacientes")
    fig.update_traces(textposition="outside", cliponaxis=False)
    fig.update_layout(xaxis_title="Estadiamento", yaxis_title="Pacientes")
    chart_estadiamento = _fig_html(_apply_plotly_theme(fig, height=460))

    pat = dff["_estadio_pat_clean"].fillna("Não").astype(str)
    pat_vc = pat.value_counts().reset_index()
    pat_vc.columns = ["Estádio patológico", "Pacientes"]
    pat_vc["__ord"] = pat_vc["Estádio patológico"].map(_stage_sort_key)
    pat_vc = pat_vc.sort_values("__ord").drop(columns="__ord")
    fig2 = px.bar(pat_vc, x="Estádio patológico", y="Pacientes", text="Pacientes")
    fig2.update_traces(textposition="outside", cliponaxis=False)
    chart_estadio_pat = _fig_html(_apply_plotly_theme(fig2, height=460))

    # -------- Distribuições --------
    candidates = []
    for c in dff.columns:
        if c in ("CBL",) or c.startswith("_") or c in date_cols:
            continue
        if dff[c].dropna().empty:
            continue
        candidates.append(c)
    priority = ["Estadiamento", "_estadio_pat_clean", "_estadio_clin_clean", "Sexo", "ASA"]
    ordered = [p for p in priority if p in candidates]
    for c in candidates:
        if c not in ordered:
            ordered.append(c)

    dist_var = filters.get("dist_var") or (ordered[0] if ordered else None)
    chart_dist = None
    if dist_var and dist_var in dff.columns:
        s = dff[dist_var]
        if pd.api.types.is_numeric_dtype(s):
            tmp = pd.DataFrame({dist_var: s}).dropna()
            if not tmp.empty:
                figd = px.histogram(tmp, x=dist_var, nbins=24)
                figd.update_layout(xaxis_title=dist_var, yaxis_title="Pacientes")
                chart_dist = _fig_html(_apply_plotly_theme(figd, height=520))
        else:
            tmp = s.where(s.notna(), "Não").astype(str).replace({"nan": "Não", "na": "Não", "": "Não"})
            vc = tmp.value_counts().head(40).reset_index()
            vc.columns = [dist_var, "Pacientes"]
            figd = px.bar(vc, x=dist_var, y="Pacientes", text="Pacientes")
            figd.update_traces(textposition="outside", cliponaxis=False)
            figd.update_layout(xaxis_title=dist_var, yaxis_title="Pacientes")
            chart_dist = _fig_html(_apply_plotly_theme(figd, height=520))

    # -------- Coorte --------
    BASE = [
        "CBL", "Sexo", "Idade na entrada", "Idade na entrada (robust)",
        "HF 1° grau", "HF 2° grau", "HAS", "Diabetes ", "Tabagismo", "Etilismo",
        "localização", "Sincronico / metacronico", "TNM", "Estadiamento",
        "_estadio_pat_clean", "_estadio_clin_clean", "grau de dif", "mucinoso",
        "angiolinf", "perineural", "perfuração", "margem", "budding",
        "CEA pré", "CEA pré (robust)", "CEA pós", "CEA pós (robust)",
        "cfDNA BxL1", "cfDNA BxL1 (robust)", "cfDNA BxLC1 ", "cfDNA BxLC1  (robust)",
        "cfDNA BxL2", "cfDNA BxL2 (robust)", "Sum MAF% BXL1", "Sum MAF% BXL1 (robust)",
        "Sum MAF% BXL2", "Sum MAF% BXL2 (robust)", "neoadjuvancia",
        "Esquema RT (0 short, 1 long)", "Esquema QT", "Oxaliplatina", "Adjuvância",
    ]
    cols = [c for c in BASE if c in dff.columns]
    include_dates = bool(filters.get("include_dates"))
    if include_dates:
        cols += [c for c in date_cols if c in dff.columns]

    coorte = dff[cols].copy()
    for c in coorte.columns:
        if c in date_cols:
            coorte[c] = coorte[c].map(_format_date_only)
            continue
        if c in ("Estadiamento", "_estadio_pat_clean", "_estadio_clin_clean"):
            coorte[c] = coorte[c].fillna("Não").astype(str)
            continue
        if c in SIMNAO_COLS:
            coorte[c] = coorte[c].map(_to_sim_nao)
            continue
        if c == "Etilismo":
            coorte[c] = coorte[c].map(_to_etilismo)
            continue
        if not pd.api.types.is_numeric_dtype(coorte[c]):
            coorte[c] = coorte[c].where(coorte[c].notna(), "Não").astype(str)
            coorte[c] = coorte[c].replace({"nan": "Não", "na": "Não", "": "Não"})

    coorte = coorte.rename(columns={
        "_estadio_pat_clean": "Estádio patológico (limpo)",
        "_estadio_clin_clean": "Estádio clínico/radiológico (limpo)",
    })

    ctx.update({
        "ok": True,
        "est_opts": est_opts, "sel_est": sel_est,
        "sex_opts": sex_opts, "sel_sex": filters.get("sexo") or [],
        "asa_opts": asa_opts, "sel_asa": filters.get("asa") or [],
        "include_dates": include_dates,
        "n_patients": n_patients, "cbls": cbls,
        "metrics": metrics,
        "chart_estadiamento": chart_estadiamento,
        "chart_estadio_pat": chart_estadio_pat,
        "dist_options": ordered, "dist_var": dist_var, "chart_dist": chart_dist,
        "coorte_columns": list(coorte.columns),
        "coorte_rows": coorte.head(400).values.tolist(),
        "coorte_csv": coorte.to_csv(index=False),
    })
    return ctx


# =====================================================================
# PAINEL AGENDA BxL  (eventos / segmentos / pendências / whatsapp)
# =====================================================================
BLOCKED_TEXTS_EXACT = {
    "", "-", "--", "na", "n/a", "nan", "nc", "sp", "f",
    "sem cirurgia", "sem tumor", "none",
}
BLOCKED_TEXTS_CONTAINS = ["nao coletado", "não coletado", "cancelado", "sem cirurgia", "sem tumor"]

_RE_CBL_STRICT = re.compile(r"^CBL\d+$", flags=re.IGNORECASE)
_RE_CBL_IN_TEXT = re.compile(r"\bCBL\s*(\d+)\b", flags=re.IGNORECASE)
RE_YMD = re.compile(r"(?<!\d)(\d{4})[\/\-.](\d{1,2})[\/\-.](\d{1,2})(?!\d)")
RE_DMY = re.compile(r"(?<!\d)(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})(?!\d)")
RE_DM = re.compile(r"(?<!\d)(\d{1,2})[\/\-.](\d{1,2})(?!\d)")


def _classify_segmento(raw):
    s = _norm(raw)
    if not s or s in {"nan", "none", "-", "--"}:
        return "INDEFINIDO"
    colon_markers = ["colon", "colico", "colica", "ceco", "cecal", "sigmoide",
                     "ascendente", "descendente", "transverso", "angulo esplenico", "angulo hepatico"]
    reto_markers = ["reto", "retal", "retossigmoide", "reto-sigmoide", "recto", "rectal"]
    if any(m in s for m in colon_markers):
        return "CÓLON"
    if any(m in s for m in reto_markers):
        return "RETO"
    return "INDEFINIDO"


def _fmt_date(d):
    return "—" if d is None else d.strftime("%d/%m/%y")


def _safe_int(x, default=0):
    try:
        if pd.isna(x):
            return default
        return int(float(x))
    except Exception:
        return default


def _valid_year(d):
    return isinstance(d, date) and 1900 <= d.year <= 2200


def _normalize_cbl(raw):
    s = str(raw or "").strip()
    m = _RE_CBL_IN_TEXT.search(s)
    if m:
        return f"CBL{int(m.group(1))}"
    return s.replace(" ", "")


def _is_active_study_id(raw):
    s = str(raw or "").strip().replace(" ", "")
    return bool(_RE_CBL_STRICT.match(s))


def _classify_patient_from_id(raw_id):
    raw = str(raw_id or "")
    low = _norm(raw)
    if "+" in raw:
        return {"EventType": "FALECIMENTO", "IsActiveStudy": False, "IsDeath": True, "IsEvent": True}
    if "exc" in low or "saiu" in low:
        return {"EventType": "EVENTO NÃO ESPECIFICADO", "IsActiveStudy": False, "IsDeath": False, "IsEvent": True}
    return {"EventType": "", "IsActiveStudy": True, "IsDeath": False, "IsEvent": False}


def _find_main_columns(headers):
    cols = {"id": None, "telefone": [], "segmento": None, "obito": None}
    for idx, h in enumerate(headers, start=1):
        nh = _norm(h)
        if cols["id"] is None and (("id" in nh and "paciente" in nh) or ("cbl" in nh)):
            cols["id"] = idx
        if any(k in nh for k in ["telefone", "celular", "fone", "whats", "whatsapp", "contato"]):
            cols["telefone"].append(idx)
        if cols["segmento"] is None and "segment" in nh:
            cols["segmento"] = idx
        if cols["obito"] is None and any(k in nh for k in ["obito", "óbito", "falecimento"]):
            cols["obito"] = idx
    return cols


def _find_event_columns(headers):
    cols = []
    for idx, h in enumerate(headers, start=1):
        nh = _norm(h)
        if not nh or "obs" in nh:
            continue
        if "bxl" in nh or "cirurg" in nh or re.search(r"\brm\b", nh) or re.search(r"\btc\b", nh):
            cols.append(idx)
    return cols


def _coleta_short_label(col_name):
    cn = _norm(col_name)
    m = re.search(r"\bbxl\s*(\d+)\b", cn)
    if m:
        return f"BxL {int(m.group(1))}"
    if "pos" in cn and "rt" in cn:
        return "BxL pós-RT"
    if "cirurg" in cn:
        return "Cirurgia"
    if re.search(r"\brm\b", cn):
        return "RM"
    if re.search(r"\btc\b", cn):
        return "TC"
    return str(col_name).strip()


def _two_digit_year_to_yyyy(yy):
    return 2000 + yy if yy <= 69 else 1900 + yy


def _infer_year_for_ddmm(d, m, today_):
    y = today_.year
    try:
        cand = date(y, m, d)
    except Exception:
        return None
    if cand <= today_:
        if (today_ - cand).days <= 30:
            return cand
        try:
            return date(y + 1, m, d)
        except Exception:
            return None
    return cand


def _clean_date_text(s):
    s = str(s or "").strip()
    if not s:
        return ""
    s = s.replace("\\", "/").replace("–", "-").replace("—", "-")
    s = re.sub(r"[;,]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"/{2,}", "/", s)
    s = re.sub(r"-{2,}", "-", s)
    s = re.sub(r"\.{2,}", ".", s)
    s = re.sub(r"(\d)\s*/\s*/\s*(\d)", r"\1/\2", s)
    return s


def _extract_first_date_from_string(s, today_):
    if s is None:
        return None
    s0 = _clean_date_text(s)
    if not s0:
        return None
    low = _norm(s0)
    if low in BLOCKED_TEXTS_EXACT or any(x in low for x in BLOCKED_TEXTS_CONTAINS):
        return None
    m = RE_YMD.search(s0)
    if m:
        y, mo, d = m.groups()
        try:
            dd = date(int(y), int(mo), int(d))
            return dd if _valid_year(dd) else None
        except Exception:
            pass
    m = RE_DMY.search(s0)
    if m:
        d, mo, y = m.groups()
        try:
            yy = _two_digit_year_to_yyyy(int(y)) if len(y) == 2 else int(y)
            dd = date(yy, int(mo), int(d))
            return dd if _valid_year(dd) else None
        except Exception:
            pass
    m = RE_DM.search(s0)
    if m:
        d, mo = m.groups()
        try:
            return _infer_year_for_ddmm(int(d), int(mo), today_)
        except Exception:
            return None
    return None


def _extract_date(x, today_, epoch):
    from openpyxl.utils.datetime import from_excel
    if x is None:
        return None
    try:
        if pd.isna(x):
            return None
    except Exception:
        pass
    if isinstance(x, datetime):
        d = x.date()
        return d if _valid_year(d) else None
    if isinstance(x, date):
        return x if _valid_year(x) else None
    if isinstance(x, (int, float, np.integer, np.floating)):
        try:
            v = float(x)
            if 1 <= v <= 60000:
                dtv = from_excel(v, epoch=epoch)
                if isinstance(dtv, datetime):
                    d = dtv.date()
                    return d if _valid_year(d) else None
        except Exception:
            pass
    return _extract_first_date_from_string(str(x), today_)


def _has_collected_tag(raw):
    s = "" if raw is None else str(raw)
    return bool(re.search(r"\(\s*coletado\s*\)", s, flags=re.IGNORECASE))


def _should_ignore_cell(raw):
    low = _norm("" if raw is None else str(raw).strip())
    return low in BLOCKED_TEXTS_EXACT or any(x in low for x in BLOCKED_TEXTS_CONTAINS)


def _load_events_and_patients(file_path, sheet_name=None):
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name is None or sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    epoch = wb.epoch
    today_ = date.today()

    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append("" if v is None else str(v).strip())

    cols = _find_main_columns(headers)
    cbl_col = cols["id"] or 1
    phone_cols = cols["telefone"]
    segment_col = cols["segmento"]
    obito_col = cols["obito"]
    event_cols = _find_event_columns(headers)

    patient_rows, phones_map, events = [], {}, []
    for r in range(2, ws.max_row + 1):
        cbl_val = ws.cell(row=r, column=cbl_col).value
        if cbl_val is None or str(cbl_val).strip() == "":
            continue
        cbl_raw = str(cbl_val).strip()
        cbl = _normalize_cbl(cbl_raw)
        pclass = _classify_patient_from_id(cbl_raw)

        segmento = ""
        if segment_col:
            seg_val = ws.cell(row=r, column=segment_col).value
            segmento = "" if seg_val is None else str(seg_val).strip()

        phone_raw_by_col = {}
        for pc in phone_cols:
            hname = headers[pc - 1] or f"Telefone_{pc}"
            phone_raw_by_col[hname] = ws.cell(row=r, column=pc).value
        phones_map[cbl] = phone_raw_by_col

        obito_date = None
        if obito_col:
            obito_date = _extract_date(ws.cell(row=r, column=obito_col).value, today_, epoch)

        valid_dates = []
        for c in event_cols:
            d_any = _extract_date(ws.cell(row=r, column=c).value, today_, epoch)
            if d_any:
                valid_dates.append(d_any)
        event_date = obito_date if obito_date else (max(valid_dates) if valid_dates else None)

        patient_rows.append({
            "CBL": cbl, "CBL_RAW": cbl_raw, "Segmento": segmento,
            "SegmentoGrupo": _classify_segmento(segmento),
            "EventType": pclass["EventType"],
            "EventDate_str": _fmt_date(event_date) if event_date else "—",
            "IsEvent": pclass["IsEvent"], "IsDeath": pclass["IsDeath"],
            "IsActiveStudy": pclass["IsActiveStudy"],
        })

        if not pclass["IsActiveStudy"]:
            continue

        for bc in event_cols:
            head = headers[bc - 1] or f"Col_{bc}"
            cell = ws.cell(row=r, column=bc)
            raw = cell.value
            raw_str = "" if raw is None else str(raw).strip()
            if raw is None or raw_str == "" or _should_ignore_cell(raw_str) or _has_collected_tag(raw_str):
                continue
            dd = _extract_date(raw, today_, epoch)
            if not dd:
                continue
            events.append({
                "CBL": cbl, "CBL_RAW": cbl_raw, "IN_STUDY": _is_active_study_id(cbl_raw),
                "Segmento": segmento, "SegmentoGrupo": _classify_segmento(segmento),
                "Coleta_short": _coleta_short_label(str(head)),
                "Data": dd, "Data_str": _fmt_date(dd), "Dias": (dd - today_).days,
                "Sheet": sheet_name, "Cell": cell.coordinate,
            })

    df_events = pd.DataFrame(events)
    df_pat = pd.DataFrame(patient_rows).drop_duplicates(subset=["CBL", "CBL_RAW"])
    return df_events, df_pat, phones_map, headers, phone_cols, wb.sheetnames, sheet_name


def build_painel_agenda(file_path, filters: dict) -> dict:
    ctx = {"ok": False}
    if not zipfile.is_zipfile(file_path):
        ctx["error"] = "O arquivo não está em formato .xlsx válido."
        return ctx
    try:
        df_all, df_pat, phones_map, headers, phone_cols, sheetnames, sheet = \
            _load_events_and_patients(file_path, filters.get("sheet"))
    except Exception as e:
        ctx["error"] = f"Não consegui abrir o Excel: {e}"
        return ctx

    # ----- Eventos -----
    eventos_ctx = {"has": False}
    if df_pat is not None and not df_pat.empty:
        dfu = df_pat.drop_duplicates(subset=["CBL"]).copy()
        eventos_ctx = {
            "has": True,
            "total": int(dfu["CBL"].nunique()),
            "ativos": int(dfu[dfu["IsActiveStudy"] == True]["CBL"].nunique()),
            "eventos": int(dfu[dfu["IsEvent"] == True]["CBL"].nunique()),
            "falec": int(dfu[dfu["IsDeath"] == True]["CBL"].nunique()),
        }
        df_evt = dfu[dfu["IsEvent"] == True].copy()
        if not df_evt.empty:
            vc = (df_evt["EventType"].replace("", "EVENTO NÃO ESPECIFICADO")
                  .fillna("EVENTO NÃO ESPECIFICADO").value_counts().reset_index())
            vc.columns = ["Tipo de evento", "Qtd"]
            eventos_ctx["resumo"] = vc.values.tolist()
            det = df_evt[["CBL", "EventType", "EventDate_str", "SegmentoGrupo", "Segmento", "CBL_RAW"]]
            eventos_ctx["detalhes"] = det.values.tolist()

    # ----- Segmentos -----
    seg_ctx = {"has": False}
    chart_segmento = None
    if df_pat is not None and not df_pat.empty:
        df_ativos = df_pat[df_pat["IsActiveStudy"] == True].drop_duplicates(subset=["CBL"]).copy()
        ordem = ["CÓLON", "RETO", "INDEFINIDO"]
        cont = (df_ativos["SegmentoGrupo"].fillna("INDEFINIDO").replace("", "INDEFINIDO")
                .value_counts().reindex(ordem, fill_value=0).reset_index())
        cont.columns = ["Segmento", "Pacientes"]
        seg_ctx = {
            "has": True,
            "colon": int(cont.loc[cont["Segmento"] == "CÓLON", "Pacientes"].iloc[0]),
            "reto": int(cont.loc[cont["Segmento"] == "RETO", "Pacientes"].iloc[0]),
            "indef": int(cont.loc[cont["Segmento"] == "INDEFINIDO", "Pacientes"].iloc[0]),
        }
        graf = cont[cont["Segmento"].isin(["CÓLON", "RETO"])].copy()
        figs = px.bar(graf, x="Segmento", y="Pacientes", text="Pacientes",
                      title="Pacientes ativos por segmento", template="plotly_white")
        figs.update_traces(textposition="outside")
        figs.update_layout(height=340, margin=dict(l=20, r=20, t=60, b=20),
                           yaxis_title="Número de pacientes", xaxis_title="", showlegend=False)
        chart_segmento = _fig_html(figs)

    # ----- Filtros / pendências -----
    segmento_sel = filters.get("segmento") or "Todos"
    paciente_sel = filters.get("paciente") or "Todos"
    hide_events = filters.get("hide_events", True)
    cbl_search = (filters.get("cbl_search") or "").strip()

    pacientes_options = ["Todos"]
    if df_all is not None and not df_all.empty:
        pacientes_options += sorted(df_all["CBL"].dropna().astype(str).unique().tolist())
    elif df_pat is not None and not df_pat.empty:
        pacientes_options += sorted(df_pat["CBL"].dropna().astype(str).unique().tolist())

    pend_ctx = {"empty": True}
    tables = {}
    if df_all is not None and not df_all.empty:
        df = df_all.copy()
        if segmento_sel != "Todos":
            df = df[df["SegmentoGrupo"].astype(str) == segmento_sel].copy()
        if paciente_sel != "Todos":
            df = df[df["CBL"].astype(str) == str(paciente_sel)].copy()
        if hide_events and "IN_STUDY" in df.columns:
            df = df[df["IN_STUDY"] == True].copy()
        if cbl_search:
            df = df[df["CBL"].astype(str).str.contains(cbl_search, case=False, na=False)].copy()

        if not df.empty:
            df["Dias_num"] = pd.to_numeric(df["Dias"], errors="coerce")
            df["Status_display"] = df["Dias_num"].apply(lambda d: "ATRASADA" if pd.notna(d) and d < 0 else "PENDENTE")
            df["Coleta_show"] = df.apply(
                lambda r: (f"⏰ {r['Coleta_short']}" if r["Status_display"] == "ATRASADA" else f"📌 {r['Coleta_short']}"),
                axis=1)
            pend = df.sort_values(["Data", "CBL", "Coleta_short"]).reset_index(drop=True)
            atras = pend[pend["Dias_num"] < 0]
            fut = pend[pend["Dias_num"] >= 0]
            fut_0_7 = fut[(fut["Dias_num"] >= 0) & (fut["Dias_num"] <= 7)]
            fut_8_14 = fut[(fut["Dias_num"] >= 8) & (fut["Dias_num"] <= 14)]
            fut_15_30 = fut[(fut["Dias_num"] >= 15) & (fut["Dias_num"] <= 30)]
            fut_31 = fut[fut["Dias_num"] >= 31]

            SHOW = ["CBL", "SegmentoGrupo", "Segmento", "Coleta_show", "Data_str", "Dias_num", "Status_display", "Cell"]

            def _rows(d):
                return d[SHOW].values.tolist()

            tables = {
                "pend": _rows(pend), "atras": _rows(atras),
                "f07": _rows(fut_0_7), "f814": _rows(fut_8_14),
                "f1530": _rows(fut_15_30), "f31": _rows(fut_31),
            }
            pend_ctx = {
                "empty": False,
                "n_pend": int(len(pend)),
                "n_atras": int(len(atras)),
                "n_0_30": int(len(fut[(fut["Dias_num"] >= 0) & (fut["Dias_num"] <= 30)])),
                "n_31": int(len(fut_31)),
            }

    ctx.update({
        "ok": True,
        "sheetnames": sheetnames, "sheet": sheet,
        "segmento_options": ["Todos", "CÓLON", "RETO", "INDEFINIDO"],
        "segmento_sel": segmento_sel,
        "pacientes_options": pacientes_options, "paciente_sel": paciente_sel,
        "hide_events": hide_events, "cbl_search": cbl_search,
        "eventos": eventos_ctx, "seg": seg_ctx, "chart_segmento": chart_segmento,
        "pend": pend_ctx, "tables": tables,
    })
    return ctx
